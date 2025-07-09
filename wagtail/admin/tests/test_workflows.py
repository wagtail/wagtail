import io
import json
import logging
from unittest import expectedFailure, mock, skip

from django.conf import settings
from django.contrib.admin.utils import quote
from django.contrib.auth.models import Group, Permission
from django.contrib.contenttypes.models import ContentType
from django.core import mail
from django.core.mail import EmailMultiAlternatives
from django.test import TestCase, override_settings
from django.urls import reverse
from freezegun import freeze_time
from openpyxl import load_workbook

from wagtail.admin.admin_url_finder import AdminURLFinder
from wagtail.admin.mail import (
    BaseWorkflowStateEmailNotifier,
    WorkflowStateApprovalEmailNotifier,
    WorkflowStateRejectionEmailNotifier,
)
from wagtail.admin.staticfiles import versioned_static
from wagtail.admin.utils import (
    get_admin_base_url,
    get_latest_str,
    get_user_display_name,
)
from wagtail.models import (
    GroupApprovalTask,
    GroupPagePermission,
    Page,
    PageViewRestriction,
    Task,
    TaskState,
    Workflow,
    WorkflowContentType,
    WorkflowPage,
    WorkflowState,
    WorkflowTask,
)
from wagtail.signals import page_published, published
from wagtail.test.testapp.models import (
    FullFeaturedSnippet,
    ModeratedModel,
    MultiPreviewModesPage,
    SimplePage,
    SimpleTask,
    UserApprovalTask,
)
from wagtail.test.utils import WagtailTestUtils
from wagtail.test.utils.template_tests import AdminTemplateTestUtils
from wagtail.users.models import UserProfile


def delete_existing_workflows():
    WorkflowPage.objects.all().delete()
    Workflow.objects.all().delete()
    Task.objects.all().delete()
    WorkflowTask.objects.all().delete()


class TestWorkflowMenus(WagtailTestUtils, TestCase):
    def setUp(self):
        self.login()

        self.editor = self.create_user(
            username="editor",
            email="editor@email.com",
            password="password",
        )
        editors = Group.objects.get(name="Editors")
        editors.user_set.add(self.editor)

    def test_workflow_settings_and_reports_menus_are_shown_to_admin(self):
        response = self.client.get("/admin/")
        self.assertContains(response, '"url": "/admin/workflows/list/"')
        self.assertContains(response, '"url": "/admin/workflows/tasks/index/"')
        self.assertContains(response, '"url": "/admin/reports/workflow/"')
        self.assertContains(response, '"url": "/admin/reports/workflow_tasks/"')

    def test_workflow_settings_menus_are_not_shown_to_editor(self):
        self.login(user=self.editor)
        response = self.client.get("/admin/")
        self.assertNotContains(response, '"url": "/admin/workflows/list/"')
        self.assertNotContains(response, '"url": "/admin/workflows/tasks/index/"')
        self.assertContains(response, '"url": "/admin/reports/workflow/"')
        self.assertContains(response, '"url": "/admin/reports/workflow_tasks/"')

    @override_settings(WAGTAIL_WORKFLOW_ENABLED=False)
    def test_workflow_menus_are_hidden_when_workflows_are_disabled(self):
        response = self.client.get("/admin/")
        self.assertNotContains(response, '"url": "/admin/workflows/list/"')
        self.assertNotContains(response, '"url": "/admin/workflows/tasks/index/"')
        self.assertNotContains(response, '"url": "/admin/reports/workflow/"')
        self.assertNotContains(response, '"url": "/admin/reports/workflow_tasks/"')


class TestWorkflowsIndexView(AdminTemplateTestUtils, WagtailTestUtils, TestCase):
    def setUp(self):
        delete_existing_workflows()
        self.login()

        self.editor = self.create_user(
            username="editor",
            email="editor@email.com",
            password="password",
        )
        editors = Group.objects.get(name="Editors")
        editors.user_set.add(self.editor)

        self.moderator = self.create_user(
            username="moderator",
            email="moderator@email.com",
            password="password",
        )
        moderators = Group.objects.get(name="Moderators")
        moderators.user_set.add(self.moderator)
        moderators.permissions.add(Permission.objects.get(codename="add_workflow"))

    def create_workflows(self):
        home_page = Page.objects.get(depth=2)
        workflows = [
            Workflow.objects.create(name=f"test_workflow_{i}", active=True)
            for i in range(5)
        ]
        task = SimpleTask.objects.create(name="test_task")
        workflow_tasks = [
            WorkflowTask(workflow=workflow, task=task) for workflow in workflows
        ]
        WorkflowTask.objects.bulk_create(workflow_tasks)
        workflow_pages = [
            WorkflowPage(
                workflow=workflow,
                page=home_page.add_child(
                    instance=SimplePage(title="Simple", content="Very simple")
                ),
            )
            for workflow in workflows
        ]
        WorkflowPage.objects.bulk_create(workflow_pages)

    def get(self, params={}):
        return self.client.get(reverse("wagtailadmin_workflows:index"), params)

    def test_simple(self):
        response = self.get()
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "wagtailadmin/workflows/index.html")
        self.assertBreadcrumbsItemsRendered(
            [{"url": "", "label": "Workflows"}],
            response.content,
        )

        # Initially there should be no workflows listed
        self.assertContains(response, "There are no enabled workflows.")

        Workflow.objects.create(name="test_workflow", active=True)

        # Now the listing should contain our workflow
        response = self.get()
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "wagtailadmin/workflows/index.html")
        self.assertNotContains(response, "There are no enabled workflows.")
        self.assertContains(response, "test_workflow")

    def test_multiple_snippets_assigned_to_workflow(self):
        Workflow.objects.create(name="Nocontenttypes")
        multi_ct_workflow = Workflow.objects.create(name="Multicontenttypes")
        for model in [FullFeaturedSnippet, ModeratedModel]:
            WorkflowContentType.objects.create(
                workflow=multi_ct_workflow,
                content_type=ContentType.objects.get_for_model(model),
            )

        response = self.get()
        self.assertEqual(response.status_code, 200)
        soup = self.get_soup(response.content)
        cells = [
            text
            for td in soup.select("main table td")
            if (text := td.get_text(separator=" | ", strip=True))
        ]
        self.assertEqual(
            cells,
            [
                "Multicontenttypes",
                "0 pages | 2 snippet types",
                "Nocontenttypes",
                "0 pages | 0 snippet types",
            ],
        )

    def test_num_queries(self):
        self.create_workflows()
        self.get()
        with self.assertNumQueries(23):
            self.get()
        self.create_workflows()
        with self.assertNumQueries(33):
            self.get()

    def test_deactivated(self):
        Workflow.objects.create(name="test_workflow", active=False)

        # The listing should contain our workflow, as well as marking it as disabled
        response = self.get(params={"show_disabled": "true"})
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "No workflows have been created.")
        self.assertContains(response, "test_workflow")
        self.assertContains(
            response, '<span class="w-status">Disabled</span>', html=True
        )
        # Should display the "Show disabled" option as a filter
        soup = self.get_soup(response.content)
        active_filter = soup.select_one('[data-w-active-filter-id="id_show_disabled"]')
        self.assertIsNotNone(active_filter)
        self.assertEqual(
            active_filter.get_text(separator=" ", strip=True),
            "Show disabled: Yes",
        )
        show_disabled_yes = soup.select_one('input[name="show_disabled"][value="true"]')
        self.assertIsNotNone(show_disabled_yes)
        self.assertTrue(show_disabled_yes.has_attr("checked"))

        # If we set 'show_disabled' to 'False', the workflow should not be displayed
        response = self.get(params={})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "There are no enabled workflows.")
        # Should not display any active filters,
        # and the "Show disabled" option should be set to "No" by default
        soup = self.get_soup(response.content)
        active_filter = soup.select_one('[data-w-active-filter-id="id_show_disabled"]')
        self.assertIsNone(active_filter)
        show_disabled_no = soup.select_one('input[name="show_disabled"][value="false"]')
        self.assertIsNotNone(show_disabled_no)
        self.assertTrue(show_disabled_no.has_attr("checked"))

    def test_permissions(self):
        self.login(user=self.editor)
        response = self.get()
        self.assertEqual(response.status_code, 302)
        full_context = {
            key: value for context in response.context for key, value in context.items()
        }
        self.assertEqual(
            full_context["message"],
            "Sorry, you do not have permission to access this area.",
        )

        self.login(user=self.moderator)
        response = self.get()
        self.assertEqual(response.status_code, 200)

    def test_ordering(self):
        workflows = sorted(
            [
                # Mix up the creation order to ensure we're not ordering by PK
                Workflow.objects.create(name="workflow_1"),
                Workflow.objects.create(name="workflow_3"),
                Workflow.objects.create(name="workflow_2"),
            ],
            key=lambda workflow: workflow.name,
        )

        response = self.get()
        self.assertEqual(response.status_code, 200)
        self.assertSequenceEqual(response.context["object_list"], workflows)
        self.assertEqual(response.context["object_list"].query.order_by, ("name",))

        response = self.get(params={"ordering": "name"})
        self.assertEqual(response.status_code, 200)
        self.assertSequenceEqual(response.context["object_list"], workflows)
        self.assertEqual(response.context["object_list"].query.order_by, ("name",))

        response = self.get(params={"ordering": "-name"})
        self.assertEqual(response.status_code, 200)
        self.assertSequenceEqual(response.context["object_list"], workflows[::-1])
        self.assertEqual(response.context["object_list"].query.order_by, ("-name",))

    def test_search(self):
        Workflow.objects.create(name="foo workflow")
        Workflow.objects.create(name="bar workflow")
        Workflow.objects.create(name="bar world workflow")

        response = self.get(params={"q": "bAr"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "bar workflow")
        self.assertContains(response, "bar world workflow")
        self.assertNotContains(response, "foo workflow")

    def test_search_results(self):
        Workflow.objects.create(name="foo workflow")
        Workflow.objects.create(name="bar workflow")
        Workflow.objects.create(name="bar world workflow")

        response = self.client.get(
            reverse("wagtailadmin_workflows:index_results"),
            {"q": "AR"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertBreadcrumbsNotRendered(response.content)
        self.assertContains(response, "bar workflow")
        self.assertContains(response, "bar world workflow")
        self.assertNotContains(response, "foo workflow")

    def test_pagination(self):
        Workflow.objects.bulk_create(
            [Workflow(name=f"workflow_{i}") for i in range(1, 50)]
        )

        url = reverse("wagtailadmin_workflows:index")

        response = self.get({"p": 2})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context["object_list"]), 20)
        self.assertContains(response, url + "?p=1")
        self.assertContains(response, url + "?p=2")
        self.assertContains(response, url + "?p=3")

        response = self.get({"p": 4})
        # Check response
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "wagtailadmin/workflows/index.html")

        # Check that we got the last page
        self.assertEqual(
            response.context["page_obj"].number,
            response.context["paginator"].num_pages,
        )


class TestWorkflowPermissions(WagtailTestUtils, TestCase):
    url_name = "wagtailadmin_reports:workflow"

    def setUp(self):
        self.user = self.login()

    def get(self, params={}):
        return self.client.get(reverse(self.url_name), params)

    def test_simple(self):
        response = self.get()
        self.assertEqual(response.status_code, 200)

    def test_get_with_no_permission(self):
        group = Group.objects.create(name="test group")
        self.user.is_superuser = False
        self.user.save()
        self.user.groups.add(group)
        self.user.user_permissions.add(
            Permission.objects.get(
                content_type__app_label="wagtailadmin", codename="access_admin"
            )
        )
        # No GroupPagePermission created

        response = self.get()
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("wagtailadmin_home"))

    def test_get_with_minimal_permissions(self):
        group = Group.objects.create(name="test group")
        self.user.is_superuser = False
        self.user.save()
        self.user.groups.add(group)
        self.user.user_permissions.add(
            Permission.objects.get(
                content_type__app_label="wagtailadmin", codename="access_admin"
            )
        )
        GroupPagePermission.objects.create(
            group=group,
            page=Page.objects.first(),
            permission_type="change",
        )

        response = self.get()

        self.assertEqual(response.status_code, 200)


class TestWorkflowTaskPermissions(TestWorkflowPermissions):
    url_name = "wagtailadmin_reports:workflow_tasks"


class TestWorkflowsCreateView(AdminTemplateTestUtils, WagtailTestUtils, TestCase):
    def setUp(self):
        delete_existing_workflows()
        self.login()
        self.task_1 = SimpleTask.objects.create(name="first_task")
        self.task_2 = SimpleTask.objects.create(name="second_task")

        self.editor = self.create_user(
            username="editor",
            email="editor@email.com",
            password="password",
        )
        editors = Group.objects.get(name="Editors")
        editors.user_set.add(self.editor)

        self.moderator = self.create_user(
            username="moderator",
            email="moderator@email.com",
            password="password",
        )
        moderators = Group.objects.get(name="Moderators")
        moderators.user_set.add(self.moderator)
        moderators.permissions.add(Permission.objects.get(codename="add_workflow"))

        self.root_page = Page.objects.get(depth=1)
        self.snippet = FullFeaturedSnippet.objects.create(text="foo")
        self.snippet_content_type = ContentType.objects.get_for_model(
            FullFeaturedSnippet
        )

    def get(self, params={}):
        return self.client.get(reverse("wagtailadmin_workflows:add"), params)

    def post(self, post_data={}):
        return self.client.post(reverse("wagtailadmin_workflows:add"), post_data)

    def test_get(self):
        response = self.get()
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "wagtailadmin/workflows/create.html")
        self.assertBreadcrumbsItemsRendered(
            [
                {"label": "Workflows", "url": "/admin/workflows/list/"},
                {"label": "New: Workflow", "url": ""},
            ],
            response.content,
        )

        # Check the correct data attributes have been set on the form
        soup = self.get_soup(response.content)
        workflow_pages_panel = soup.find(id="workflow-pages-section")
        self.assertIn(
            "w-formset",
            workflow_pages_panel.attrs["data-controller"],
        )
        self.assertEqual(
            "totalFormsInput",
            workflow_pages_panel.find(id="id_pages-TOTAL_FORMS").attrs[
                "data-w-formset-target"
            ],
        )
        self.assertEqual(
            "template",
            workflow_pages_panel.find("template").attrs["data-w-formset-target"],
        )

        tbody = workflow_pages_panel.find("table").find("tbody")
        self.assertEqual(
            "forms",
            tbody.attrs["data-w-formset-target"],
        )

        row = tbody.find("tr")
        self.assertEqual(
            "child",
            row.attrs["data-w-formset-target"],
        )
        self.assertEqual(
            "deleteInput",
            row.find(id="id_pages-0-DELETE").attrs["data-w-formset-target"],
        )

    def test_post(self):
        response = self.post(
            {
                "name": ["test_workflow"],
                "active": ["on"],
                "workflow_tasks-TOTAL_FORMS": ["2"],
                "workflow_tasks-INITIAL_FORMS": ["0"],
                "workflow_tasks-MIN_NUM_FORMS": ["0"],
                "workflow_tasks-MAX_NUM_FORMS": ["1000"],
                "workflow_tasks-0-task": [str(self.task_1.id)],
                "workflow_tasks-0-id": [""],
                "workflow_tasks-0-ORDER": ["1"],
                "workflow_tasks-0-DELETE": [""],
                "workflow_tasks-1-task": [str(self.task_2.id)],
                "workflow_tasks-1-id": [""],
                "workflow_tasks-1-ORDER": ["2"],
                "workflow_tasks-1-DELETE": [""],
                "pages-TOTAL_FORMS": ["2"],
                "pages-INITIAL_FORMS": ["1"],
                "pages-MIN_NUM_FORMS": ["0"],
                "pages-MAX_NUM_FORMS": ["1000"],
                "pages-0-page": [str(self.root_page.id)],
                "pages-0-DELETE": [""],
                "pages-1-page": [""],
                "pages-1-DELETE": [""],
                "content_types": [str(self.snippet_content_type.id)],
            }
        )

        # Should redirect back to index
        self.assertRedirects(response, reverse("wagtailadmin_workflows:index"))

        # Check that the workflow was created
        workflows = Workflow.objects.filter(name="test_workflow", active=True)
        self.assertEqual(workflows.count(), 1)

        workflow = workflows.first()

        # Check that the tasks are associated with the workflow
        self.assertEqual(
            [self.task_1.task_ptr, self.task_2.task_ptr], list(workflow.tasks)
        )

        # Check that the tasks have sort_order set on WorkflowTask correctly
        self.assertEqual(
            WorkflowTask.objects.get(
                workflow=workflow, task=self.task_1.task_ptr
            ).sort_order,
            0,
        )
        self.assertEqual(
            WorkflowTask.objects.get(
                workflow=workflow, task=self.task_2.task_ptr
            ).sort_order,
            1,
        )

        # Check that the page has the workflow assigned
        self.assertEqual(self.root_page.get_workflow(), workflow)

        # Check that the snippet model has the default workflow assigned
        self.assertEqual(FullFeaturedSnippet.get_default_workflow(), workflow)

        # Check that the instance of the snippet model has the workflow assigned
        self.assertEqual(self.snippet.get_workflow(), workflow)

    def test_permissions(self):
        self.login(user=self.editor)
        response = self.get()
        self.assertEqual(response.status_code, 302)
        full_context = {
            key: value for context in response.context for key, value in context.items()
        }
        self.assertEqual(
            full_context["message"],
            "Sorry, you do not have permission to access this area.",
        )

        self.login(user=self.moderator)
        response = self.get()
        self.assertEqual(response.status_code, 200)

    def test_page_already_has_workflow_check(self):
        workflow = Workflow.objects.create(name="existing_workflow")
        WorkflowPage.objects.create(workflow=workflow, page=self.root_page)

        response = self.post(
            {
                "name": ["test_workflow"],
                "active": ["on"],
                "workflow_tasks-TOTAL_FORMS": ["2"],
                "workflow_tasks-INITIAL_FORMS": ["0"],
                "workflow_tasks-MIN_NUM_FORMS": ["0"],
                "workflow_tasks-MAX_NUM_FORMS": ["1000"],
                "workflow_tasks-0-task": [str(self.task_1.id)],
                "workflow_tasks-0-id": [""],
                "workflow_tasks-0-ORDER": ["1"],
                "workflow_tasks-0-DELETE": [""],
                "workflow_tasks-1-task": [str(self.task_2.id)],
                "workflow_tasks-1-id": [""],
                "workflow_tasks-1-ORDER": ["2"],
                "workflow_tasks-1-DELETE": [""],
                "pages-TOTAL_FORMS": ["2"],
                "pages-INITIAL_FORMS": ["1"],
                "pages-MIN_NUM_FORMS": ["0"],
                "pages-MAX_NUM_FORMS": ["1000"],
                "pages-0-page": [str(self.root_page.id)],
                "pages-0-DELETE": [""],
                "pages-1-page": [""],
                "pages-1-DELETE": [""],
            }
        )

        self.assertEqual(response.status_code, 200)
        self.assertFormSetError(
            response.context["pages_formset"],
            0,
            "page",
            ["This page already has workflow 'existing_workflow' assigned."],
        )

    def test_snippet_already_has_workflow_check(self):
        workflow = Workflow.objects.create(name="existing_workflow")
        WorkflowContentType.objects.create(
            workflow=workflow, content_type=self.snippet_content_type
        )

        response = self.post(
            {
                "name": ["test_workflow"],
                "active": ["on"],
                "workflow_tasks-TOTAL_FORMS": ["2"],
                "workflow_tasks-INITIAL_FORMS": ["0"],
                "workflow_tasks-MIN_NUM_FORMS": ["0"],
                "workflow_tasks-MAX_NUM_FORMS": ["1000"],
                "workflow_tasks-0-task": [str(self.task_1.id)],
                "workflow_tasks-0-id": [""],
                "workflow_tasks-0-ORDER": ["1"],
                "workflow_tasks-0-DELETE": [""],
                "workflow_tasks-1-task": [str(self.task_2.id)],
                "workflow_tasks-1-id": [""],
                "workflow_tasks-1-ORDER": ["2"],
                "workflow_tasks-1-DELETE": [""],
                "pages-TOTAL_FORMS": ["2"],
                "pages-INITIAL_FORMS": ["1"],
                "pages-MIN_NUM_FORMS": ["0"],
                "pages-MAX_NUM_FORMS": ["1000"],
                "pages-0-page": [str(self.root_page.id)],
                "pages-0-DELETE": [""],
                "pages-1-page": [""],
                "pages-1-DELETE": [""],
                "content_types": [str(self.snippet_content_type.id)],
            }
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            "Snippet 'Full-featured snippet' already has workflow 'existing_workflow' assigned.",
            count=1,
            html=True,
        )

        # Check that the assigned workflow is not changed
        link = WorkflowContentType.objects.get(content_type=self.snippet_content_type)
        self.assertEqual(link.workflow, workflow)


class TestWorkflowsEditView(AdminTemplateTestUtils, WagtailTestUtils, TestCase):
    def setUp(self):
        delete_existing_workflows()
        self.user = self.login()
        self.workflow = Workflow.objects.create(name="workflow_to_edit")
        self.task_1 = SimpleTask.objects.create(name="first_task")
        self.task_2 = SimpleTask.objects.create(name="second_task")
        self.inactive_task = SimpleTask.objects.create(
            name="inactive_task", active=False
        )
        self.workflow_task = WorkflowTask.objects.create(
            workflow=self.workflow, task=self.task_1.task_ptr, sort_order=0
        )
        self.page = Page.objects.first()
        self.snippet_content_type = ContentType.objects.get_for_model(
            FullFeaturedSnippet
        )
        WorkflowPage.objects.create(workflow=self.workflow, page=self.page)
        WorkflowContentType.objects.create(
            workflow=self.workflow, content_type=self.snippet_content_type
        )

        self.editor = self.create_user(
            username="editor",
            email="editor@email.com",
            password="password",
        )
        editors = Group.objects.get(name="Editors")
        editors.user_set.add(self.editor)

        self.moderator = self.create_user(
            username="moderator",
            email="moderator@email.com",
            password="password",
        )
        moderators = Group.objects.get(name="Moderators")
        moderators.user_set.add(self.moderator)
        moderators.permissions.add(Permission.objects.get(codename="change_workflow"))

    def get(self, params={}):
        return self.client.get(
            reverse("wagtailadmin_workflows:edit", args=[self.workflow.id]), params
        )

    def post(self, post_data={}):
        return self.client.post(
            reverse("wagtailadmin_workflows:edit", args=[self.workflow.id]), post_data
        )

    def test_get(self):
        response = self.get()
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "wagtailadmin/workflows/edit.html")
        self.assertBreadcrumbsItemsRendered(
            [
                {"url": "/admin/workflows/list/", "label": "Workflows"},
                {"url": "", "label": str(self.workflow)},
            ],
            response.content,
        )

        # Check that the list of pages has the page to which this workflow is assigned
        self.assertContains(response, self.page.title)

        # Check the correct data attributes have been set on the form
        soup = self.get_soup(response.content)

        workflow_pages_panel = soup.find(id="workflow-pages-section")
        self.assertIn(
            "w-formset",
            workflow_pages_panel.attrs["data-controller"],
        )
        self.assertEqual(
            "totalFormsInput",
            workflow_pages_panel.find(id="id_pages-TOTAL_FORMS").attrs[
                "data-w-formset-target"
            ],
        )
        self.assertEqual(
            "template",
            workflow_pages_panel.find("template").attrs["data-w-formset-target"],
        )

        tbody = workflow_pages_panel.find("table").find("tbody")
        self.assertEqual(
            "forms",
            tbody.attrs["data-w-formset-target"],
        )

        row = tbody.find("tr")
        self.assertEqual(
            "child",
            row.attrs["data-w-formset-target"],
        )
        self.assertEqual(
            "deleteInput",
            row.find(id="id_pages-0-DELETE").attrs["data-w-formset-target"],
        )

    def test_post(self):
        response = self.post(
            {
                "name": ["Edited workflow"],
                "active": ["on"],
                "workflow_tasks-TOTAL_FORMS": ["2"],
                "workflow_tasks-INITIAL_FORMS": ["1"],
                "workflow_tasks-MIN_NUM_FORMS": ["0"],
                "workflow_tasks-MAX_NUM_FORMS": ["1000"],
                "workflow_tasks-0-task": [str(self.task_1.id)],
                "workflow_tasks-0-id": [str(self.workflow_task.id)],
                "workflow_tasks-0-ORDER": ["1"],
                "workflow_tasks-0-DELETE": [""],
                "workflow_tasks-1-task": [str(self.task_2.id)],
                "workflow_tasks-1-id": [""],
                "workflow_tasks-1-ORDER": ["2"],
                "workflow_tasks-1-DELETE": [""],
                "pages-TOTAL_FORMS": ["2"],
                "pages-INITIAL_FORMS": ["1"],
                "pages-MIN_NUM_FORMS": ["0"],
                "pages-MAX_NUM_FORMS": ["1000"],
                "pages-0-page": [str(self.page.id)],
                "pages-0-DELETE": [""],
                "pages-1-page": [""],
                "pages-1-DELETE": [""],
                "content_types": [str(self.snippet_content_type.id)],
            }
        )

        # Should redirect back to index
        self.assertRedirects(response, reverse("wagtailadmin_workflows:index"))

        # Check that the workflow was edited
        workflows = Workflow.objects.filter(name="Edited workflow", active=True)
        self.assertEqual(workflows.count(), 1)

        workflow = workflows.first()

        # Check that the tasks are associated with the workflow
        self.assertEqual(
            [self.task_1.task_ptr, self.task_2.task_ptr], list(workflow.tasks)
        )

        # Check that the tasks have sort_order set on WorkflowTask correctly
        self.assertEqual(
            WorkflowTask.objects.get(
                workflow=workflow, task=self.task_1.task_ptr
            ).sort_order,
            0,
        )
        self.assertEqual(
            WorkflowTask.objects.get(
                workflow=workflow, task=self.task_2.task_ptr
            ).sort_order,
            1,
        )

        # Check that the page has the workflow assigned
        self.assertEqual(self.page.get_workflow(), workflow)

        # Check that the snippet model has the default workflow assigned
        self.assertEqual(FullFeaturedSnippet.get_default_workflow(), workflow)

        # Check that the instance of the snippet model has the workflow assigned
        snippet = FullFeaturedSnippet.objects.create(text="foo")
        self.assertEqual(snippet.get_workflow(), workflow)

    def test_permissions(self):
        self.login(user=self.editor)
        response = self.get()
        self.assertEqual(response.status_code, 302)
        full_context = {
            key: value for context in response.context for key, value in context.items()
        }
        self.assertEqual(
            full_context["message"],
            "Sorry, you do not have permission to access this area.",
        )

        self.login(user=self.moderator)
        response = self.get()
        self.assertEqual(response.status_code, 200)

    def test_admin_url_finder(self):
        editor_url_finder = AdminURLFinder(self.editor)
        self.assertIsNone(editor_url_finder.get_edit_url(self.workflow))
        moderator_url_finder = AdminURLFinder(self.moderator)
        expected_url = "/admin/workflows/edit/%d/" % self.workflow.pk
        self.assertEqual(moderator_url_finder.get_edit_url(self.workflow), expected_url)

    def test_duplicate_page_check(self):
        response = self.post(
            {
                "name": [str(self.workflow.name)],
                "active": ["on"],
                "workflow_tasks-TOTAL_FORMS": ["2"],
                "workflow_tasks-INITIAL_FORMS": ["1"],
                "workflow_tasks-MIN_NUM_FORMS": ["0"],
                "workflow_tasks-MAX_NUM_FORMS": ["1000"],
                "workflow_tasks-0-task": [str(self.task_1.id)],
                "workflow_tasks-0-id": [str(self.workflow_task.id)],
                "workflow_tasks-0-ORDER": ["1"],
                "workflow_tasks-0-DELETE": [""],
                "workflow_tasks-1-task": [str(self.task_2.id)],
                "workflow_tasks-1-id": [""],
                "workflow_tasks-1-ORDER": ["2"],
                "workflow_tasks-1-DELETE": [""],
                "pages-TOTAL_FORMS": ["2"],
                "pages-INITIAL_FORMS": ["1"],
                "pages-MIN_NUM_FORMS": ["0"],
                "pages-MAX_NUM_FORMS": ["1000"],
                "pages-0-page": [str(self.page.id)],
                "pages-0-DELETE": [""],
                "pages-1-page": [str(self.page.id)],
                "pages-1-DELETE": [""],
            }
        )

        self.assertEqual(response.status_code, 200)
        self.assertFormSetError(
            response.context["pages_formset"],
            None,
            None,
            ["You cannot assign this workflow to the same page multiple times."],
        )

    def test_snippet_already_has_workflow_check(self):
        # Change the workflow for the snippet content type to another workflow
        other_workflow = Workflow.objects.create(name="other_workflow")
        WorkflowContentType.objects.filter(
            content_type=self.snippet_content_type
        ).update(workflow=other_workflow)

        # Try to save the workflow with the snippet content type assigned
        response = self.post(
            {
                "name": [str(self.workflow.name)],
                "active": ["on"],
                "workflow_tasks-TOTAL_FORMS": ["2"],
                "workflow_tasks-INITIAL_FORMS": ["1"],
                "workflow_tasks-MIN_NUM_FORMS": ["0"],
                "workflow_tasks-MAX_NUM_FORMS": ["1000"],
                "workflow_tasks-0-task": [str(self.task_1.id)],
                "workflow_tasks-0-id": [str(self.workflow_task.id)],
                "workflow_tasks-0-ORDER": ["1"],
                "workflow_tasks-0-DELETE": [""],
                "workflow_tasks-1-task": [str(self.task_2.id)],
                "workflow_tasks-1-id": [""],
                "workflow_tasks-1-ORDER": ["2"],
                "workflow_tasks-1-DELETE": [""],
                "pages-TOTAL_FORMS": ["2"],
                "pages-INITIAL_FORMS": ["1"],
                "pages-MIN_NUM_FORMS": ["0"],
                "pages-MAX_NUM_FORMS": ["1000"],
                "pages-0-page": [str(self.page.id)],
                "pages-0-DELETE": [""],
                "pages-1-page": [""],
                "pages-1-DELETE": [""],
                "content_types": [str(self.snippet_content_type.id)],
            }
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            "Snippet 'Full-featured snippet' already has workflow 'other_workflow' assigned.",
            count=1,
            html=True,
        )

        # Check that the assigned workflow is not changed
        link = WorkflowContentType.objects.get(content_type=self.snippet_content_type)
        self.assertEqual(link.workflow, other_workflow)

    def test_render_enable_button_if_workflow_disabled(self):
        self.workflow.active = False
        self.workflow.save()
        response = self.get()
        soup = self.get_soup(response.content)
        enable_url = reverse("wagtailadmin_workflows:enable", args=(self.workflow.pk,))
        enable_button = soup.find("button", {"data-w-action-url-value": enable_url})
        self.assertIsNotNone(enable_button)

    def test_render_enable_button_if_workflow_disabled_minimal_permissions(self):
        self.user.is_superuser = False
        self.user.save()
        self.user.user_permissions.add(
            Permission.objects.get(
                content_type__app_label="wagtailadmin",
                codename="access_admin",
            ),
            Permission.objects.get(codename="add_workflow"),
            Permission.objects.get(codename="change_workflow"),
        )
        self.workflow.active = False
        self.workflow.save()
        response = self.get()
        soup = self.get_soup(response.content)
        enable_url = reverse("wagtailadmin_workflows:enable", args=(self.workflow.pk,))
        enable_button = soup.find("button", {"data-w-action-url-value": enable_url})
        self.assertIsNotNone(enable_button)

    def test_render_enable_button_if_workflow_disabled_no_permissions(self):
        self.user.is_superuser = False
        self.user.save()
        self.user.user_permissions.add(
            Permission.objects.get(
                content_type__app_label="wagtailadmin",
                codename="access_admin",
            ),
            Permission.objects.get(codename="change_workflow"),
        )
        self.workflow.active = False
        self.workflow.save()
        response = self.get()
        soup = self.get_soup(response.content)
        enable_url = reverse("wagtailadmin_workflows:enable", args=(self.workflow.pk,))
        enable_button = soup.find("button", {"data-w-action-url-value": enable_url})
        self.assertIsNone(enable_button)

    def test_pages_and_content_types_ignored_if_workflow_disabled(self):
        self.workflow.active = False
        self.workflow.save()
        self.workflow.workflow_pages.all().delete()
        self.workflow.workflow_content_types.all().delete()

        response = self.post(
            {
                "name": [str(self.workflow.name)],
                "active": ["on"],
                "workflow_tasks-TOTAL_FORMS": ["2"],
                "workflow_tasks-INITIAL_FORMS": ["1"],
                "workflow_tasks-MIN_NUM_FORMS": ["0"],
                "workflow_tasks-MAX_NUM_FORMS": ["1000"],
                "workflow_tasks-0-task": [str(self.task_1.id)],
                "workflow_tasks-0-id": [str(self.workflow_task.id)],
                "workflow_tasks-0-ORDER": ["1"],
                "workflow_tasks-0-DELETE": [""],
                "workflow_tasks-1-task": [str(self.task_2.id)],
                "workflow_tasks-1-id": [""],
                "workflow_tasks-1-ORDER": ["2"],
                "workflow_tasks-1-DELETE": [""],
                "pages-TOTAL_FORMS": ["2"],
                "pages-INITIAL_FORMS": ["1"],
                "pages-MIN_NUM_FORMS": ["0"],
                "pages-MAX_NUM_FORMS": ["1000"],
                "pages-0-page": [str(self.page.id)],
                "pages-0-DELETE": [""],
                "pages-1-page": [""],
                "pages-1-DELETE": [""],
                "content_types": [str(self.snippet_content_type.id)],
            }
        )

        # Should redirect back to index
        self.assertRedirects(response, reverse("wagtailadmin_workflows:index"))

        # Check that the pages weren't added to the workflow
        self.workflow.refresh_from_db()
        self.assertFalse(self.workflow.workflow_pages.exists())

        # Check that the workflow is not assigned to any snippet model
        self.assertFalse(self.workflow.workflow_content_types.exists())


class TestRemoveWorkflow(WagtailTestUtils, TestCase):
    fixtures = ["test.json"]

    def setUp(self):
        delete_existing_workflows()
        self.login()
        self.workflow = Workflow.objects.create(name="workflow")
        self.page = Page.objects.first()
        WorkflowPage.objects.create(workflow=self.workflow, page=self.page)

        self.editor = self.create_user(
            username="editor",
            email="editor@email.com",
            password="password",
        )
        editors = Group.objects.get(name="Editors")
        editors.user_set.add(self.editor)

        self.moderator = self.create_user(
            username="moderator",
            email="moderator@email.com",
            password="password",
        )
        moderators = Group.objects.get(name="Moderators")
        moderators.user_set.add(self.moderator)
        moderators.permissions.add(Permission.objects.get(codename="change_workflow"))

    def post(self, post_data={}):
        return self.client.post(
            reverse(
                "wagtailadmin_workflows:remove", args=[self.page.id, self.workflow.id]
            ),
            post_data,
        )

    def test_post(self):
        # Check that a WorkflowPage instance is removed correctly
        self.post()
        self.assertEqual(
            WorkflowPage.objects.filter(workflow=self.workflow, page=self.page).count(),
            0,
        )

    def test_no_permissions(self):
        self.login(user=self.editor)
        response = self.post()
        # The WorkflowPage instance should not be removed
        self.assertEqual(
            WorkflowPage.objects.filter(workflow=self.workflow, page=self.page).count(),
            1,
        )
        self.assertEqual(response.status_code, 302)

    def test_post_with_permission(self):
        self.login(user=self.moderator)
        response = self.post()
        self.assertEqual(response.status_code, 302)


class TestTaskIndexView(AdminTemplateTestUtils, WagtailTestUtils, TestCase):
    def setUp(self):
        delete_existing_workflows()
        self.login()

        self.editor = self.create_user(
            username="editor",
            email="editor@email.com",
            password="password",
        )
        editors = Group.objects.get(name="Editors")
        editors.user_set.add(self.editor)

        self.moderator = self.create_user(
            username="moderator",
            email="moderator@email.com",
            password="password",
        )
        moderators = Group.objects.get(name="Moderators")
        moderators.user_set.add(self.moderator)
        moderators.permissions.add(Permission.objects.get(codename="change_task"))

    def get(self, params={}):
        return self.client.get(reverse("wagtailadmin_workflows:task_index"), params)

    def test_simple(self):
        response = self.get()
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "wagtailadmin/workflows/task_index.html")
        self.assertBreadcrumbsItemsRendered(
            [{"url": "", "label": "Workflow tasks"}],
            response.content,
        )

        # Initially there should be no tasks listed
        self.assertContains(response, "There are no enabled tasks")

        SimpleTask.objects.create(name="test_task", active=True)

        # Now the listing should contain our task
        response = self.get()
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "wagtailadmin/workflows/task_index.html")
        self.assertNotContains(response, "There are no enabled tasks")
        self.assertContains(response, "test_task")

    def test_deactivated(self):
        Task.objects.create(name="test_task", active=False)

        # The listing should contain our task, as well as marking it as disabled
        response = self.get(params={"show_disabled": "true"})
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "No tasks have been created.")
        self.assertContains(response, "test_task")
        self.assertContains(
            response, '<span class="w-status">Disabled</span>', html=True
        )
        # Should display the "Show disabled" option as a filter
        soup = self.get_soup(response.content)
        active_filter = soup.select_one('[data-w-active-filter-id="id_show_disabled"]')
        self.assertIsNotNone(active_filter)
        self.assertEqual(
            active_filter.get_text(separator=" ", strip=True),
            "Show disabled: Yes",
        )
        show_disabled_yes = soup.select_one('input[name="show_disabled"][value="true"]')
        self.assertIsNotNone(show_disabled_yes)
        self.assertTrue(show_disabled_yes.has_attr("checked"))

        # The listing should not contain task if show_disabled query parameter is 'False'
        response = self.get(params={})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "There are no enabled tasks")
        self.assertNotContains(response, "test_task")

        # Should not display any active filters,
        # and the "Show disabled" option should be set to "No" by default
        soup = self.get_soup(response.content)
        active_filter = soup.select_one('[data-w-active-filter-id="id_show_disabled"]')
        self.assertIsNone(active_filter)
        show_disabled_no = soup.select_one('input[name="show_disabled"][value="false"]')
        self.assertIsNotNone(show_disabled_no)
        self.assertTrue(show_disabled_no.has_attr("checked"))

    def test_permissions(self):
        self.login(user=self.editor)
        response = self.get()
        self.assertEqual(response.status_code, 302)
        full_context = {
            key: value for context in response.context for key, value in context.items()
        }
        self.assertEqual(
            full_context["message"],
            "Sorry, you do not have permission to access this area.",
        )

        self.login(user=self.moderator)
        response = self.get()
        self.assertEqual(response.status_code, 200)

    def test_ordering(self):
        tasks = sorted(
            [
                # Mix up the creation order to ensure we're not ordering by PK
                Task.objects.create(name="task_1"),
                Task.objects.create(name="task_3"),
                Task.objects.create(name="task_2"),
            ],
            key=lambda task: task.name,
        )

        response = self.get()
        self.assertEqual(response.status_code, 200)
        self.assertSequenceEqual(response.context["object_list"], tasks)
        self.assertEqual(response.context["object_list"].query.order_by, ("name",))

        response = self.get(params={"ordering": "name"})
        self.assertEqual(response.status_code, 200)
        self.assertSequenceEqual(response.context["object_list"], tasks)
        self.assertEqual(response.context["object_list"].query.order_by, ("name",))

        response = self.get(params={"ordering": "-name"})
        self.assertEqual(response.status_code, 200)
        self.assertSequenceEqual(response.context["object_list"], tasks[::-1])
        self.assertEqual(response.context["object_list"].query.order_by, ("-name",))

    def test_search(self):
        Task.objects.create(name="foo task")
        Task.objects.create(name="bar task")
        Task.objects.create(name="bar world task")

        response = self.get(params={"q": "bAr"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "bar task")
        self.assertContains(response, "bar world task")
        self.assertNotContains(response, "foo task")

    def test_search_results(self):
        Task.objects.create(name="foo task")
        Task.objects.create(name="bar task")
        Task.objects.create(name="bar world task")

        response = self.client.get(
            reverse("wagtailadmin_workflows:task_index_results"),
            {"q": "AR"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertBreadcrumbsNotRendered(response.content)
        self.assertContains(response, "bar task")
        self.assertContains(response, "bar world task")
        self.assertNotContains(response, "foo task")

    def test_task_type_filter(self):
        SimpleTask.objects.create(name="easy task")
        SimpleTask.objects.create(name="medium task")
        GroupApprovalTask.objects.create(name="complex task")

        simple_ct = ContentType.objects.get_for_model(SimpleTask).pk
        group_approval_ct = ContentType.objects.get_for_model(GroupApprovalTask).pk

        response = self.get(params={"content_type": [simple_ct]})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "easy task")
        self.assertContains(response, "medium task")
        self.assertNotContains(response, "complex task")

        # Should display the active filter
        soup = self.get_soup(response.content)
        active_filter = soup.select_one('[data-w-active-filter-id="id_content_type"]')
        self.assertIsNotNone(active_filter)
        self.assertEqual(
            active_filter.get_text(separator=" ", strip=True),
            "Type: Simple task",
        )
        simple_ct_box = soup.select_one(
            f'input[name="content_type"][value="{simple_ct}"]'
        )
        self.assertIsNotNone(simple_ct_box)
        self.assertTrue(simple_ct_box.has_attr("checked"))
        group_approval_ct_box = soup.select_one(
            f'input[name="content_type"][value="{group_approval_ct}"]'
        )
        self.assertIsNotNone(group_approval_ct_box)
        self.assertFalse(group_approval_ct_box.has_attr("checked"))

        # Should allow multiple content types to be selected
        response = self.get(params={"content_type": [simple_ct, group_approval_ct]})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "easy task")
        self.assertContains(response, "medium task")
        self.assertContains(response, "complex task")

        # Should display the active filters
        soup = self.get_soup(response.content)
        active_filters = soup.select('[data-w-active-filter-id="id_content_type"]')
        self.assertCountEqual(
            [filter.get_text(separator=" ", strip=True) for filter in active_filters],
            {"Type: Simple task", "Type: Group approval task"},
        )
        simple_ct_box = soup.select_one(
            f'input[name="content_type"][value="{simple_ct}"]'
        )
        self.assertIsNotNone(simple_ct_box)
        self.assertTrue(simple_ct_box.has_attr("checked"))
        group_approval_ct_box = soup.select_one(
            f'input[name="content_type"][value="{group_approval_ct}"]'
        )
        self.assertIsNotNone(group_approval_ct_box)
        self.assertTrue(group_approval_ct_box.has_attr("checked"))

    def test_task_type_filter_hidden_if_single_task_type(self):
        SimpleTask.objects.create(name="easy task")
        SimpleTask.objects.create(name="medium task")
        GroupApprovalTask.objects.create(name="complex task")

        simple_ct = ContentType.objects.get_for_model(SimpleTask).pk

        with mock.patch(
            "wagtail.admin.views.workflows.get_task_types"
        ) as get_task_types:
            get_task_types.return_value = [SimpleTask]
            response = self.get({"content_type": [simple_ct]})

        # Should not be filtered
        self.assertContains(response, "easy task")
        self.assertContains(response, "medium task")
        self.assertContains(response, "complex task")

        # Should not display the content type filter
        soup = self.get_soup(response.content)
        active_filters = soup.select_one(".w-active_filters")
        self.assertIsNone(active_filters)
        content_type_filter = soup.select_one('input[name="content_type"]')
        self.assertIsNone(content_type_filter)

    def test_pagination(self):
        Task.objects.bulk_create([Task(name=f"task_{i}") for i in range(1, 120)])

        url = reverse("wagtailadmin_workflows:task_index")

        response = self.get({"p": 2})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context["object_list"]), 50)
        self.assertContains(response, url + "?p=1")
        self.assertContains(response, url + "?p=2")
        self.assertContains(response, url + "?p=3")

        response = self.get({"p": 4})
        # Check response
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "wagtailadmin/workflows/task_index.html")

        # Check that we got the last page
        self.assertEqual(
            response.context["page_obj"].number,
            response.context["paginator"].num_pages,
        )

    def test_num_queries(self):
        workflows = [Workflow.objects.create(name=f"workflow_{i}") for i in range(7)]
        tasks = [Task.objects.create(name=f"task_{i}") for i in range(20)]
        WorkflowTask.objects.bulk_create(
            [
                WorkflowTask(workflow=workflow, task=task, sort_order=0)
                for workflow in workflows
                for task in tasks
            ]
        )
        self.get()

        with self.assertNumQueries(12):
            response = self.get()
        self.assertContains(response, "+2 more", count=20)

        tasks = [Task.objects.create(name=f"task_{i}") for i in range(21, 41)]
        WorkflowTask.objects.bulk_create(
            [
                WorkflowTask(workflow=workflow, task=task, sort_order=0)
                for workflow in workflows
                for task in tasks
            ]
        )

        with self.assertNumQueries(12):
            response = self.get()
        self.assertContains(response, "+2 more", count=40)


class TestCreateTaskView(AdminTemplateTestUtils, WagtailTestUtils, TestCase):
    def setUp(self):
        delete_existing_workflows()
        self.login()

        self.editor = self.create_user(
            username="editor",
            email="editor@email.com",
            password="password",
        )
        editors = Group.objects.get(name="Editors")
        editors.user_set.add(self.editor)

        self.moderator = self.create_user(
            username="moderator",
            email="moderator@email.com",
            password="password",
        )
        moderators = Group.objects.get(name="Moderators")
        moderators.user_set.add(self.moderator)
        moderators.permissions.add(Permission.objects.get(codename="add_task"))

    def get(self, url_kwargs=None, params={}):
        url_kwargs = url_kwargs or {}
        url_kwargs.setdefault("app_label", SimpleTask._meta.app_label)
        url_kwargs.setdefault("model_name", SimpleTask._meta.model_name)
        return self.client.get(
            reverse("wagtailadmin_workflows:add_task", kwargs=url_kwargs), params
        )

    def post(self, post_data={}):
        return self.client.post(
            reverse(
                "wagtailadmin_workflows:add_task",
                kwargs={
                    "app_label": SimpleTask._meta.app_label,
                    "model_name": SimpleTask._meta.model_name,
                },
            ),
            post_data,
        )

    def test_get(self):
        response = self.get()
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "wagtailadmin/workflows/create_task.html")
        self.assertBreadcrumbsItemsRendered(
            [
                {"label": "Tasks", "url": "/admin/workflows/tasks/index/"},
                {"label": "New: Simple task", "url": ""},
            ],
            response.content,
        )

    def test_get_with_non_task_model(self):
        response = self.get(
            url_kwargs={"app_label": "wagtailcore", "model_name": "Site"}
        )
        self.assertEqual(response.status_code, 404)

    def test_get_with_base_task_model(self):
        response = self.get(
            url_kwargs={"app_label": "wagtailcore", "model_name": "Task"}
        )
        self.assertEqual(response.status_code, 404)

    def test_post(self):
        response = self.post({"name": "test_task", "active": "on"})

        # Should redirect back to index
        self.assertRedirects(response, reverse("wagtailadmin_workflows:task_index"))

        # Check that the task was created
        tasks = Task.objects.filter(name="test_task", active=True)
        self.assertEqual(tasks.count(), 1)

    def test_permissions(self):
        self.login(user=self.editor)
        response = self.get()
        self.assertEqual(response.status_code, 302)
        full_context = {
            key: value for context in response.context for key, value in context.items()
        }
        self.assertEqual(
            full_context["message"],
            "Sorry, you do not have permission to access this area.",
        )

        self.login(user=self.moderator)
        response = self.get()
        self.assertEqual(response.status_code, 200)


class TestSelectTaskTypeView(WagtailTestUtils, TestCase):
    def setUp(self):
        delete_existing_workflows()
        self.login()

    def get(self):
        return self.client.get(reverse("wagtailadmin_workflows:select_task_type"))

    def test_get(self):
        response = self.get()
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(
            response, "wagtailadmin/workflows/select_task_type.html"
        )

        # Check that the list of available task types includes SimpleTask and GroupApprovalTask
        self.assertContains(response, SimpleTask.get_verbose_name())
        self.assertContains(response, GroupApprovalTask.get_verbose_name())
        self.assertContains(response, GroupApprovalTask.get_description())

    def test_get_single_task_type(self):
        with mock.patch(
            "wagtail.admin.views.workflows.get_task_types"
        ) as get_task_types:
            get_task_types.return_value = [GroupApprovalTask]
            response = self.get()

        # Should redirect to the create task view for the only available task type
        self.assertRedirects(
            response,
            reverse(
                "wagtailadmin_workflows:add_task",
                args=(
                    GroupApprovalTask._meta.app_label,
                    GroupApprovalTask._meta.model_name,
                ),
            ),
        )


class TestEditTaskView(AdminTemplateTestUtils, WagtailTestUtils, TestCase):
    def setUp(self):
        delete_existing_workflows()
        self.user = self.login()
        self.task = GroupApprovalTask.objects.create(name="test_task")

        self.editor = self.create_user(
            username="editor",
            email="editor@email.com",
            password="password",
        )
        editors = Group.objects.get(name="Editors")
        editors.user_set.add(self.editor)

        self.moderator = self.create_user(
            username="moderator",
            email="moderator@email.com",
            password="password",
        )
        moderators = Group.objects.get(name="Moderators")
        moderators.user_set.add(self.moderator)
        moderators.permissions.add(Permission.objects.get(codename="change_task"))

    def get(self, params={}):
        return self.client.get(
            reverse("wagtailadmin_workflows:edit_task", args=[self.task.id]), params
        )

    def post(self, post_data={}):
        return self.client.post(
            reverse("wagtailadmin_workflows:edit_task", args=[self.task.id]), post_data
        )

    def test_get(self):
        response = self.get()
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "wagtailadmin/workflows/edit_task.html")
        self.assertBreadcrumbsItemsRendered(
            [
                {"url": "/admin/workflows/tasks/index/", "label": "Tasks"},
                {"url": "", "label": str(self.task)},
            ],
            response.content,
        )

    def test_post(self):
        self.assertEqual(self.task.groups.count(), 0)
        editors = Group.objects.get(name="Editors")

        response = self.post(
            {"name": "test_task_modified", "active": "on", "groups": [str(editors.id)]}
        )

        # Should redirect back to index
        self.assertRedirects(response, reverse("wagtailadmin_workflows:task_index"))

        # Check that the task was updated
        task = GroupApprovalTask.objects.get(id=self.task.id)

        # The task name cannot be changed
        self.assertEqual(task.name, "test_task")

        # This request should've added a group to the task
        self.assertEqual(task.groups.count(), 1)
        self.assertTrue(task.groups.filter(id=editors.id).exists())

    def test_permissions(self):
        self.login(user=self.editor)
        response = self.get()
        self.assertEqual(response.status_code, 302)
        full_context = {
            key: value for context in response.context for key, value in context.items()
        }
        self.assertEqual(
            full_context["message"],
            "Sorry, you do not have permission to access this area.",
        )

        self.login(user=self.moderator)
        response = self.get()
        self.assertEqual(response.status_code, 200)

    def test_render_enable_button_if_task_disabled(self):
        self.task.active = False
        self.task.save()
        response = self.get()
        soup = self.get_soup(response.content)
        enable_url = reverse("wagtailadmin_workflows:enable_task", args=(self.task.pk,))
        enable_button = soup.find("button", {"data-w-action-url-value": enable_url})
        self.assertIsNotNone(enable_button)

    def test_render_enable_button_if_task_disabled_minimal_permissions(self):
        self.user.is_superuser = False
        self.user.save()
        self.user.user_permissions.add(
            Permission.objects.get(
                content_type__app_label="wagtailadmin",
                codename="access_admin",
            ),
            Permission.objects.get(codename="add_task"),
            Permission.objects.get(codename="change_task"),
        )
        self.task.active = False
        self.task.save()
        response = self.get()
        soup = self.get_soup(response.content)
        enable_url = reverse("wagtailadmin_workflows:enable_task", args=(self.task.pk,))
        enable_button = soup.find("button", {"data-w-action-url-value": enable_url})
        self.assertIsNotNone(enable_button)

    def test_render_enable_button_if_task_disabled_no_permissions(self):
        self.user.is_superuser = False
        self.user.save()
        self.user.user_permissions.add(
            Permission.objects.get(
                content_type__app_label="wagtailadmin",
                codename="access_admin",
            ),
            Permission.objects.get(codename="change_task"),
        )
        self.task.active = False
        self.task.save()
        response = self.get()
        soup = self.get_soup(response.content)
        enable_url = reverse("wagtailadmin_workflows:enable_task", args=(self.task.pk,))
        enable_button = soup.find("button", {"data-w-action-url-value": enable_url})
        self.assertIsNone(enable_button)

    def test_admin_url_finder(self):
        editor_url_finder = AdminURLFinder(self.editor)
        self.assertIsNone(editor_url_finder.get_edit_url(self.task))
        moderator_url_finder = AdminURLFinder(self.moderator)
        expected_url = "/admin/workflows/tasks/edit/%d/" % self.task.pk
        self.assertEqual(moderator_url_finder.get_edit_url(self.task), expected_url)


class BasePageWorkflowTests(AdminTemplateTestUtils, WagtailTestUtils, TestCase):
    model_name = "page"

    def setUp(self):
        delete_existing_workflows()
        self.submitter = self.create_user(
            username="submitter",
            email="submitter@email.com",
            password="password",
        )
        editors = Group.objects.get(name="Editors")
        editors.user_set.add(self.submitter)
        self.moderator = self.create_user(
            username="moderator",
            email="moderator@email.com",
            password="password",
        )
        moderators = Group.objects.get(name="Moderators")
        moderators.user_set.add(self.moderator)

        self.superuser = self.create_superuser(
            username="superuser",
            email="superuser@email.com",
            password="password",
        )

        self.login(user=self.submitter)

        self.setup_workflow_and_tasks()
        self.setup_object()

    def setup_workflow_and_tasks(self):
        self.workflow = Workflow.objects.create(name="test_workflow")
        self.task_1 = GroupApprovalTask.objects.create(name="test_task_1")
        self.task_2 = GroupApprovalTask.objects.create(name="test_task_2")
        self.task_1.groups.set(Group.objects.filter(name="Moderators"))
        self.task_2.groups.set(Group.objects.filter(name="Moderators"))
        WorkflowTask.objects.create(
            workflow=self.workflow, task=self.task_1, sort_order=1
        )
        WorkflowTask.objects.create(
            workflow=self.workflow, task=self.task_2, sort_order=2
        )

    def setup_object(self):
        # Create a page
        root_page = Page.objects.get(id=2)
        self.object = SimplePage(
            title="Hello world!",
            slug="hello-world",
            content="hello",
            live=False,
            has_unpublished_changes=True,
        )
        root_page.add_child(instance=self.object)
        self.object_class = self.object.specific_class

        # Assign to workflow
        WorkflowPage.objects.create(workflow=self.workflow, page=self.object)

    def get_url(self, view, args=None):
        return reverse(
            f"wagtailadmin_pages:{view}",
            args=(self.object.id,) if args is None else args,
        )

    def post(self, action, data=None, **kwargs):
        post_data = {
            "title": str(self.object.title),
            "slug": str(self.object.slug),
            "content": str(self.object.content),
            f"action-{action}": "True",
        }
        if data:
            post_data.update(data)
        return self.client.post(self.get_url("edit"), post_data, **kwargs)

    def workflow_action(self, action, data=None, **kwargs):
        return self.client.post(
            self.get_url(
                "workflow_action",
                args=(
                    quote(self.object.pk),
                    action,
                    self.object.current_workflow_task_state.id,
                ),
            ),
            data,
            follow=True,
            **kwargs,
        )

    def approve(self, data=None, **kwargs):
        return self.workflow_action("approve", data, **kwargs)

    def reject(self, data=None, **kwargs):
        return self.workflow_action("reject", data, **kwargs)


class BaseSnippetWorkflowTests(BasePageWorkflowTests):
    model = FullFeaturedSnippet

    def setUp(self):
        super().setUp()
        self.edit_permission = Permission.objects.get(
            content_type__app_label="tests",
            codename=f"change_{self.model._meta.model_name}",
        )
        self.publish_permission = Permission.objects.get(
            content_type__app_label="tests",
            codename=f"publish_{self.model._meta.model_name}",
        )
        self.lock_permission = Permission.objects.filter(
            content_type__app_label="tests",
            codename=f"lock_{self.model._meta.model_name}",
        ).first()
        self.submitter.user_permissions.add(self.edit_permission)
        self.moderator.user_permissions.add(
            self.edit_permission, self.publish_permission
        )
        if self.lock_permission:
            self.moderator.user_permissions.add(self.lock_permission)

    @property
    def model_name(self):
        return self.model._meta.verbose_name

    def setup_object(self):
        self.object = self.model.objects.create(
            text="Hello world!",
            live=False,
            has_unpublished_changes=True,
        )
        self.object_class = type(self.object)

        # Assign to workflow
        WorkflowContentType.objects.create(
            workflow=self.workflow,
            content_type=ContentType.objects.get_for_model(self.model),
        )

    def get_url(self, view, args=None):
        return reverse(
            self.model.snippet_viewset.get_url_name(view),
            args=(quote(self.object.pk),) if args is None else args,
        )

    def post(self, action, data=None, **kwargs):
        post_data = {
            "text": self.object.text,
            f"action-{action}": "True",
        }
        if data:
            post_data.update(data)
        return self.client.post(self.get_url("edit"), post_data, **kwargs)


class TestSubmitPageToWorkflow(BasePageWorkflowTests):
    def setUp(self):
        super().setUp()
        # Ensure a revision exists
        self.object.save_revision()

    def test_submit_for_approval_creates_states(self):
        """Test that WorkflowState and TaskState objects are correctly created when an object is submitted for approval"""

        self.post("submit")

        workflow_state = self.object.current_workflow_state

        self.assertEqual(type(workflow_state), WorkflowState)
        self.assertEqual(workflow_state.workflow, self.workflow)
        self.assertEqual(workflow_state.status, workflow_state.STATUS_IN_PROGRESS)
        self.assertEqual(workflow_state.requested_by, self.submitter)

        task_state = workflow_state.current_task_state

        self.assertEqual(type(task_state), TaskState)
        self.assertEqual(task_state.task.specific, self.task_1)
        self.assertEqual(task_state.status, task_state.STATUS_IN_PROGRESS)

    def test_submit_for_approval_changes_status_in_status_side_panel_meta(self):
        edit_url = self.get_url("edit")

        response = self.client.get(edit_url)
        self.assertContains(response, "Draft", count=1)

        # submit for approval
        self.post("submit")

        response = self.client.get(edit_url)

        # Should show the moderation status
        self.assertRegex(
            response.content.decode("utf-8"),
            rf"Sent to[\s|\n]+{self.object.current_workflow_task.name}",
        )
        self.assertContains(response, "In Moderation")
        self.assertNotContains(response, "Draft")

    def test_submit_for_approval_changes_lock_status(self):
        edit_url = self.get_url("edit")

        response = self.client.get(edit_url)

        # Should show the lock information as unlocked
        self.assertContains(response, "Unlocked", count=1)
        self.assertContains(
            response, f"Anyone can edit this {self.model_name}", count=1
        )
        self.assertNotContains(response, "Locked by workflow")
        self.assertNotContains(
            response,
            f"Only reviewers can edit and approve the {self.model_name}",
        )
        self.assertNotContains(response, self.get_url("lock"))

        # submit for approval
        self.post("submit")

        response = self.client.get(edit_url)

        # Should show the lock information as locked
        self.assertContains(response, "Locked by workflow", count=1)
        self.assertContains(
            response,
            f"Only reviewers can edit and approve the {self.model_name}",
            count=1,
        )

        self.assertNotContains(response, "Unlocked")
        self.assertNotContains(
            response,
            f"Anyone can edit this {self.model_name}",
        )

        # Should be unable to unlock
        self.assertNotContains(response, self.get_url("unlock"))

    def test_can_manual_lock_while_in_workflow(self):
        edit_url = self.get_url("edit")

        # Submit to workflow as submitter
        self.post("submit")

        # Login as a moderator to have lock permission
        self.login(self.moderator)

        response = self.client.get(edit_url)

        # Should show the lock information as unlocked
        self.assertContains(response, "Unlocked", count=1)
        self.assertContains(
            response,
            f"Reviewers can edit this {self.model_name} – lock it to prevent other reviewers from editing",
            count=1,
        )
        self.assertContains(response, self.get_url("lock"), count=1)
        self.assertNotContains(response, "Locked by workflow")
        self.assertNotContains(
            response,
            f"Only reviewers can edit and approve the {self.model_name}",
        )
        self.assertNotContains(response, self.get_url("unlock"))

    def test_can_unlock_manual_lock_while_in_workflow(self):
        edit_url = self.get_url("edit")

        # Submit to workflow as submitter
        self.post("submit")

        # Login as a moderator to have lock permission
        self.login(self.moderator)

        # Lock the object
        self.client.post(self.get_url("lock"))

        response = self.client.get(edit_url)

        # Should show the lock information as locked
        self.assertContains(response, "Locked by you", count=1)
        self.assertContains(
            response,
            f"Only you can make changes while the {self.model_name} is locked",
            count=1,
        )
        # One in the side panel, one in the message banner
        self.assertContains(response, self.get_url("unlock"), count=2)
        self.assertNotContains(response, "Locked by workflow")
        self.assertNotContains(
            response,
            f"Only reviewers can edit and approve the {self.model_name}",
        )
        self.assertNotContains(response, self.get_url("lock"))

    def test_can_unlock_other_users_manual_lock_while_in_workflow(self):
        edit_url = self.get_url("edit")

        # Submit to workflow as submitter
        self.post("submit")

        # Login as a moderator to have lock permission
        self.login(self.moderator)

        # Lock the object
        self.client.post(self.get_url("lock"))

        # Login as a superuser that has unlock permission
        self.login(self.superuser)

        response = self.client.get(edit_url)

        display_name = get_user_display_name(self.moderator)
        # Should show the lock information as locked
        self.assertContains(response, "Locked by another user", count=1)
        self.assertContains(
            response,
            f"Only {display_name} can make changes while the {self.model_name} is locked",
            count=1,
        )
        # One in the side panel, one in the message banner
        self.assertContains(response, self.get_url("unlock"), count=2)
        self.assertNotContains(response, "Locked by workflow")
        self.assertNotContains(
            response,
            f"Only reviewers can edit and approve the {self.model_name}",
        )
        self.assertNotContains(response, self.get_url("lock"))

    def test_cannot_unlock_other_users_manual_lock_while_in_workflow(self):
        edit_url = self.get_url("edit")

        # Submit to workflow as submitter
        self.post("submit")

        # Login as a superuser to have lock permission
        self.login(self.superuser)

        # Lock the object
        self.client.post(self.get_url("lock"))

        # Login as a moderator that does not have unlock permission
        # according to the workflow
        self.login(self.moderator)

        response = self.client.get(edit_url)

        display_name = get_user_display_name(self.superuser)
        # Should show the lock information as locked
        self.assertContains(response, "Locked by another user", count=1)
        self.assertContains(
            response,
            f"Only {display_name} can make changes while the {self.model_name} is locked",
            count=1,
        )
        # Has no permission to unlock
        self.assertNotContains(response, self.get_url("unlock"))
        self.assertNotContains(response, "Locked by workflow")
        self.assertNotContains(
            response,
            f"Only reviewers can edit and approve the {self.model_name}",
        )
        self.assertNotContains(response, self.get_url("lock"))

    def test_workflow_action_menu_items(self):
        edit_url = self.get_url("edit")

        # Initial view as the submitter, should only see save and submit buttons
        response = self.client.get(edit_url)
        self.assertContains(response, "Save draft")
        self.assertContains(response, "Submit to test_workflow")
        self.assertNotContains(response, "Cancel workflow")
        self.assertNotContains(response, "Restart workflow")
        self.assertNotContains(response, "Approve")
        self.assertNotContains(response, "Request changes")
        self.assertNotContains(
            response,
            '<button type="submit" class="button action-save" disabled>',
        )

        # submit for approval
        self.post("submit")

        # After submit, as a submitter, should only see cancel and locked buttons
        response = self.client.get(edit_url)
        self.assertNotContains(response, "Save draft")
        self.assertNotContains(response, "Submit to test_workflow")
        self.assertContains(response, "Cancel workflow")
        self.assertNotContains(response, "Restart workflow")
        self.assertNotContains(response, "Approve")
        self.assertNotContains(response, "Request changes")
        self.assertContains(
            response,
            '<button type="submit" class="button action-save" disabled>',
        )

        # After submit, as a moderator, should only see save, approve, and reject buttons
        self.login(self.moderator)
        response = self.client.get(edit_url)
        self.assertContains(response, "Save draft")
        self.assertNotContains(response, "Submit to test_workflow")
        self.assertNotContains(response, "Cancel workflow")
        self.assertNotContains(response, "Restart workflow")
        self.assertContains(response, "Approve")
        self.assertContains(response, "Request changes")
        self.assertNotContains(
            response,
            '<button type="submit" class="button action-save" disabled>',
        )

        self.reject()

        # After reject, as a submitter, should only see save, cancel, and restart buttons
        self.login(self.submitter)
        response = self.client.get(edit_url)
        self.assertContains(response, "Save draft")
        self.assertNotContains(response, "Submit to test_workflow")
        self.assertContains(response, "Cancel workflow")
        self.assertContains(response, "Restart workflow")
        self.assertNotContains(response, "Approve")
        self.assertNotContains(response, "Request changes")
        self.assertNotContains(
            response,
            '<button type="submit" class="button action-save" disabled>',
        )

        # After cancel, as a submitter, should only see save and submit buttons
        response = self.post("cancel-workflow", follow=True)
        self.assertContains(response, "Save draft")
        self.assertContains(response, "Submit to test_workflow")
        self.assertNotContains(response, "Cancel workflow")
        self.assertNotContains(response, "Restart workflow")
        self.assertNotContains(response, "Approve")
        self.assertNotContains(response, "Request changes")
        self.assertNotContains(
            response,
            '<button type="submit" class="button action-save" disabled>',
        )

    def test_workflow_action_menu_items_when_reverting(self):
        old_revision = self.object.latest_revision
        revert_url = self.get_url(
            "revisions_revert",
            args=(quote(self.object.pk), old_revision.id),
        )

        # Initial view as the submitter, should only see save button
        response = self.client.get(revert_url)
        self.assertContains(response, "Replace current draft")
        self.assertNotContains(response, "Submit to test_workflow")
        self.assertNotContains(response, "Cancel workflow")
        self.assertNotContains(response, "Restart workflow")
        self.assertNotContains(response, "Approve")
        self.assertNotContains(response, "Request changes")
        self.assertNotContains(
            response,
            '<button type="submit" class="button action-save" disabled>',
        )

        # submit for approval
        self.post("submit")

        # After submit, as a submitter, should only see locked button
        response = self.client.get(revert_url)
        self.assertNotContains(response, "Replace current draft")
        self.assertNotContains(response, "Submit to test_workflow")
        self.assertNotContains(response, "Cancel workflow")
        self.assertNotContains(response, "Restart workflow")
        self.assertNotContains(response, "Approve")
        self.assertNotContains(response, "Request changes")
        self.assertContains(
            response,
            '<button type="submit" class="button action-save" disabled>',
        )

        # After submit, as a moderator, should only see save button
        self.login(self.moderator)
        response = self.client.get(revert_url)
        self.assertContains(response, "Replace current draft")
        self.assertNotContains(response, "Submit to test_workflow")
        self.assertNotContains(response, "Cancel workflow")
        self.assertNotContains(response, "Restart workflow")
        self.assertNotContains(response, "Approve")
        self.assertNotContains(response, "Request changes")
        self.assertNotContains(
            response,
            '<button type="submit" class="button action-save" disabled>',
        )

        self.reject()

        # After reject, as a submitter, should only see save button
        self.login(self.submitter)
        response = self.client.get(revert_url)
        self.assertContains(response, "Replace current draft")
        self.assertNotContains(response, "Submit to test_workflow")
        self.assertNotContains(response, "Cancel workflow")
        self.assertNotContains(response, "Restart workflow")
        self.assertNotContains(response, "Approve")
        self.assertNotContains(response, "Request changes")
        self.assertNotContains(
            response,
            '<button type="submit" class="button action-save" disabled>',
        )

        # After cancel, as a submitter, should only see save button
        self.post("cancel-workflow")
        response = self.client.get(revert_url)
        self.assertContains(response, "Replace current draft")
        self.assertNotContains(response, "Submit to test_workflow")
        self.assertNotContains(response, "Cancel workflow")
        self.assertNotContains(response, "Restart workflow")
        self.assertNotContains(response, "Approve")
        self.assertNotContains(response, "Request changes")
        self.assertNotContains(
            response,
            '<button type="submit" class="button action-save" disabled>',
        )

    @override_settings(WAGTAILADMIN_BASE_URL="http://admin.example.com")
    def test_submit_sends_mail(self):
        self.post("submit")
        # 3 emails sent:
        # - to moderator - submitted for approval in moderation stage test_task_1
        # - to superuser - submitted for approval in moderation stage test_task_1
        # - to superuser - submitted to workflow test_workflow
        self.assertEqual(len(mail.outbox), 3)

        # the 'submitted to workflow' email should include the submitter's name
        workflow_message = None
        email_subject = (
            f'The {self.model_name} "{get_latest_str(self.object)}" has been submitted '
            'to workflow "test_workflow"'
        )
        for msg in mail.outbox:
            if msg.subject == email_subject:
                workflow_message = msg
                break

        self.assertTrue(workflow_message)
        self.assertIn(
            (
                f'The {self.model_name} "{get_latest_str(self.object)}" has been submitted '
                'for moderation to workflow "test_workflow" by submitter'
            ),
            workflow_message.body,
        )
        self.assertIn("http://admin.example.com/admin/", workflow_message.body)

    @override_settings(WAGTAILADMIN_NOTIFICATION_USE_HTML=True)
    def test_submit_sends_html_mail(self):
        self.test_submit_sends_mail()

    @override_settings(WAGTAILADMIN_BASE_URL=None)
    def test_submit_sends_mail_without_base_url(self):
        # With a missing WAGTAILADMIN_BASE_URL setting, we won't be able to construct absolute URLs
        # for the email, but we don't want it to fail outright either

        self.post("submit")
        # 3 emails sent:
        # - to moderator - submitted for approval in moderation stage test_task_1
        # - to superuser - submitted for approval in moderation stage test_task_1
        # - to superuser - submitted to workflow test_workflow
        self.assertEqual(len(mail.outbox), 3)

        # the 'submitted to workflow' email should include the submitter's name
        workflow_message = None
        email_subject = (
            f'The {self.model_name} "{get_latest_str(self.object)}" has been submitted '
            'to workflow "test_workflow"'
        )
        for msg in mail.outbox:
            if msg.subject == email_subject:
                workflow_message = msg
                break

        self.assertTrue(workflow_message)
        self.assertIn(
            (
                f'The {self.model_name} "{get_latest_str(self.object)}" has been submitted '
                'for moderation to workflow "test_workflow" by submitter'
            ),
            workflow_message.body,
        )

    @override_settings(WAGTAILADMIN_NOTIFICATION_USE_HTML=True)
    def test_submit_sends_html_mail_without_base_url(self):
        self.test_submit_sends_mail_without_base_url()

    @mock.patch.object(
        EmailMultiAlternatives, "send", side_effect=IOError("Server down")
    )
    def test_email_send_error(self, mock_fn):
        logging.disable(logging.CRITICAL)

        response = self.post("submit")
        logging.disable(logging.NOTSET)

        # An email that fails to send should return a message rather than crash the page
        self.assertEqual(response.status_code, 302)
        self.client.get(reverse("wagtailadmin_home"))

    def test_resume_rejected_workflow(self):
        # test that an existing workflow can be resumed by submitting when rejected
        self.workflow.start(self.object, user=self.submitter)
        workflow_state = self.object.current_workflow_state
        workflow_state.current_task_state.approve(user=self.superuser)
        workflow_state.refresh_from_db()
        workflow_state.current_task_state.reject(user=self.superuser)
        workflow_state.refresh_from_db()
        self.assertEqual(workflow_state.current_task_state.task.specific, self.task_2)
        self.assertEqual(workflow_state.status, WorkflowState.STATUS_NEEDS_CHANGES)

        self.post("submit")
        workflow_state.refresh_from_db()

        # check that the same workflow state's status is now in progress
        self.assertEqual(workflow_state.status, WorkflowState.STATUS_IN_PROGRESS)

        # check that the workflow remains on the rejecting task, rather than resetting
        self.assertEqual(workflow_state.current_task_state.task.specific, self.task_2)

    def test_restart_rejected_workflow(self):
        # test that an existing workflow can be restarted when rejected
        self.workflow.start(self.object, user=self.submitter)
        workflow_state = self.object.current_workflow_state
        workflow_state.current_task_state.approve(user=self.superuser)
        workflow_state.refresh_from_db()
        workflow_state.current_task_state.reject(user=self.superuser)
        workflow_state.refresh_from_db()
        self.assertEqual(workflow_state.current_task_state.task.specific, self.task_2)
        self.assertEqual(workflow_state.status, WorkflowState.STATUS_NEEDS_CHANGES)

        self.post("restart-workflow")
        workflow_state.refresh_from_db()

        # check that the same workflow state's status is now cancelled
        self.assertEqual(workflow_state.status, WorkflowState.STATUS_CANCELLED)

        # check that the new workflow has started on the first task
        new_workflow_state = self.object.current_workflow_state
        self.assertEqual(new_workflow_state.status, WorkflowState.STATUS_IN_PROGRESS)
        self.assertEqual(
            new_workflow_state.current_task_state.task.specific, self.task_1
        )

    def test_cancel_workflow(self):
        # test that an existing workflow can be cancelled after submission by the submitter
        self.workflow.start(self.object, user=self.submitter)
        workflow_state = self.object.current_workflow_state
        self.assertEqual(workflow_state.current_task_state.task.specific, self.task_1)
        self.assertEqual(workflow_state.status, WorkflowState.STATUS_IN_PROGRESS)
        response = self.post("cancel-workflow", follow=True)
        workflow_state.refresh_from_db()

        # check that the workflow state's status is now cancelled
        self.assertEqual(workflow_state.status, WorkflowState.STATUS_CANCELLED)
        self.assertEqual(
            workflow_state.current_task_state.status, TaskState.STATUS_CANCELLED
        )

        self.assertNotContains(
            response,
            '<button type="submit" class="button action-save" disabled>',
        )
        self.assertNotContains(
            response,
            f"The {self.model_name} could not be saved as it is locked",
        )
        self.assertNotContains(
            response,
            f"The {self.model_name} could not be saved due to validation errors",
        )
        self.assertNotContains(
            response,
            f"The {self.model_name} could not be saved due to errors",
        )

    def test_email_headers(self):
        # Submit
        self.post("submit")

        message = mail.outbox[0].message()
        msg_headers = set(message.items())
        headers = {("Auto-Submitted", "auto-generated")}
        self.assertTrue(
            headers.issubset(msg_headers),
            msg="Message is missing the Auto-Submitted header.",
        )

        self.assertFalse(message.is_multipart())

    @override_settings(WAGTAILADMIN_NOTIFICATION_USE_HTML=True)
    def test_html_email_headers(self):
        self.post("submit")

        message = mail.outbox[0].message()
        msg_headers = set(message.items())
        headers = {("Auto-Submitted", "auto-generated")}
        self.assertTrue(
            headers.issubset(msg_headers),
            msg="Message is missing the Auto-Submitted header.",
        )

        self.assertTrue(mail.outbox[0].message().is_multipart())


class TestSubmitSnippetToWorkflow(TestSubmitPageToWorkflow, BaseSnippetWorkflowTests):
    pass


# Do the same tests without LockableMixin
class TestSubmitSnippetToWorkflowNotLockable(TestSubmitSnippetToWorkflow):
    model = ModeratedModel

    def test_workflow_action_menu_items(self):
        edit_url = self.get_url("edit")

        # Initial view as the submitter, should only see save and submit buttons
        response = self.client.get(edit_url)
        self.assertContains(response, "Save draft")
        self.assertContains(response, "Submit to test_workflow")
        self.assertNotContains(response, "Cancel workflow")
        self.assertNotContains(response, "Restart workflow")
        self.assertNotContains(response, "Approve")
        self.assertNotContains(response, "Request changes")
        self.assertNotContains(
            response,
            '<button type="submit" class="button action-save" disabled>',
        )

        # submit for approval
        self.post("submit")

        # After submit, as a submitter, should only see save and cancel buttons
        # Save button is visible because the model is not lockable
        response = self.client.get(edit_url)
        self.assertContains(response, "Save draft")
        self.assertNotContains(response, "Submit to test_workflow")
        self.assertContains(response, "Cancel workflow")
        self.assertNotContains(response, "Restart workflow")
        self.assertNotContains(response, "Approve")
        self.assertNotContains(response, "Request changes")
        self.assertNotContains(
            response,
            '<button type="submit" class="button action-save" disabled>',
        )

        # After submit, as a moderator, should only see save, approve, and reject buttons
        self.login(self.moderator)
        response = self.client.get(edit_url)
        self.assertContains(response, "Save draft")
        self.assertNotContains(response, "Submit to test_workflow")
        self.assertNotContains(response, "Cancel workflow")
        self.assertNotContains(response, "Restart workflow")
        self.assertContains(response, "Approve")
        self.assertContains(response, "Request changes")
        self.assertNotContains(
            response,
            '<button type="submit" class="button action-save" disabled>',
        )

        self.reject()

        # After reject, as a submitter, should only see save, cancel, and restart buttons
        self.login(self.submitter)
        response = self.client.get(edit_url)
        self.assertContains(response, "Save draft")
        self.assertNotContains(response, "Submit to test_workflow")
        self.assertContains(response, "Cancel workflow")
        self.assertContains(response, "Restart workflow")
        self.assertNotContains(response, "Approve")
        self.assertNotContains(response, "Request changes")
        self.assertNotContains(
            response,
            '<button type="submit" class="button action-save" disabled>',
        )

    def test_workflow_action_menu_items_when_reverting(self):
        old_revision = self.object.latest_revision
        revert_url = self.get_url(
            "revisions_revert",
            args=(quote(self.object.pk), old_revision.id),
        )

        # Initial view as the submitter, should only see save button
        response = self.client.get(revert_url)
        self.assertContains(response, "Replace current draft")
        self.assertNotContains(response, "Submit to test_workflow")
        self.assertNotContains(response, "Cancel workflow")
        self.assertNotContains(response, "Restart workflow")
        self.assertNotContains(response, "Approve")
        self.assertNotContains(response, "Request changes")
        self.assertNotContains(
            response,
            '<button type="submit" class="button action-save" disabled>',
        )

        # submit for approval
        self.post("submit")

        # After submit, as a submitter, should only see save button
        # Save button is visible because the model is not lockable
        response = self.client.get(revert_url)
        self.assertContains(response, "Replace current draft")
        self.assertNotContains(response, "Submit to test_workflow")
        self.assertNotContains(response, "Cancel workflow")
        self.assertNotContains(response, "Restart workflow")
        self.assertNotContains(response, "Approve")
        self.assertNotContains(response, "Request changes")
        self.assertNotContains(
            response,
            '<button type="submit" class="button action-save" disabled>',
        )

        # After submit, as a moderator, should only see save button
        self.login(self.moderator)
        response = self.client.get(revert_url)
        self.assertContains(response, "Replace current draft")
        self.assertNotContains(response, "Submit to test_workflow")
        self.assertNotContains(response, "Cancel workflow")
        self.assertNotContains(response, "Restart workflow")
        self.assertNotContains(response, "Approve")
        self.assertNotContains(response, "Request changes")
        self.assertNotContains(
            response,
            '<button type="submit" class="button action-save" disabled>',
        )

        self.reject()

        # After reject, as a submitter, should only see save button
        self.login(self.submitter)
        response = self.client.get(revert_url)
        self.assertContains(response, "Replace current draft")
        self.assertNotContains(response, "Submit to test_workflow")
        self.assertNotContains(response, "Cancel workflow")
        self.assertNotContains(response, "Restart workflow")
        self.assertNotContains(response, "Approve")
        self.assertNotContains(response, "Request changes")
        self.assertNotContains(
            response,
            '<button type="submit" class="button action-save" disabled>',
        )

    def test_submit_for_approval_changes_lock_status(self):
        edit_url = self.get_url("edit")

        # submit for approval
        self.post("submit")

        # Login as a moderator
        self.login(self.moderator)

        response = self.client.get(edit_url)

        # Model is not lockable, should not show any lock information or buttons

        self.assertNotContains(response, "Unlocked")
        self.assertNotContains(response, f"Anyone can edit this {self.model_name}")
        self.assertNotContains(
            response,
            f"Reviewers can edit this {self.model_name} – lock it to prevent other reviewers from editing",
        )
        self.assertNotContains(response, "Locked by workflow")
        self.assertNotContains(
            response,
            f"Only reviewers can edit and approve the {self.model_name}",
        )
        self.assertNotContains(response, "Locked by another user")

    @skip("Model is not lockable")
    def test_can_manual_lock_while_in_workflow(self):
        pass

    @skip("Model is not lockable")
    def test_can_unlock_manual_lock_while_in_workflow(self):
        pass

    @skip("Model is not lockable")
    def test_can_unlock_other_users_manual_lock_while_in_workflow(self):
        pass

    @skip("Model is not lockable")
    def test_cannot_unlock_other_users_manual_lock_while_in_workflow(self):
        pass


@freeze_time("2020-03-31 12:00:00")
class TestApproveRejectPageWorkflow(BasePageWorkflowTests):
    published_signal = page_published
    title_field = "title"

    def setUp(self):
        super().setUp()
        self.submitter.first_name = "Sebastian"
        self.submitter.last_name = "Mitter"
        self.submitter.save()
        self.post("submit")
        self.login(user=self.moderator)

    def setup_workflow_and_tasks(self):
        self.workflow = Workflow.objects.create(name="test_workflow")
        self.task_1 = GroupApprovalTask.objects.create(name="test_task_1")
        self.task_1.groups.set(Group.objects.filter(name="Moderators"))
        WorkflowTask.objects.create(
            workflow=self.workflow, task=self.task_1, sort_order=1
        )

    @override_settings(WAGTAIL_FINISH_WORKFLOW_ACTION="")
    def test_approve_task_and_workflow(self):
        """
        This posts to the approve task view and checks that the object was approved and published
        """
        # Unset WAGTAIL_FINISH_WORKFLOW_ACTION - default action should be to publish
        del settings.WAGTAIL_FINISH_WORKFLOW_ACTION
        # Connect a mock signal handler to published signal
        mock_handler = mock.MagicMock()
        self.published_signal.connect(mock_handler)

        try:
            # Post
            response = self.approve({"comment": "my comment"})
            self.assertRedirects(response, self.get_url("edit"))

            # Check that the workflow was approved

            workflow_state = WorkflowState.objects.for_instance(self.object).get(
                requested_by=self.submitter
            )

            self.assertEqual(workflow_state.status, workflow_state.STATUS_APPROVED)

            # Check that the task was approved

            task_state = workflow_state.current_task_state

            self.assertEqual(task_state.status, task_state.STATUS_APPROVED)

            # Check that the comment was added to the task state correctly

            self.assertEqual(task_state.comment, "my comment")

            self.object.refresh_from_db()
            # Object must be live
            self.assertTrue(
                self.object.live, msg="Approving moderation failed to set live=True"
            )
            # Object should now have no unpublished changes
            self.assertFalse(
                self.object.has_unpublished_changes,
                msg="Approving moderation failed to set has_unpublished_changes=False",
            )

            # Check that the published signal was fired
            self.assertEqual(mock_handler.call_count, 1)
            mock_call = mock_handler.mock_calls[0][2]

            self.assertEqual(mock_call["sender"], self.object_class)
            self.assertEqual(mock_call["instance"], self.object)
            self.assertIsInstance(mock_call["instance"], self.object_class)
        finally:
            self.published_signal.disconnect(mock_handler)

    def test_approve_task_and_workflow_with_ajax(self):
        """
        This posts to the approve task view and checks that the object was approved and published
        """
        # Post
        response = self.approve(
            {"comment": "my comment"}, HTTP_X_REQUESTED_WITH="XMLHttpRequest"
        )
        self.assertJSONEqual(
            response.content.decode(),
            {"step": "success", "redirect": self.get_url("edit")},
        )

        # Check that the workflow was approved

        workflow_state = WorkflowState.objects.for_instance(self.object).get(
            requested_by=self.submitter
        )

        self.assertEqual(workflow_state.status, workflow_state.STATUS_APPROVED)

        # Check that the task was approved

        task_state = workflow_state.current_task_state

        self.assertEqual(task_state.status, task_state.STATUS_APPROVED)

        # Check that the comment was added to the task state correctly

        self.assertEqual(task_state.comment, "my comment")

        self.object.refresh_from_db()
        # Object must be live
        self.assertTrue(
            self.object.live, msg="Approving moderation failed to set live=True"
        )
        # Object should now have no unpublished changes
        self.assertFalse(
            self.object.has_unpublished_changes,
            msg="Approving moderation failed to set has_unpublished_changes=False",
        )

    def test_workflow_dashboard_panel(self):
        response = self.client.get(reverse("wagtailadmin_home"))
        self.assertContains(response, "Awaiting your review")
        soup = self.get_soup(response.content)
        # check that the workflow-action script is present with the correct data-activate attribute
        workflow_action_js = versioned_static("wagtailadmin/js/workflow-action.js")
        scripts = soup.select(f"script[src='{workflow_action_js}']")
        self.assertEqual(len(scripts), 1)
        script = scripts[0]
        self.assertIsNotNone(script)
        self.assertEqual(script.get("data-activate"), "dashboard")
        # Should no longer contain inline JS for activating the workflow actions
        self.assertNotContains(response, "ActivateWorkflowActionsForDashboard")

    def test_workflow_action_script_included(self):
        response = self.client.get(self.get_url("edit"))
        self.assertEqual(response.status_code, 200)
        soup = self.get_soup(response.content)
        # check that the workflow-action script is present with the correct
        # data-activate and data-confirm-cancellation-url attributes
        workflow_action_js = versioned_static("wagtailadmin/js/workflow-action.js")
        scripts = soup.select(f"script[src='{workflow_action_js}']")
        self.assertEqual(len(scripts), 1)
        script = scripts[0]
        self.assertIsNotNone(script)
        self.assertEqual(script.get("data-activate"), "editor")
        self.assertEqual(
            script.get("data-confirm-cancellation-url"),
            self.get_url("confirm_workflow_cancellation"),
        )
        # Should no longer contain inline JS for activating the workflow actions
        self.assertNotContains(response, "ActivateWorkflowActionsForEditView")

    @override_settings(WAGTAIL_WORKFLOW_CANCEL_ON_PUBLISH=False)
    def test_workflow_action_script_included_without_cancel_confirmation(self):
        response = self.client.get(self.get_url("edit"))
        self.assertEqual(response.status_code, 200)
        soup = self.get_soup(response.content)
        # check that the workflow-action script is present with the correct data-activate attribute
        workflow_action_js = versioned_static("wagtailadmin/js/workflow-action.js")
        scripts = soup.select(f"script[src='{workflow_action_js}']")
        self.assertEqual(len(scripts), 1)
        script = scripts[0]
        self.assertIsNotNone(script)
        self.assertEqual(script.get("data-activate"), "editor")
        # data-confirm-cancellation-url attribute should not be present as
        # WAGTAIL_WORKFLOW_CANCEL_ON_PUBLISH is set to False
        self.assertIsNone(script.get("data-confirm-cancellation-url"))
        # Should no longer contain inline JS for activating the workflow actions
        self.assertNotContains(response, "ActivateWorkflowActionsForEditView")

    def test_workflow_action_get(self):
        """
        This tests that a GET request to the workflow action view (for the approve action) returns a modal with a form for extra data entry:
        adding a comment
        """
        response = self.client.get(
            self.get_url(
                "workflow_action",
                args=(
                    quote(self.object.pk),
                    "approve",
                    self.object.current_workflow_task_state.id,
                ),
            ),
        )
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(
            response, "wagtailadmin/shared/workflow_action_modal.html"
        )
        html = json.loads(response.content)["html"]
        self.assertTagInHTML(
            '<form action="'
            + self.get_url(
                "workflow_action",
                args=(
                    quote(self.object.pk),
                    "approve",
                    self.object.current_workflow_task_state.id,
                ),
            )
            + '" method="POST" novalidate>',
            html,
        )
        self.assertIn("Comment", html)

    def test_workflow_action_get_custom_template(self):
        """
        https://github.com/wagtail/wagtail/issues/12222
        Custom tasks can override Task.get_template_for_action() to use a custom
        template for the workflow action modal.
        """
        # Add a custom task to the workflow
        custom_task = UserApprovalTask.objects.create(
            name="user_approval_1",
            user=self.moderator,
        )
        WorkflowTask.objects.create(
            workflow=self.workflow,
            task=custom_task,
            sort_order=2,
        )
        self.approve()  # Approve the GroupApprovalTask

        # Refresh from DB
        self.object = self.object_class.objects.get(pk=self.object.pk)

        response = self.client.get(
            self.get_url(
                "workflow_action",
                args=(
                    quote(self.object.pk),
                    "approve",
                    self.object.current_workflow_task_state.id,
                ),
            ),
        )
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "tests/workflows/approve_with_style.html")
        self.assertTemplateNotUsed(
            response, "wagtailadmin/shared/workflow_action_modal.html"
        )
        html = json.loads(response.content)["html"]
        soup = self.get_soup(html)
        form = soup.select_one("form")
        self.assertIsNotNone(form)
        self.assertEqual(
            form["action"],
            self.get_url(
                "workflow_action",
                args=(
                    quote(self.object.pk),
                    "approve",
                    self.object.current_workflow_task_state.id,
                ),
            ),
        )
        submit = form.select_one("button[type=submit]")
        self.assertIsNotNone(submit)
        self.assertEqual(submit.text.strip(), "Ship it!")
        self.assertNotIn("Comment", html)

    def test_workflow_action_view_bad_id(self):
        """
        This tests that the workflow action view handles invalid object ids correctly
        """
        # Post
        response = self.client.post(
            self.get_url(
                "workflow_action",
                args=(
                    127777777777,
                    "approve",
                    self.object.current_workflow_task_state.id,
                ),
            ),
        )

        # Check that the user received a 404 response
        self.assertEqual(response.status_code, 404)

    def test_workflow_action_view_not_in_group(self):
        """
        This tests that the workflow action view for a GroupApprovalTask won't allow approval from a user not in the
        specified group/a superuser
        """
        # Remove privileges from user
        self.login(user=self.submitter)

        # Post
        response = self.approve()
        # Check that the user received a permission denied response
        self.assertRedirects(response, "/admin/")

    def test_workflow_action_view_not_in_moderation(self):
        """
        This tests that the workflow action view won't allow an action
        for an object that's not in moderation. For example, the submitter
        cancelled the workflow before the moderator could approve it.
        """
        self.login(user=self.submitter)

        # Keep reference to the current workflow state so we can get the URL
        current_workflow_task_state = self.object.current_workflow_task_state

        # Cancel the workflow
        response = self.client.post(
            self.get_url("edit"),
            {"action-cancel-workflow": "True"},
        )

        self.login(self.moderator)

        # Try to approve
        response = self.client.post(
            self.get_url(
                "workflow_action",
                args=(
                    quote(self.object.pk),
                    "approve",
                    current_workflow_task_state.id,
                ),
            ),
            follow=True,
        )
        # Check that the user is redirected to the edit page
        # and received an error message
        self.assertRedirects(response, self.get_url("edit"))
        self.assertContains(
            response,
            f"The {self.model_name} &#x27;{get_latest_str(self.object)}&#x27; "
            "is not currently awaiting moderation.",
        )

    def test_edit_view_workflow_cancellation_not_in_group(self):
        """
        This tests that the object edit view for a GroupApprovalTask, locked to a user not in the
        specified group/a superuser, still allows the submitter to cancel workflows
        """
        self.login(user=self.submitter)

        # Post
        response = self.client.post(
            self.get_url("edit"),
            {"action-cancel-workflow": "True"},
        )

        # Check that the user received a 200 response
        self.assertEqual(response.status_code, 200)

        # Check that the workflow state was marked as cancelled
        workflow_state = WorkflowState.objects.for_instance(self.object).get(
            requested_by=self.submitter
        )
        self.assertEqual(workflow_state.status, WorkflowState.STATUS_CANCELLED)

    def test_reject_task_and_workflow(self):
        """
        This posts to the reject task view and checks that the object was rejected and not published
        """
        # Post
        self.reject()

        # Check that the workflow was marked as needing changes

        workflow_state = WorkflowState.objects.for_instance(self.object).get(
            requested_by=self.submitter
        )

        self.assertEqual(workflow_state.status, workflow_state.STATUS_NEEDS_CHANGES)

        # Check that the task was rejected

        task_state = workflow_state.current_task_state

        self.assertEqual(task_state.status, task_state.STATUS_REJECTED)

        self.object.refresh_from_db()
        # Object must not be live
        self.assertFalse(self.object.live)

    def test_reject_task_and_workflow_without_form(self):
        """
        This posts to the reject task view for a task without a form and checks that the object can still be rejected and not published
        """
        # Post
        with mock.patch("wagtail.models.Task.get_form_for_action") as get_form:
            get_form.return_value = None
            response = self.reject()

        self.assertRedirects(response, self.get_url("edit"))

        # Check that the workflow was marked as needing changes

        workflow_state = WorkflowState.objects.for_instance(self.object).get(
            requested_by=self.submitter
        )

        self.assertEqual(workflow_state.status, workflow_state.STATUS_NEEDS_CHANGES)

        # Check that the task was rejected

        task_state = workflow_state.current_task_state

        self.assertEqual(task_state.status, task_state.STATUS_REJECTED)

        self.object.refresh_from_db()
        # Object must not be live
        self.assertFalse(self.object.live)

    def test_reject_task_and_workflow_with_invalid_form_ajax(self):
        """
        This posts to the reject task view with invalid form data and checks that the object was not rejected and not published
        """
        # Post
        with mock.patch("wagtail.forms.TaskStateCommentForm.is_valid") as is_valid:
            is_valid.return_value = False
            response = self.reject(HTTP_X_REQUESTED_WITH="XMLHttpRequest")

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(
            response, "wagtailadmin/shared/workflow_action_modal.html"
        )

        # Check that the workflow was not marked as needing changes

        workflow_state = WorkflowState.objects.for_instance(self.object).get(
            requested_by=self.submitter
        )

        self.assertNotEqual(workflow_state.status, workflow_state.STATUS_NEEDS_CHANGES)

        # Check that the task was not rejected

        task_state = workflow_state.current_task_state

        self.assertNotEqual(task_state.status, task_state.STATUS_REJECTED)

        self.object.refresh_from_db()
        # Object must not be live
        self.assertFalse(self.object.live)

    def test_workflow_action_view_rejection_not_in_group(self):
        """
        This tests that the workflow action view for a GroupApprovalTask won't allow rejection from a user not in the
        specified group/a superuser
        """
        # Remove privileges from user
        self.login(user=self.submitter)

        # Post
        response = self.reject()

        # Check that the user received a permission denied response
        self.assertRedirects(response, "/admin/")

    def test_collect_workflow_action_data_get(self):
        """
        This tests that a GET request to the collect_workflow_action_data view (for the approve action) returns a modal with a form for extra data entry:
        adding a comment
        """
        response = self.client.get(
            self.get_url(
                "collect_workflow_action_data",
                args=(
                    quote(self.object.pk),
                    "approve",
                    self.object.current_workflow_task_state.id,
                ),
            )
        )
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(
            response, "wagtailadmin/shared/workflow_action_modal.html"
        )
        self.assertTemplateUsed(response, "wagtailadmin/shared/non_field_errors.html")
        html = json.loads(response.content)["html"]
        self.assertTagInHTML(
            '<form action="'
            + self.get_url(
                "collect_workflow_action_data",
                args=(
                    quote(self.object.pk),
                    "approve",
                    self.object.current_workflow_task_state.id,
                ),
            )
            + '" method="POST" novalidate>',
            html,
        )
        self.assertIn("Comment", html)

    def test_collect_workflow_action_data_get_custom_template(self):
        """
        https://github.com/wagtail/wagtail/issues/12222
        Custom tasks can override Task.get_template_for_action() to use a custom
        template for the workflow action modal.
        """
        # Add a custom task to the workflow
        custom_task = UserApprovalTask.objects.create(
            name="user_approval_1",
            user=self.moderator,
        )
        WorkflowTask.objects.create(
            workflow=self.workflow,
            task=custom_task,
            sort_order=2,
        )
        self.approve()  # Approve the GroupApprovalTask

        # Refresh from DB
        self.object = self.object_class.objects.get(pk=self.object.pk)

        response = self.client.get(
            self.get_url(
                "collect_workflow_action_data",
                args=(
                    quote(self.object.pk),
                    "approve",
                    self.object.current_workflow_task_state.id,
                ),
            ),
        )
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "tests/workflows/approve_with_style.html")
        self.assertTemplateNotUsed(
            response, "wagtailadmin/shared/workflow_action_modal.html"
        )
        html = json.loads(response.content)["html"]
        soup = self.get_soup(html)
        form = soup.select_one("form")
        self.assertIsNotNone(form)
        self.assertEqual(
            form["action"],
            self.get_url(
                "collect_workflow_action_data",
                args=(
                    quote(self.object.pk),
                    "approve",
                    self.object.current_workflow_task_state.id,
                ),
            ),
        )
        submit = form.select_one("button[type=submit]")
        self.assertIsNotNone(submit)
        self.assertEqual(submit.text.strip(), "Ship it!")
        self.assertNotIn("Comment", html)

    def test_collect_workflow_action_data_post(self):
        """
        This tests that a POST request to the collect_workflow_action_data view (for the approve action) returns a modal response with the validated data
        """
        response = self.client.post(
            self.get_url(
                "collect_workflow_action_data",
                args=(
                    quote(self.object.pk),
                    "approve",
                    self.object.current_workflow_task_state.id,
                ),
            ),
            {"comment": "This is my comment"},
        )
        self.assertEqual(response.status_code, 200)
        response_json = json.loads(response.content)
        self.assertEqual(response_json["step"], "success")
        self.assertEqual(
            response_json["cleaned_data"], {"comment": "This is my comment"}
        )

    def test_collect_workflow_action_data_post_invalid_form(self):
        """
        This tests that a POST request to the collect_workflow_action_data view with an invalid form data returns a redirect
        """
        with mock.patch("wagtail.forms.TaskStateCommentForm.is_valid") as is_valid:
            is_valid.return_value = False
            response = self.client.post(
                self.get_url(
                    "collect_workflow_action_data",
                    args=(
                        quote(self.object.pk),
                        "approve",
                        self.object.current_workflow_task_state.id,
                    ),
                ),
                {"comment": "This is my comment"},
            )
        self.assertRedirects(response, self.get_url("edit"))

    def test_collect_workflow_action_data_post_invalid_form_ajax(self):
        """
        This tests that a POST request to the collect_workflow_action_data view with an invalid form data returns the form with errors
        """
        with mock.patch("wagtail.forms.TaskStateCommentForm.is_valid") as is_valid:
            is_valid.return_value = False
            response = self.client.post(
                self.get_url(
                    "collect_workflow_action_data",
                    args=(
                        quote(self.object.pk),
                        "approve",
                        self.object.current_workflow_task_state.id,
                    ),
                ),
                {"comment": "This is my comment"},
                HTTP_X_REQUESTED_WITH="XMLHttpRequest",
            )
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(
            response, "wagtailadmin/shared/workflow_action_modal.html"
        )

    def test_workflow_action_via_edit_view(self):
        """
        Posting to the 'edit' view with 'action-workflow-action' set should perform the given workflow action in addition to updating object content
        """
        # Post
        self.post(
            "workflow-action",
            {
                self.title_field: "This title was edited while approving",
                "workflow-action-name": "approve",
                "workflow-action-extra-data": '{"comment": "my comment"}',
            },
        )

        # Check that the workflow was approved

        workflow_state = WorkflowState.objects.for_instance(self.object).get(
            requested_by=self.submitter
        )

        self.assertEqual(workflow_state.status, workflow_state.STATUS_APPROVED)

        # Check that the task was approved

        task_state = workflow_state.current_task_state

        self.assertEqual(task_state.status, task_state.STATUS_APPROVED)

        # Check that the comment was added to the task state correctly

        self.assertEqual(task_state.comment, "my comment")

        # Check that object edits made at the same time as the action have been saved
        self.object.refresh_from_db()
        self.assertEqual(
            getattr(self.object.get_latest_revision_as_object(), self.title_field),
            "This title was edited while approving",
        )


class TestApproveRejectSnippetWorkflow(
    TestApproveRejectPageWorkflow, BaseSnippetWorkflowTests
):
    published_signal = published
    title_field = "text"


# Do the same tests without LockableMixin
class TestApproveRejectSnippetWorkflowNotLockable(TestApproveRejectSnippetWorkflow):
    model = ModeratedModel


@freeze_time("2020-03-31 12:00:00")
class TestPageWorkflowReport(BasePageWorkflowTests):
    export_formats = ["xlsx", "csv"]
    workflow_url_name = "wagtailadmin_reports:workflow"
    workflow_tasks_url_name = "wagtailadmin_reports:workflow_tasks"
    header_buttons_parent_selector = "#w-slim-header-buttons"
    drilldown_selector = ".w-drilldown"
    extra_params = ""

    def setUp(self):
        super().setUp()
        self.submitter.first_name = "Sebastian"
        self.submitter.last_name = "Mitter"
        self.submitter.save()
        self.post("submit", follow=True)
        self.login(user=self.moderator)

    def assertBreadcrumbs(self, breadcrumbs, html):
        self.assertBreadcrumbsItemsRendered(breadcrumbs, html)

    def assertPageTitle(self, soup, title):
        self.assertEqual(soup.select_one("title").text.strip(), title)

    def get(self, url, params=None):
        return self.client.get(url, params)

    def setup_workflow_and_tasks(self):
        self.workflow = Workflow.objects.create(name="test_workflow")
        self.task_1 = GroupApprovalTask.objects.create(name="test_task_1")
        self.task_1.groups.set(Group.objects.filter(name="Moderators"))
        WorkflowTask.objects.create(
            workflow=self.workflow, task=self.task_1, sort_order=1
        )

    def get_file_content(self, response, format):
        if format == "xlsx":
            workbook = load_workbook(io.BytesIO(response.getvalue()))
            worksheet = workbook.active
            return "".join(
                str(worksheet.cell(row=i, column=j).value)
                for j in range(1, worksheet.max_column + 1)
                for i in range(1, worksheet.max_row + 1)
            )
        return response.getvalue().decode()

    def test_workflow_report(self):
        response = self.get(reverse(self.workflow_url_name))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Hello world!")
        self.assertContains(response, "test_workflow")
        self.assertContains(response, "Sebastian Mitter")
        self.assertContains(response, "March 31, 2020")
        self.assertBreadcrumbs(
            [{"url": "", "label": "Workflows"}],
            response.content,
        )
        soup = self.get_soup(response.content)
        by_task_link = soup.select_one(
            f"{self.header_buttons_parent_selector} .w-header-button"
        )
        self.assertIsNotNone(by_task_link)
        self.assertEqual(
            by_task_link.get("href"),
            reverse("wagtailadmin_reports:workflow_tasks"),
        )
        self.assertEqual(list(by_task_link.children)[-1].strip(), "By task")
        self.assertIsNone(soup.select_one(".w-active-filters"))
        self.assertPageTitle(soup, "Workflows - Wagtail")

        response = self.get(reverse(self.workflow_tasks_url_name))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Hello world!")
        self.assertBreadcrumbs(
            [{"url": "", "label": "Workflow tasks"}],
            response.content,
        )
        soup = self.get_soup(response.content)
        by_task_link = soup.select_one(
            f"{self.header_buttons_parent_selector} .w-header-button"
        )
        self.assertIsNotNone(by_task_link)
        self.assertEqual(
            by_task_link.get("href"),
            reverse("wagtailadmin_reports:workflow"),
        )
        self.assertEqual(list(by_task_link.children)[-1].strip(), "By workflow")
        self.assertIsNone(soup.select_one(".w-active-filters"))
        self.assertPageTitle(soup, "Workflow tasks - Wagtail")

    def test_workflow_report_filtered(self):
        # the moderator can review the task, so the workflow state should show up even when reports are filtered by reviewable
        response = self.get(reverse(self.workflow_url_name), {"reviewable": "true"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Hello world!")
        self.assertContains(response, "test_workflow")
        self.assertContains(response, "Sebastian Mitter")
        self.assertContains(response, "March 31, 2020")

        # Should render the export buttons inside the header "more" dropdown
        # with the filtered URL
        soup = self.get_soup(response.content)
        links = soup.select(f"{self.header_buttons_parent_selector} .w-dropdown a")
        unfiltered_url = reverse(self.workflow_url_name)
        filtered_url = f"{unfiltered_url}?reviewable=true{self.extra_params}"
        self.assertEqual(len(links), 2)
        self.assertEqual(
            [link.get("href") for link in links],
            [f"{filtered_url}&export=xlsx", f"{filtered_url}&export=csv"],
        )

        # Should render the active filter pill
        active_filter = soup.select_one(".w-active-filters .w-pill__content")
        clear_button = soup.select_one(".w-active-filters .w-pill__remove")
        self.assertIsNotNone(active_filter)
        self.assertIsNotNone(clear_button)
        self.assertNotIn("reviewable", clear_button.attrs.get("data-w-swap-src-value"))
        self.assertEqual(clear_button.attrs.get("data-w-swap-reflect-value"), "true")

        # Should render the filter inside the drilldown component
        inputs = soup.select(
            f"{self.drilldown_selector} input[name='reviewable'][type='radio']"
        )
        self.assertEqual(len(inputs), 2)
        self.assertEqual(inputs[0].get("value"), "")
        self.assertIsNone(inputs[0].get("checked"))
        self.assertEqual(inputs[1].get("value"), "true")
        self.assertEqual(inputs[1].get("checked"), "")

        response = self.get(
            reverse(self.workflow_tasks_url_name),
            {"reviewable": "true"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Hello world!")

        # Should render the export buttons inside the header "more" dropdown
        # with the filtered URL
        soup = self.get_soup(response.content)
        links = soup.select(f"{self.header_buttons_parent_selector} .w-dropdown a")
        unfiltered_url = reverse(self.workflow_tasks_url_name)
        filtered_url = f"{unfiltered_url}?reviewable=true{self.extra_params}"
        self.assertEqual(len(links), 2)
        self.assertEqual(
            [link.get("href") for link in links],
            [f"{filtered_url}&export=xlsx", f"{filtered_url}&export=csv"],
        )

        # Should render the active filter pill
        active_filter = soup.select_one(".w-active-filters .w-pill__content")
        clear_button = soup.select_one(".w-active-filters .w-pill__remove")
        self.assertIsNotNone(active_filter)
        self.assertIsNotNone(clear_button)
        self.assertNotIn("reviewable", clear_button.attrs.get("data-w-swap-src-value"))
        self.assertEqual(clear_button.attrs.get("data-w-swap-reflect-value"), "true")

        # Should render the filter inside the drilldown component
        inputs = soup.select(
            f"{self.drilldown_selector} input[name='reviewable'][type='radio']"
        )
        self.assertEqual(len(inputs), 2)
        self.assertEqual(inputs[0].get("value"), "")
        self.assertIsNone(inputs[0].get("checked"))
        self.assertEqual(inputs[1].get("value"), "true")
        self.assertEqual(inputs[1].get("checked"), "")

        # the submitter cannot review the task, so the workflow state shouldn't show up when reports are filtered by reviewable
        self.login(self.submitter)
        response = self.get(reverse(self.workflow_url_name), {"reviewable": "true"})
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Hello world!")
        self.assertNotContains(response, "Sebastian Mitter")
        self.assertNotContains(response, "March 31, 2020")

        response = self.get(
            reverse(self.workflow_tasks_url_name),
            {"reviewable": "true"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Hello world!")

    def test_workflow_report_export(self):
        for export_format in self.export_formats:
            with self.subTest(export_format=export_format):
                response = self.get(
                    reverse(self.workflow_url_name),
                    {"export": export_format},
                )
                content = self.get_file_content(response, export_format)
                self.assertEqual(response.status_code, 200)
                self.assertIn("Hello world!", content)
                self.assertIn("test_workflow", content)
                self.assertIn("submitter", content)
                self.assertIn("2020-03-31", content)

                response = self.get(
                    reverse(self.workflow_tasks_url_name),
                    {"export": export_format},
                )
                content = self.get_file_content(response, export_format)
                self.assertEqual(response.status_code, 200)
                self.assertIn("Hello world!", content)

    def test_workflow_report_filtered_export(self):
        for export_format in self.export_formats:
            with self.subTest(export_format=export_format):
                # the moderator can review the task, so the workflow state should show up even when reports are filtered by reviewable
                self.login(self.moderator)
                response = self.get(
                    reverse(self.workflow_url_name),
                    {"reviewable": "true", "export": export_format},
                )
                content = self.get_file_content(response, export_format)
                self.assertEqual(response.status_code, 200)
                self.assertIn("Hello world!", content)
                self.assertIn("test_workflow", content)
                self.assertIn("submitter", content)
                self.assertIn("2020-03-31", content)

                response = self.get(
                    reverse(self.workflow_tasks_url_name),
                    {"reviewable": "true", "export": export_format},
                )
                content = self.get_file_content(response, export_format)
                self.assertEqual(response.status_code, 200)
                self.assertIn("Hello world!", content)

                # the submitter cannot review the task, so the workflow state shouldn't show up when reports are filtered by reviewable
                self.login(self.submitter)
                response = self.get(
                    reverse(self.workflow_url_name),
                    {"reviewable": "true", "export": export_format},
                )
                content = self.get_file_content(response, export_format)
                self.assertEqual(response.status_code, 200)
                self.assertNotIn("Hello world!", content)
                self.assertNotIn("submitter", content)
                self.assertNotIn("2020-03-31", content)

                response = self.get(
                    reverse(self.workflow_tasks_url_name),
                    {"reviewable": "true", "export": export_format},
                )
                content = self.get_file_content(response, export_format)
                self.assertEqual(response.status_code, 200)
                self.assertNotIn("Hello world!", content)

    def test_workflow_report_deleted(self):
        self.object.delete()
        response = self.get(reverse(self.workflow_url_name))
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Hello world!")
        # test_workflow is only rendered in the filter, not the results
        self.assertContains(response, "test_workflow", count=1)
        self.assertNotContains(response, "Sebastian Mitter")
        self.assertNotContains(response, "March 31, 2020")

        response = self.get(reverse(self.workflow_tasks_url_name))
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Hello world!")


class TestPageWorkflowReportResults(TestPageWorkflowReport):
    workflow_url_name = "wagtailadmin_reports:workflow_results"
    workflow_tasks_url_name = "wagtailadmin_reports:workflow_tasks_results"
    header_buttons_parent_selector = (
        '[data-controller="w-teleport"]'
        '[data-w-teleport-target-value="#w-slim-header-buttons"]'
    )
    drilldown_selector = (
        '[data-controller="w-teleport"]'
        '[data-w-teleport-target-value="#filters-drilldown"]'
    )
    extra_params = "&_w_filter_fragment=true"

    def assertBreadcrumbs(self, breadcrumbs, html):
        self.assertBreadcrumbsNotRendered(html)

    def assertPageTitle(self, soup, title):
        self.assertIsNone(soup.select_one("title"))

    def get(self, url, params=None):
        params = params or {}
        params["_w_filter_fragment"] = "true"
        return super().get(url, params)


class TestSnippetWorkflowReport(TestPageWorkflowReport, BaseSnippetWorkflowTests):
    pass


class TestSnippetWorkflowReportResults(
    TestPageWorkflowReportResults, BaseSnippetWorkflowTests
):
    pass


class TestNonLockableSnippetWorkflowReport(
    TestPageWorkflowReport, BaseSnippetWorkflowTests
):
    # This model does not use LockableMixin, and it also does not have a
    # GenericRelation to WorkflowState and Revision, but it should not break
    # the report page.
    # See https://github.com/wagtail/wagtail/issues/11300 for more details.
    model = ModeratedModel


class TestNonLockableSnippetWorkflowReportResults(
    TestPageWorkflowReportResults, BaseSnippetWorkflowTests
):
    model = ModeratedModel


class TestPageNotificationPreferences(BasePageWorkflowTests):
    def setUp(self):
        super().setUp()
        self.moderator2 = self.create_user(
            username="moderator2",
            email="moderator2@email.com",
            password="password",
        )
        moderators = Group.objects.get(name="Moderators")
        moderators.user_set.add(self.moderator2)

        self.superuser_profile = UserProfile.get_for_user(self.superuser)
        self.moderator2_profile = UserProfile.get_for_user(self.moderator2)
        self.submitter_profile = UserProfile.get_for_user(self.submitter)

    def setup_workflow_and_tasks(self):
        self.workflow = Workflow.objects.create(name="test_workflow")
        self.task_1 = GroupApprovalTask.objects.create(name="test_task_1")
        self.task_1.groups.set(Group.objects.filter(name="Moderators"))
        WorkflowTask.objects.create(
            workflow=self.workflow, task=self.task_1, sort_order=1
        )

    def test_vanilla_profile(self):
        # Check that the vanilla profile has rejected notifications on
        self.assertIs(self.submitter_profile.rejected_notifications, True)

        # Check that the vanilla profile has approved notifications on
        self.assertIs(self.submitter_profile.approved_notifications, True)

    @override_settings(WAGTAILADMIN_NOTIFICATION_INCLUDE_SUPERUSERS=True)
    def test_submitted_email_notifications_sent(self):
        """Test that 'submitted' notifications for WorkflowState and TaskState are both sent correctly"""
        self.login(self.submitter)
        self.post("submit")

        self.assertEqual(len(mail.outbox), 4)

        task_submission_emails = [
            email for email in mail.outbox if "task" in email.subject
        ]
        task_submission_emailed_addresses = [
            address for email in task_submission_emails for address in email.to
        ]
        workflow_submission_emails = [
            email for email in mail.outbox if "workflow" in email.subject
        ]
        workflow_submission_emailed_addresses = [
            address for email in workflow_submission_emails for address in email.to
        ]
        workflow_submission_email_body = workflow_submission_emails[0].body

        self.assertEqual(len(task_submission_emails), 3)
        # the moderator is in the Group assigned to the GroupApproval task, so should get an email
        self.assertIn(self.moderator.email, task_submission_emailed_addresses)
        self.assertIn(self.moderator2.email, task_submission_emailed_addresses)
        # with `WAGTAILADMIN_NOTIFICATION_INCLUDE_SUPERUSERS`, the superuser should get a task email
        self.assertIn(self.superuser.email, task_submission_emailed_addresses)
        # the submitter triggered this workflow update, so should not get an email
        self.assertNotIn(self.submitter.email, task_submission_emailed_addresses)

        self.assertEqual(len(workflow_submission_emails), 1)
        # the moderator should not get a workflow email
        self.assertNotIn(self.moderator.email, workflow_submission_emailed_addresses)
        self.assertNotIn(self.moderator2.email, workflow_submission_emailed_addresses)
        # with `WAGTAILADMIN_NOTIFICATION_INCLUDE_SUPERUSERS`, the superuser should get a workflow email
        self.assertIn(self.superuser.email, workflow_submission_emailed_addresses)
        # as the submitter was the triggering user, the submitter should not get an email notification
        self.assertNotIn(self.submitter.email, workflow_submission_emailed_addresses)
        # check that the email contains the ability to show the notifications tab in the admin account preferences
        account_notifications_url = (
            get_admin_base_url()
            + reverse("wagtailadmin_account")
            + "#tab-notifications"
        )
        self.assertIn(
            "Edit your notification preferences here: %s" % account_notifications_url,
            workflow_submission_email_body,
        )

    @override_settings(WAGTAILADMIN_NOTIFICATION_INCLUDE_SUPERUSERS=False)
    def test_submitted_email_notifications_superuser_settings(self):
        """Test that 'submitted' notifications for WorkflowState and TaskState are not sent to superusers if
        `WAGTAILADMIN_NOTIFICATION_INCLUDE_SUPERUSERS=False`"""
        self.login(self.submitter)
        self.post("submit")

        task_submission_emails = [
            email for email in mail.outbox if "task" in email.subject
        ]
        task_submission_emailed_addresses = [
            address for email in task_submission_emails for address in email.to
        ]
        workflow_submission_emails = [
            email for email in mail.outbox if "workflow" in email.subject
        ]
        workflow_submission_emailed_addresses = [
            address for email in workflow_submission_emails for address in email.to
        ]

        # with `WAGTAILADMIN_NOTIFICATION_INCLUDE_SUPERUSERS` off, the superuser should not get a task email
        self.assertNotIn(self.superuser.email, task_submission_emailed_addresses)

        # with `WAGTAILADMIN_NOTIFICATION_INCLUDE_SUPERUSERS` off, the superuser should not get a workflow email
        self.assertNotIn(self.superuser.email, workflow_submission_emailed_addresses)

    @override_settings(WAGTAILADMIN_NOTIFICATION_INCLUDE_SUPERUSERS=True)
    def test_submit_notification_active_users_only(self):
        # moderator2 is inactive
        self.moderator2.is_active = False
        self.moderator2.save()

        # superuser is inactive
        self.superuser.is_active = False
        self.superuser.save()

        # Submit
        self.login(self.submitter)
        self.post("submit")

        workflow_submission_emails = [
            email for email in mail.outbox if "workflow" in email.subject
        ]
        workflow_submission_emailed_addresses = [
            address for email in workflow_submission_emails for address in email.to
        ]
        task_submission_emails = [
            email for email in mail.outbox if "task" in email.subject
        ]
        task_submission_emailed_addresses = [
            address for email in task_submission_emails for address in email.to
        ]

        # Check that moderator2 didn't receive a task submitted email
        self.assertNotIn(self.moderator2.email, task_submission_emailed_addresses)

        # Check that the superuser didn't receive a workflow or task email
        self.assertNotIn(self.superuser.email, task_submission_emailed_addresses)
        self.assertNotIn(self.superuser.email, workflow_submission_emailed_addresses)

    @override_settings(WAGTAILADMIN_NOTIFICATION_INCLUDE_SUPERUSERS=True)
    def test_submit_notification_preferences_respected(self):
        # moderator2 doesn't want emails
        self.moderator2_profile.submitted_notifications = False
        self.moderator2_profile.save()

        # superuser doesn't want emails
        self.superuser_profile.submitted_notifications = False
        self.superuser_profile.save()

        # Submit
        self.login(self.submitter)
        self.post("submit")

        # Check that only one moderator got a task submitted email
        workflow_submission_emails = [
            email for email in mail.outbox if "workflow" in email.subject
        ]
        workflow_submission_emailed_addresses = [
            address for email in workflow_submission_emails for address in email.to
        ]
        task_submission_emails = [
            email for email in mail.outbox if "task" in email.subject
        ]
        task_submission_emailed_addresses = [
            address for email in task_submission_emails for address in email.to
        ]
        self.assertNotIn(self.moderator2.email, task_submission_emailed_addresses)

        # Check that the superuser didn't receive a workflow or task email
        self.assertNotIn(self.superuser.email, task_submission_emailed_addresses)
        self.assertNotIn(self.superuser.email, workflow_submission_emailed_addresses)

    def test_approved_notifications(self):
        self.login(self.submitter)
        self.post("submit")
        # Approve
        self.login(self.moderator)
        self.approve()

        # Submitter must receive a workflow approved email
        workflow_approved_emails = [
            email
            for email in mail.outbox
            if ("workflow" in email.subject and "approved" in email.subject)
        ]
        self.assertEqual(len(workflow_approved_emails), 1)
        self.assertIn(self.submitter.email, workflow_approved_emails[0].to)

    def test_approved_notifications_preferences_respected(self):
        # Submitter doesn't want 'approved' emails
        self.submitter_profile.approved_notifications = False
        self.submitter_profile.save()

        self.login(self.submitter)
        self.post("submit")
        # Approve
        self.login(self.moderator)
        self.approve()

        # Submitter must not receive a workflow approved email, so there should be no emails in workflow_approved_emails
        workflow_approved_emails = [
            email
            for email in mail.outbox
            if ("workflow" in email.subject and "approved" in email.subject)
        ]
        self.assertEqual(len(workflow_approved_emails), 0)

    def test_rejected_notifications(self):
        self.login(self.submitter)
        self.post("submit")
        # Reject
        self.login(self.moderator)
        self.reject()

        # Submitter must receive a workflow rejected email
        workflow_rejected_emails = [
            email
            for email in mail.outbox
            if ("workflow" in email.subject and "rejected" in email.subject)
        ]
        self.assertEqual(len(workflow_rejected_emails), 1)
        self.assertIn(self.submitter.email, workflow_rejected_emails[0].to)

    def test_rejected_notification_preferences_respected(self):
        # Submitter doesn't want 'rejected' emails
        self.submitter_profile.rejected_notifications = False
        self.submitter_profile.save()

        self.login(self.submitter)
        self.post("submit")
        # Reject
        self.login(self.moderator)
        self.reject()

        # Submitter must not receive a workflow rejected email
        workflow_rejected_emails = [
            email
            for email in mail.outbox
            if ("workflow" in email.subject and "rejected" in email.subject)
        ]
        self.assertEqual(len(workflow_rejected_emails), 0)


@override_settings(WAGTAILADMIN_NOTIFICATION_USE_HTML=True)
class TestPageNotificationPreferencesHTML(TestPageNotificationPreferences):
    pass


class TestSnippetNotificationPreferences(
    TestPageNotificationPreferences, BaseSnippetWorkflowTests
):
    def setUp(self):
        super().setUp()
        self.moderator2.user_permissions.add(
            self.edit_permission,
            self.publish_permission,
        )


@override_settings(WAGTAILADMIN_NOTIFICATION_USE_HTML=True)
class TestSnippetNotificationPreferencesHTML(TestSnippetNotificationPreferences):
    pass


class TestDisableViews(BasePageWorkflowTests):
    def test_disable_workflow(self):
        """Test that deactivating a workflow sets it to inactive and cancels in progress states"""
        self.login(self.submitter)
        self.post("submit")
        self.login(self.superuser)
        self.approve()

        response = self.client.post(
            reverse("wagtailadmin_workflows:disable", args=(self.workflow.pk,))
        )
        self.assertEqual(response.status_code, 302)
        self.workflow.refresh_from_db()
        self.assertIs(self.workflow.active, False)
        states = WorkflowState.objects.for_instance(self.object).filter(
            workflow=self.workflow
        )
        self.assertEqual(
            states.filter(status=WorkflowState.STATUS_IN_PROGRESS).count(), 0
        )
        self.assertEqual(
            states.filter(status=WorkflowState.STATUS_CANCELLED).count(), 1
        )

        self.assertEqual(
            TaskState.objects.filter(
                workflow_state__workflow=self.workflow,
                status=TaskState.STATUS_IN_PROGRESS,
            ).count(),
            0,
        )

    def test_get_disable_workflow_shows_warning(self):
        """Test that deactivating a workflow shows a warning if there are in progress states"""
        self.login(self.submitter)
        self.post("submit")
        self.login(self.superuser)
        self.approve()

        response = self.client.get(
            reverse("wagtailadmin_workflows:disable", args=(self.workflow.pk,))
        )
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "wagtailadmin/workflows/confirm_disable.html")
        self.assertContains(
            response,
            "This workflow is in progress on 1 page/snippet. Disabling this workflow will cancel moderation on this page/snippet.",
        )
        self.assertBreadcrumbsNotRendered(response.content)

    def test_get_disable_workflow_no_warning(self):
        """Test that deactivating a workflow does not show a warning if there are no in progress states"""
        self.login(self.superuser)

        response = self.client.get(
            reverse("wagtailadmin_workflows:disable", args=(self.workflow.pk,))
        )
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "wagtailadmin/workflows/confirm_disable.html")
        self.assertNotContains(response, "This workflow is in progress")
        self.assertNotContains(
            response,
            "Disabling this workflow will cancel moderation on this page/snippet.",
        )
        self.assertBreadcrumbsNotRendered(response.content)

    def test_disable_task_view(self):
        """Test that a view is shown before disabling a task that shows a warning"""
        self.login(self.submitter)
        self.post("submit")
        self.login(self.superuser)

        response = self.client.get(
            reverse("wagtailadmin_workflows:disable_task", args=(self.task_1.pk,))
        )

        self.assertTemplateUsed(
            response, "wagtailadmin/workflows/confirm_disable_task.html"
        )
        self.assertEqual(
            response.context["warning_message"],
            "This task is in progress on 1 page/snippet. Disabling this task will cause it to be skipped in the moderation workflow and not be listed for selection when editing a workflow.",
        )
        self.assertContains(
            response,
            "This task is in progress on 1 page/snippet. Disabling this task will cause it to be skipped in the moderation workflow and not be listed for selection when editing a workflow.",
        )
        self.assertBreadcrumbsNotRendered(response.content)

        # create a new, unused, task and check there is no warning message
        unused_task = GroupApprovalTask.objects.create(name="unused_task_3")
        unused_task.groups.set(Group.objects.filter(name="Moderators"))

        response = self.client.get(
            reverse("wagtailadmin_workflows:disable_task", args=(unused_task.pk,))
        )

        self.assertNotIn("warning_message", response.context)
        self.assertNotContains(response, "This task is in progress")
        self.assertNotContains(
            response,
            "Disabling this task will cause it to be skipped in the moderation workflow "
            "and not be listed for selection when editing a workflow.",
        )
        self.assertBreadcrumbsNotRendered(response.content)

        unused_task.delete()  # clean up

    def test_disable_task(self):
        """Test that deactivating a task sets it to inactive and cancels in progress states"""
        self.login(self.submitter)
        self.post("submit")
        self.login(self.superuser)

        response = self.client.post(
            reverse("wagtailadmin_workflows:disable_task", args=(self.task_1.pk,))
        )
        self.assertEqual(response.status_code, 302)
        self.task_1.refresh_from_db()
        self.assertIs(self.task_1.active, False)
        states = TaskState.objects.for_instance(self.object).filter(
            task=self.task_1.task_ptr
        )
        self.assertEqual(states.filter(status=TaskState.STATUS_IN_PROGRESS).count(), 0)
        self.assertEqual(states.filter(status=TaskState.STATUS_CANCELLED).count(), 1)

        # Check that the object's WorkflowState has moved on to the next active task
        self.assertEqual(
            self.object.current_workflow_state.current_task_state.task.specific,
            self.task_2,
        )

    def test_enable_workflow(self):
        self.login(self.superuser)
        self.workflow.active = False
        self.workflow.save()

        response = self.client.post(
            reverse("wagtailadmin_workflows:enable", args=(self.workflow.pk,))
        )
        self.assertEqual(response.status_code, 302)
        self.workflow.refresh_from_db()
        self.assertIs(self.workflow.active, True)

    def test_enable_workflow_minimal_permissions(self):
        self.superuser.is_superuser = False
        self.superuser.save()
        self.superuser.user_permissions.add(
            Permission.objects.get(
                content_type__app_label="wagtailadmin",
                codename="access_admin",
            ),
            Permission.objects.get(codename="add_workflow"),
        )
        self.login(self.superuser)
        self.workflow.active = False
        self.workflow.save()

        response = self.client.post(
            reverse("wagtailadmin_workflows:enable", args=(self.workflow.pk,))
        )
        self.assertEqual(response.status_code, 302)
        self.workflow.refresh_from_db()
        self.assertIs(self.workflow.active, True)

    def test_enable_workflow_no_permissions(self):
        self.superuser.is_superuser = False
        self.superuser.save()
        self.superuser.user_permissions.add(
            Permission.objects.get(
                content_type__app_label="wagtailadmin",
                codename="access_admin",
            ),
        )
        self.login(self.superuser)
        self.workflow.active = False
        self.workflow.save()

        response = self.client.post(
            reverse("wagtailadmin_workflows:enable", args=(self.workflow.pk,))
        )
        self.assertRedirects(response, reverse("wagtailadmin_home"))
        self.workflow.refresh_from_db()
        self.assertIs(self.workflow.active, False)

    def test_enable_task(self):
        self.login(self.superuser)
        self.task_1.active = False
        self.task_1.save()

        response = self.client.post(
            reverse("wagtailadmin_workflows:enable_task", args=(self.task_1.pk,))
        )
        self.assertEqual(response.status_code, 302)
        self.task_1.refresh_from_db()
        self.assertIs(self.task_1.active, True)

    def test_enable_task_minimal_permissions(self):
        self.superuser.is_superuser = False
        self.superuser.save()
        self.superuser.user_permissions.add(
            Permission.objects.get(
                content_type__app_label="wagtailadmin",
                codename="access_admin",
            ),
            Permission.objects.get(codename="add_task"),
        )
        self.login(self.superuser)
        self.task_1.active = False
        self.task_1.save()

        response = self.client.post(
            reverse("wagtailadmin_workflows:enable_task", args=(self.task_1.pk,))
        )
        self.assertEqual(response.status_code, 302)
        self.task_1.refresh_from_db()
        self.assertIs(self.task_1.active, True)

    def test_enable_task_no_permissions(self):
        self.superuser.is_superuser = False
        self.superuser.save()
        self.superuser.user_permissions.add(
            Permission.objects.get(
                content_type__app_label="wagtailadmin",
                codename="access_admin",
            ),
        )
        self.login(self.superuser)
        self.task_1.active = False
        self.task_1.save()

        response = self.client.post(
            reverse("wagtailadmin_workflows:enable_task", args=(self.task_1.pk,))
        )
        self.assertRedirects(response, reverse("wagtailadmin_home"))
        self.task_1.refresh_from_db()
        self.assertIs(self.task_1.active, False)


class TestDisableViewsWithSnippetWorkflows(TestDisableViews, BaseSnippetWorkflowTests):
    pass


class TestPageWorkflowPreview(BasePageWorkflowTests):
    preview_template = "tests/simple_page.html"
    preview_content = "Simple page"
    new_content = "Not-so-simple object"

    def setUp(self):
        super().setUp()
        self.edit_object()
        self.workflow.start(self.object, self.submitter)
        self.login(self.moderator)

    def edit_object(self):
        self.object.title = self.new_content
        self.object.save_revision()

    def test_preview_workflow(self):
        preview_url = self.get_url(
            "workflow_preview",
            args=(quote(self.object.pk), self.object.current_workflow_task.id),
        )
        response = self.client.get(preview_url)

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, self.preview_template)
        self.assertContains(response, self.preview_content)

    def test_preview_workflow_show_edit_link_in_userbar(self):
        preview_url = self.get_url(
            "workflow_preview",
            args=(quote(self.object.pk), self.object.current_workflow_task.id),
        )
        response = self.client.get(preview_url)

        # Should show edit link in the userbar
        # https://github.com/wagtail/wagtail/issues/10002
        self.assertContains(response, "Edit this page")
        self.assertContains(
            response, reverse("wagtailadmin_pages:edit", args=(quote(self.object.pk),))
        )

    def test_preview_workflow_by_submitter(self):
        self.login(self.submitter)
        preview_url = self.get_url(
            "workflow_preview",
            args=(quote(self.object.pk), self.object.current_workflow_task.id),
        )
        response = self.client.get(preview_url)

        self.assertRedirects(response, reverse("wagtailadmin_home"))

    def test_preview_workflow_bad_task_id(self):
        preview_url = self.get_url(
            "workflow_preview",
            args=(quote(self.object.pk), self.task_2.id),
        )
        response = self.client.get(preview_url, follow=True)

        self.assertRedirects(response, reverse("wagtailadmin_home"))
        self.assertContains(
            response,
            f"The {self.model_name} &#x27;{get_latest_str(self.object)}&#x27; is not "
            f"currently awaiting moderation in task &#x27;{self.task_2.name}&#x27;.",
        )
        self.assertContains(response, self.new_content)

    def test_preview_workflow_nonexistent_ids(self):
        preview_url = self.get_url(
            "workflow_preview",
            args=(quote(self.object.pk), 123),
        )
        response = self.client.get(preview_url)

        self.assertEqual(response.status_code, 404)

        preview_url = self.get_url("workflow_preview", args=(123, self.task_1.id))
        response = self.client.get(preview_url)

        self.assertEqual(response.status_code, 404)


class TestSnippetWorkflowPreview(TestPageWorkflowPreview, BaseSnippetWorkflowTests):
    preview_template = "tests/previewable_model.html"
    preview_content = "Not-so-simple object (Default Preview)"

    def edit_object(self):
        self.object.text = self.new_content
        self.object.save_revision()

    @expectedFailure
    def test_preview_workflow_show_edit_link_in_userbar(self):
        # TODO: Add edit link to userbar on snippet previews
        super().test_preview_workflow_show_edit_link_in_userbar()


class TestTaskChooserView(WagtailTestUtils, TestCase):
    def setUp(self):
        self.login()

        self.task_enabled = GroupApprovalTask.objects.create(name="Enabled foo")
        self.task_disabled = GroupApprovalTask.objects.create(
            name="Disabled foo", active=False
        )

    def test_get(self):
        response = self.client.get(reverse("wagtailadmin_workflows:task_chooser"))

        self.assertEqual(response.status_code, 200)

        self.assertTemplateUsed(
            response, "wagtailadmin/workflows/task_chooser/chooser.html"
        )

        # Check that the "select task type" view was shown in the "new" tab
        self.assertTemplateUsed(
            response,
            "wagtailadmin/workflows/task_chooser/includes/select_task_type.html",
        )
        self.assertTemplateUsed(
            response, "wagtailadmin/workflows/task_chooser/includes/results.html"
        )
        self.assertTemplateNotUsed(
            response, "wagtailadmin/workflows/task_chooser/includes/create_form.html"
        )
        self.assertFalse(response.context["search_form"].is_searching())
        # check that only active (non-disabled) tasks are listed
        self.assertEqual(
            [task.name for task in response.context["tasks"].object_list],
            ["Enabled foo", "Moderators approval"],
        )

    def test_search(self):
        response = self.client.get(
            reverse("wagtailadmin_workflows:task_chooser_results") + "?q=foo"
        )

        self.assertEqual(response.status_code, 200)

        self.assertTemplateUsed(
            response, "wagtailadmin/workflows/task_chooser/includes/results.html"
        )
        self.assertTemplateNotUsed(
            response, "wagtailadmin/workflows/task_chooser/chooser.html"
        )
        self.assertTrue(response.context["search_form"].is_searching())
        self.assertEqual(response.context["query_string"], "foo")
        # check that only active (non-disabled) tasks are listed
        self.assertEqual(
            [task.name for task in response.context["tasks"].object_list],
            ["Enabled foo"],
        )

    def test_pagination(self):
        response = self.client.get(
            reverse("wagtailadmin_workflows:task_chooser_results") + "?p=2"
        )

        self.assertEqual(response.status_code, 200)

        # When pagination is used, only the results template should be rendered
        self.assertTemplateUsed(
            response, "wagtailadmin/workflows/task_chooser/includes/results.html"
        )
        self.assertTemplateNotUsed(
            response, "wagtailadmin/workflows/task_chooser/chooser.html"
        )
        self.assertFalse(response.context["search_form"].is_searching())

    def test_get_with_create_model_selected(self):
        response = self.client.get(
            reverse("wagtailadmin_workflows:task_chooser_create")
            + "?create_model=wagtailcore.GroupApprovalTask"
        )

        self.assertEqual(response.status_code, 200)

        # Check that the "create" view was returned
        self.assertTemplateUsed(
            response, "wagtailadmin/workflows/task_chooser/includes/create_form.html"
        )
        self.assertTemplateNotUsed(
            response,
            "wagtailadmin/workflows/task_chooser/includes/select_task_type.html",
        )

    def test_get_with_non_task_create_model_selected(self):
        response = self.client.get(
            reverse("wagtailadmin_workflows:task_chooser_create")
            + "?create_model=wagtailcore.Page"
        )

        self.assertEqual(response.status_code, 404)

    def test_get_with_base_task_create_model_selected(self):
        # Task is technically a subclass of itself so we need an extra test for it
        response = self.client.get(
            reverse("wagtailadmin_workflows:task_chooser_create")
            + "?create_model=wagtailcore.Task"
        )

        self.assertEqual(response.status_code, 404)

    @mock.patch("wagtail.admin.views.workflows.get_task_types")
    def test_get_with_single_task_model(self, get_task_types):
        # When a single task type exists there's no need to specify create_model
        get_task_types.return_value = [GroupApprovalTask]

        response = self.client.get(reverse("wagtailadmin_workflows:task_chooser"))

        self.assertEqual(response.status_code, 200)

        self.assertTemplateUsed(
            response, "wagtailadmin/workflows/task_chooser/chooser.html"
        )

        # Check that the "create" view was shown in the "new" tab
        self.assertTemplateUsed(
            response, "wagtailadmin/workflows/task_chooser/includes/create_form.html"
        )
        self.assertTemplateNotUsed(
            response,
            "wagtailadmin/workflows/task_chooser/includes/select_task_type.html",
        )

    # POST requests are for creating new tasks

    def get_post_data(self):
        return {
            "create-task-name": "Editor approval task",
            "create-task-groups": [str(Group.objects.get(name="Editors").id)],
        }

    def test_post_with_create_model_selected(self):
        response = self.client.post(
            reverse("wagtailadmin_workflows:task_chooser_create")
            + "?create_model=wagtailcore.GroupApprovalTask",
            self.get_post_data(),
        )

        self.assertEqual(response.status_code, 200)

        # Check that the task was created
        task = Task.objects.get(name="Editor approval task", active=True)

        # Check the response JSON
        self.assertEqual(
            response.json(),
            {
                "step": "task_chosen",
                "result": {
                    "id": task.id,
                    "name": "Editor approval task",
                    "edit_url": reverse(
                        "wagtailadmin_workflows:edit_task", args=[task.id]
                    ),
                },
            },
        )

    @mock.patch("wagtail.admin.views.workflows.get_task_types")
    def test_post_with_single_task_model(self, get_task_types):
        # When a single task type exists there's no need to specify create_model
        get_task_types.return_value = [GroupApprovalTask]

        response = self.client.post(
            reverse("wagtailadmin_workflows:task_chooser_create"), self.get_post_data()
        )

        self.assertEqual(response.status_code, 200)

        # Check that the task was created
        task = Task.objects.get(name="Editor approval task", active=True)

        # Check the response JSON
        self.assertEqual(
            response.json(),
            {
                "step": "task_chosen",
                "result": {
                    "id": task.id,
                    "name": "Editor approval task",
                    "edit_url": reverse(
                        "wagtailadmin_workflows:edit_task", args=[task.id]
                    ),
                },
            },
        )

    def test_post_without_create_model_selected(self):
        response = self.client.post(
            reverse("wagtailadmin_workflows:task_chooser_create"), self.get_post_data()
        )

        self.assertEqual(response.status_code, 400)

        # Check that the task wasn't created
        self.assertFalse(
            Task.objects.filter(name="Editor approval task", active=True).exists()
        )

    def test_post_with_non_task_create_model_selected(self):
        response = self.client.post(
            reverse("wagtailadmin_workflows:task_chooser_create")
            + "?create_model=wagtailcore.Page",
            self.get_post_data(),
        )

        self.assertEqual(response.status_code, 404)

        # Check that the task wasn't created
        self.assertFalse(
            Task.objects.filter(name="Editor approval task", active=True).exists()
        )

    def test_post_with_base_task_create_model_selected(self):
        # Task is technically a subclass of itself so we need an extra test for it
        response = self.client.post(
            reverse("wagtailadmin_workflows:task_chooser_create")
            + "?create_model=wagtailcore.Task",
            self.get_post_data(),
        )

        self.assertEqual(response.status_code, 404)

        # Check that the task wasn't created
        self.assertFalse(
            Task.objects.filter(name="Editor approval task", active=True).exists()
        )


class TestTaskChooserChosenView(WagtailTestUtils, TestCase):
    def setUp(self):
        delete_existing_workflows()
        self.login()
        self.task = SimpleTask.objects.create(name="test_task")

    def test_get(self):
        response = self.client.get(
            reverse("wagtailadmin_workflows:task_chosen", args=[self.task.id])
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/json")
        self.assertEqual(
            response.json(),
            {
                "result": {
                    "edit_url": reverse(
                        "wagtailadmin_workflows:edit_task", args=[self.task.id]
                    ),
                    "id": self.task.id,
                    "name": "test_task",
                },
                "step": "task_chosen",
            },
        )


class TestWorkflowUsageView(WagtailTestUtils, TestCase):
    def setUp(self):
        self.user = self.login()
        self.workflow = Workflow.objects.get()

        self.root_page = Page.objects.get(depth=1)
        self.home_page = Page.objects.get(depth=2)

        self.child_page_with_default_workflow = self.home_page.add_child(
            instance=SimplePage(title="A page", content="I'm a page")
        )
        self.child_page_with_another_workflow = self.home_page.add_child(
            instance=SimplePage(title="Another page", content="I'm another page")
        )
        self.another_workflow = Workflow.objects.create(name="Another workflow")
        self.another_workflow.workflow_pages.create(
            page=self.child_page_with_another_workflow
        )

    def test_get(self):
        response = self.client.get(
            reverse("wagtailadmin_workflows:usage", args=[self.workflow.id])
        )

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "wagtailadmin/pages/listing.html")

        object_set = {page.id for page in response.context["object_list"]}
        # Should not contain child_page_with_another_workflow. It should also
        # not contain the root page, as it's irrelevant (you'll never be able to
        # edit and submit it to the workflow)
        self.assertEqual(
            object_set,
            {self.home_page.id, self.child_page_with_default_workflow.id},
        )

    def test_with_no_permission(self):
        group = Group.objects.create(name="test group")
        self.user.is_superuser = False
        self.user.save()
        self.user.groups.add(group)
        group.permissions.add(
            Permission.objects.get(
                content_type__app_label="wagtailadmin", codename="access_admin"
            )
        )
        # No GroupPagePermission created

        response = self.client.get(
            reverse("wagtailadmin_workflows:usage", args=[self.workflow.id])
        )
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("wagtailadmin_home"))

        # Only a page permission is created, but no workflow permission, not enough
        permission = GroupPagePermission.objects.create(
            group=group,
            page=Page.objects.first(),
            permission_type="change",
        )

        response = self.client.get(
            reverse("wagtailadmin_workflows:usage", args=[self.workflow.id])
        )
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("wagtailadmin_home"))

        # Delete page permission and add workflow permission, also not enough
        permission.delete()
        group.permissions.add(
            Permission.objects.get(
                content_type__app_label="wagtailcore", codename="change_workflow"
            )
        )

        response = self.client.get(
            reverse("wagtailadmin_workflows:usage", args=[self.workflow.id])
        )
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("wagtailadmin_home"))

    def test_with_minimal_permissions(self):
        group = Group.objects.create(name="test group")
        self.user.is_superuser = False
        self.user.save()
        self.user.groups.add(group)
        group.permissions.add(
            Permission.objects.get(
                content_type__app_label="wagtailadmin", codename="access_admin"
            ),
            Permission.objects.get(
                content_type__app_label="wagtailcore", codename="change_workflow"
            ),
        )
        GroupPagePermission.objects.create(
            group=group,
            page=Page.objects.first(),
            permission_type="change",
        )

        # With a workflow permission and a page permission, the user should be
        # able to access the view
        response = self.client.get(
            reverse("wagtailadmin_workflows:usage", args=[self.workflow.id])
        )

        self.assertEqual(response.status_code, 200)

    def test_get_search_and_filtered_results(self):
        page_1 = SimplePage(title="Hello wagtail", content="test")
        page_2 = SimplePage(title="Hello django", content="test")
        self.home_page.add_child(instance=page_1)
        self.home_page.add_child(instance=page_2)
        self.home_page.add_child(instance=MultiPreviewModesPage(title="Hello python"))

        response = self.client.get(
            reverse("wagtailadmin_workflows:usage_results", args=[self.workflow.id]),
            {
                "content_type": ContentType.objects.get_for_model(SimplePage).id,
                "q": "hello",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "wagtailadmin/generic/listing_results.html")

        object_set = {page.id for page in response.context["object_list"]}
        self.assertEqual(object_set, {page_1.id, page_2.id})


@freeze_time("2020-06-01 12:00:00")
class TestPageWorkflowStatus(BasePageWorkflowTests):
    def setUp(self):
        super().setUp()
        self.login(self.superuser)

    def setup_workflow_and_tasks(self):
        self.workflow = Workflow.objects.create(name="test_workflow")
        self.task_1 = GroupApprovalTask.objects.create(name="test_task_1")
        self.task_1.groups.set(Group.objects.filter(name="Moderators"))
        WorkflowTask.objects.create(
            workflow=self.workflow, task=self.task_1, sort_order=1
        )

        self.task_2 = GroupApprovalTask.objects.create(name="test_task_2")
        self.task_2.groups.set(Group.objects.filter(name="Editors"))
        WorkflowTask.objects.create(
            workflow=self.workflow, task=self.task_2, sort_order=2
        )

    def workflow_action(self, action):
        return super().workflow_action(
            action,
            {
                "action": action,
                "comment": "good work" if action == "approve" else "needs some changes",
                "next": self.get_url("edit"),
            },
        )

    def test_workflow_status_modal(self):
        # The workflow status view should return permission denied when the object is but a draft
        response = self.client.get(self.get_url("edit"))
        html = response.content.decode("utf-8")
        self.assertNotIn('id="workflow-status-dialog"', html)

        # Submit for moderation
        self.post("submit")

        response = self.client.get(self.get_url("edit"))
        html = response.content.decode("utf-8")
        self.assertIn(
            "In progress\n        </span>\n                        {}".format(
                self.task_1.name
            ),
            html,
        )
        self.assertIn(
            "Not started\n        </span>\n                        {}".format(
                self.task_2.name
            ),
            html,
        )
        self.assertIn(self.get_url("history"), html)

    def test_status_through_workflow_cycle(self):
        self.login(self.superuser)
        response = self.client.get(self.get_url("edit"))
        self.assertContains(response, "Draft", 1)

        self.object.save_revision()
        response = self.client.get(self.get_url("edit"))
        self.assertContains(response, 'id="status-sidebar-draft"')

        self.post("submit")
        response = self.client.get(self.get_url("edit"))
        self.assertRegex(
            response.content.decode("utf-8"),
            rf"Sent to[\s|\n]+{self.task_1.name}",
        )

        response = self.workflow_action("approve")
        self.assertRegex(
            response.content.decode("utf-8"),
            rf"Sent to[\s|\n]+{self.task_2.name}",
        )

        response = self.workflow_action("reject")
        self.assertContains(response, "Changes requested")

        # resubmit
        self.post("submit")
        response = self.client.get(self.get_url("edit"))
        self.assertRegex(
            response.content.decode("utf-8"),
            rf"Sent to[\s|\n]+{self.task_2.name}",
        )

        response = self.workflow_action("approve")
        self.assertContains(response, 'id="status-sidebar-live"')

    def test_status_after_cancel(self):
        # start workflow, then cancel
        self.post("submit")
        self.post("cancel-workflow")
        response = self.client.get(self.get_url("edit"))
        self.assertContains(response, 'id="status-sidebar-draft"')

    def test_status_after_restart(self):
        self.post("submit")
        response = self.workflow_action("approve")
        self.assertRegex(
            response.content.decode("utf-8"),
            rf"Sent to[\s|\n]+{self.task_2.name}",
        )
        self.workflow_action("reject")
        self.post("restart-workflow")
        response = self.client.get(self.get_url("edit"))
        self.assertRegex(
            response.content.decode("utf-8"),
            rf"Sent to[\s|\n]+{self.task_1.name}",
        )

    def test_workflow_status_modal_task_comments(self):
        self.post("submit")
        self.workflow_action("reject")

        response = self.client.get(self.get_url("edit"))
        self.assertIn("needs some changes", response.content.decode("utf-8"))

        self.post("submit")
        self.workflow_action("approve")
        response = self.client.get(self.get_url("edit"))
        self.assertIn("good work", response.content.decode("utf-8"))

    def test_workflow_edit_locked_message(self):
        self.post("submit")
        self.login(self.submitter)
        response = self.client.get(self.get_url("edit"))

        needle = (
            f"This {self.model_name} is awaiting <b>'test_task_1'</b> in the "
            "<b>'test_workflow'</b> workflow. Only reviewers for this task "
            f"can edit the {self.model_name}."
        )
        self.assertContains(response, needle, count=1)
        self.assertNotContains(response, "Save draft")

        self.login(self.moderator)
        response = self.client.get(self.get_url("edit"))
        self.assertNotContains(response, needle)
        self.assertContains(response, "Save draft")


class TestSnippetWorkflowStatus(TestPageWorkflowStatus, BaseSnippetWorkflowTests):
    pass


class TestSnippetWorkflowStatusNotLockable(TestSnippetWorkflowStatus):
    model = ModeratedModel

    def test_workflow_edit_locked_message(self):
        # Without LockableMixin, the edit view should not be locked
        self.post("submit")
        self.login(self.submitter)
        response = self.client.get(self.get_url("edit"))

        needle = "Only reviewers for this task can edit"
        self.assertNotContains(response, needle)
        self.assertContains(response, "Save draft")

        self.login(self.moderator)
        response = self.client.get(self.get_url("edit"))
        self.assertNotContains(response, needle)
        self.assertContains(response, "Save draft")


class TestDashboardWithPages(BasePageWorkflowTests):
    def setUp(self):
        super().setUp()
        # Ensure that the presence of private pages doesn't break the dashboard -
        # https://github.com/wagtail/wagtail/issues/10819
        homepage = Page.objects.filter(depth=2).first()
        PageViewRestriction.objects.create(
            page=homepage, restriction_type=PageViewRestriction.LOGIN
        )

    def test_dashboard_for_submitter(self):
        self.login(self.submitter)
        self.post("submit")

        response = self.client.get(reverse("wagtailadmin_home"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Your pages and snippets in a workflow")

    def test_dashboard_for_moderator(self):
        self.login(self.submitter)
        self.post("submit")

        self.login(self.moderator)
        response = self.client.get(reverse("wagtailadmin_home"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Awaiting your review")

        # object had no previous revisions
        self.assertNotContains(response, "Compare with live version")
        self.assertNotContains(response, "Compare with previous version")

    def test_dashboard_for_moderator_with_previous_revisions(self):
        live_revision = self.object.save_revision()
        self.object.publish(live_revision)
        previous_revision = self.object.save_revision()

        self.login(self.submitter)
        self.post("submit")
        self.object.refresh_from_db()
        latest_revision = self.object.latest_revision

        self.login(self.moderator)
        response = self.client.get(reverse("wagtailadmin_home"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Awaiting your review")

        soup = self.get_soup(response.content)
        compare_with_live_url = self.get_url(
            "revisions_compare",
            args=(self.object.pk, "live", latest_revision.id),
        )
        compare_with_previous_url = self.get_url(
            "revisions_compare",
            args=(self.object.pk, previous_revision.id, latest_revision.id),
        )

        compare_with_live_link = soup.select_one(f"a[href='{compare_with_live_url}']")
        self.assertIsNotNone(compare_with_live_link)
        self.assertEqual(
            compare_with_live_link.text.strip(),
            "Compare with live version",
        )

        compare_with_previous_link = soup.select_one(
            f"a[href='{compare_with_previous_url}']"
        )
        self.assertIsNotNone(compare_with_previous_link)
        self.assertEqual(
            compare_with_previous_link.text.strip(),
            "Compare with previous version",
        )

    def test_dashboard_after_deleting_object_in_moderation(self):
        # WorkflowState's content_object may point to a nonexistent object
        # https://github.com/wagtail/wagtail/issues/11300
        self.login(self.submitter)
        self.post("submit")
        self.object.delete()

        response = self.client.get(reverse("wagtailadmin_home"))
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Your pages and snippets in a workflow")

        self.login(self.moderator)
        response = self.client.get(reverse("wagtailadmin_home"))
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Awaiting your review")


class TestDashboardWithSnippets(TestDashboardWithPages, BaseSnippetWorkflowTests):
    pass


class TestDashboardWithNonLockableSnippets(TestDashboardWithSnippets):
    # This model does not use LockableMixin, and it also does not have a
    # GenericRelation to WorkflowState and Revision, but it should not break
    # the dashboard.
    # See https://github.com/wagtail/wagtail/issues/11300 for more details.
    model = ModeratedModel


class TestWorkflowStateEmailNotifier(BasePageWorkflowTests):
    def setUp(self):
        super().setUp()
        # Ensure a revision exists
        self.object.save_revision()

    def test_workflowstate_email_notifier_get_recipient_users__without_triggering_user(
        self,
    ):
        self.workflow.start(self.object, user=self.submitter)
        workflow_state = self.object.current_workflow_state

        for notifier in [
            WorkflowStateApprovalEmailNotifier(),
            WorkflowStateRejectionEmailNotifier(),
        ]:
            with self.subTest(f"Testing with {notifier}"):
                self.assertSetEqual(
                    notifier.get_recipient_users(workflow_state), {self.submitter}
                )

    def test_workflowstate_email_notifier_get_recipient_users__with_triggering_user(
        self,
    ):
        self.workflow.start(self.object, user=self.submitter)
        workflow_state = self.object.current_workflow_state

        for notifier in [
            WorkflowStateApprovalEmailNotifier(),
            WorkflowStateRejectionEmailNotifier(),
        ]:
            with self.subTest(f"Testing with {notifier}"):
                self.assertSetEqual(
                    notifier.get_recipient_users(workflow_state, user=self.moderator),
                    {self.submitter},
                )

    def test_workflowstate_email_notifier_get_recipient_users__without_requested_by(
        self,
    ):
        self.workflow.start(self.object, user=self.submitter)
        workflow_state: WorkflowState = self.object.current_workflow_state
        workflow_state.requested_by = None
        workflow_state.save(update_fields=["requested_by"])

        for notifier in [
            WorkflowStateApprovalEmailNotifier(),
            WorkflowStateRejectionEmailNotifier(),
        ]:
            with self.subTest(f"Testing with {notifier}"):
                self.assertSetEqual(notifier.get_recipient_users(workflow_state), set())
                self.assertSetEqual(
                    notifier.get_recipient_users(workflow_state, user=self.moderator),
                    set(),
                )

    def test_workflowstate_email_notifier_get_recipient_users__with_same_requested_by_and_triggering_user(
        self,
    ):
        self.workflow.start(self.object, user=self.submitter)
        workflow_state: WorkflowState = self.object.current_workflow_state
        workflow_state.requested_by = None
        workflow_state.save(update_fields=["requested_by"])

        for notifier in [
            WorkflowStateApprovalEmailNotifier(),
            WorkflowStateRejectionEmailNotifier(),
        ]:
            with self.subTest(f"Testing with {notifier}"):
                self.assertSetEqual(notifier.get_recipient_users(workflow_state), set())
                self.assertSetEqual(
                    notifier.get_recipient_users(workflow_state, user=self.submitter),
                    set(),
                )

    @mock.patch("wagtail.admin.mail.BaseWorkflowStateEmailNotifier.get_recipient_users")
    def test_base_workflowstate_email_notifier_get_valid_recipients(
        self, mock_get_recipient_users
    ):
        notifier = BaseWorkflowStateEmailNotifier()

        # check with an empty set
        mock_get_recipient_users.return_value = set()
        self.assertSetEqual(notifier.get_valid_recipients(self.object), set())

        # check None values are filtered out
        mock_get_recipient_users.return_value = {None}
        self.assertSetEqual(notifier.get_valid_recipients(self.object), set())

        # check with a valid user
        mock_get_recipient_users.return_value = {self.submitter}
        notifications = ["approved", "rejected", "submitted"]
        for notification in notifications:
            with self.subTest(f"Testing with {notification}_notifications"):
                notifier.notification = notification
                self.assertSetEqual(
                    notifier.get_valid_recipients(self.object), {self.submitter}
                )

        # remove notifications and re-test
        userprofile = UserProfile.get_for_user(self.submitter)
        updated_fields = []
        for notification in notifications:
            attribute = f"{notification}_notifications"
            setattr(userprofile, attribute, False)
            updated_fields.append(attribute)
        userprofile.save(update_fields=updated_fields)

        for notification in notifications:
            with self.subTest(f"Testing with {notification}_notifications"):
                notifier.notification = notification
                self.assertSetEqual(notifier.get_valid_recipients(self.object), set())
