import datetime
from io import BytesIO
from unittest import mock

from django.conf import settings
from django.conf.locale import LANG_INFO
from django.contrib.auth.models import Group, Permission
from django.contrib.contenttypes.models import ContentType
from django.db.models import F
from django.test import RequestFactory, TestCase
from django.test.utils import override_settings
from django.urls import reverse
from django.utils import timezone, translation
from freezegun import freeze_time
from openpyxl import load_workbook

from wagtail.admin.views.mixins import ExcelDateFormatter
from wagtail.admin.views.reports import page_types_usage
from wagtail.admin.views.reports.audit_logging import LogEntriesView
from wagtail.models import (
    GroupPagePermission,
    Locale,
    ModelLogEntry,
    Page,
    PageLogEntry,
    Site,
)
from wagtail.test.testapp.models import (
    Advert,
    EventPage,
    EventPageSpeaker,
    SimplePage,
)
from wagtail.test.utils import WagtailTestUtils
from wagtail.test.utils.template_tests import AdminTemplateTestUtils


class BaseReportViewTestCase(AdminTemplateTestUtils, WagtailTestUtils, TestCase):
    url_name = None
    header_buttons_parent_selector = "#w-slim-header-buttons"
    drilldown_selector = ".w-drilldown"
    extra_params = ""
    results_only = False

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.url = reverse(cls.url_name)
        if cls.results_only:
            cls.header_buttons_parent_selector = (
                '[data-controller="w-teleport"]'
                '[data-w-teleport-target-value="#w-slim-header-buttons"]'
            )
            cls.drilldown_selector = (
                '[data-controller="w-teleport"]'
                '[data-w-teleport-target-value="#filters-drilldown"]'
            )
            cls.extra_params = "&_w_filter_fragment=true"

    def setUp(self):
        self.user = self.login()

    def get(self, params={}, **kwargs):
        if self.results_only:
            params["_w_filter_fragment"] = "true"
        return self.client.get(self.url, params, **kwargs)

    def assertActiveFilter(self, soup, name, value):
        # Should render the export buttons inside the header "more" dropdown
        # with the filtered URL. When used in a results-only view, these are
        # teleported to the correct element in the skeleton.
        links_parent = soup.select_one(self.header_buttons_parent_selector)
        self.assertIsNotNone(links_parent)
        links = links_parent.select(".w-dropdown a")
        unfiltered_url = reverse(self.url_name)
        filtered_url = f"{unfiltered_url}?{name}={value}{self.extra_params}"
        self.assertEqual(len(links), 2)
        self.assertEqual(
            [link.get("href") for link in links],
            [f"{filtered_url}&export=xlsx", f"{filtered_url}&export=csv"],
        )

        # Should render the active filter pill
        active_filter = soup.select_one(".w-active-filters .w-pill__content")
        clear_button = soup.select_one(".w-active-filters .w-pill__remove")
        self.assertIsNotNone(active_filter)
        self.assertIsNotNone(clear_button)
        self.assertNotIn(name, clear_button.attrs.get("data-w-swap-src-value"))
        self.assertEqual(clear_button.attrs.get("data-w-swap-reflect-value"), "true")

    def assertActiveFilterNotRendered(self, soup):
        self.assertIsNone(soup.select_one(".w-active-filters"))

    def assertBreadcrumbs(self, breadcrumbs, html):
        if self.results_only:
            self.assertBreadcrumbsNotRendered(html)
        else:
            self.assertBreadcrumbsItemsRendered(breadcrumbs, html)

    def assertPageTitle(self, soup, title):
        page_title = soup.select_one("title")
        if self.results_only:
            self.assertIsNone(page_title)
        else:
            self.assertIsNotNone(page_title)
            self.assertEqual(page_title.text.strip(), title)


class TestLockedPagesView(BaseReportViewTestCase):
    url_name = "wagtailadmin_reports:locked_pages"

    def test_simple(self):
        response = self.get()
        self.assertEqual(response.status_code, 200)
        self.assertTemplateNotUsed(
            response,
            "wagtailadmin/reports/base_page_report.html",
        )
        self.assertTemplateUsed(response, "wagtailadmin/reports/base_report.html")
        self.assertTemplateUsed(
            response,
            "wagtailadmin/reports/locked_pages_results.html",
        )
        self.assertBreadcrumbs(
            [{"url": "", "label": "Locked pages"}],
            response.content,
        )

        # Initially there should be no locked pages
        self.assertContains(response, "No locked pages found.")

        # Should render the filter inside the drilldown
        soup = self.get_soup(response.content)
        locked_by_options = soup.select(
            f"{self.drilldown_selector} select[name='locked_by'] option"
        )
        # No user locked anything, so there should be no option for the filter
        self.assertEqual(len(locked_by_options), 1)
        self.assertEqual(locked_by_options[0].text, "---------")
        self.assertEqual(locked_by_options[0].get("value"), "")
        self.assertActiveFilterNotRendered(soup)
        self.assertPageTitle(soup, "Locked pages - Wagtail")

        parent_page = Page.objects.first()
        parent_page.add_child(
            instance=Page(
                title="First locked page",
                locked=True,
                locked_by=self.user,
                locked_at=timezone.now(),
            )
        )
        parent_page.add_child(
            instance=Page(
                title="Second locked page",
                locked=True,
                locked_by=self.user,
                locked_at=timezone.now(),
            )
        )

        # Now the listing should contain our locked page
        response = self.get()
        self.assertEqual(response.status_code, 200)
        self.assertTemplateNotUsed(
            response,
            "wagtailadmin/reports/base_page_report.html",
        )
        self.assertTemplateUsed(response, "wagtailadmin/reports/base_report.html")
        self.assertTemplateUsed(
            response,
            "wagtailadmin/reports/locked_pages_results.html",
        )
        self.assertBreadcrumbs(
            [{"url": "", "label": "Locked pages"}],
            response.content,
        )
        self.assertNotContains(response, "No locked pages found.")
        self.assertContains(response, "First locked page")
        self.assertContains(response, "Second locked page")

        # Should render the filter inside the drilldown
        soup = self.get_soup(response.content)
        locked_by_options = soup.select(
            f"{self.drilldown_selector} select[name='locked_by'] option"
        )
        # The options should only display users who have locked pages
        self.assertEqual(len(locked_by_options), 2)
        self.assertEqual(locked_by_options[0].text, "---------")
        self.assertIsNone(locked_by_options[0].value)
        self.assertEqual(locked_by_options[1].text, str(self.user))
        self.assertEqual(locked_by_options[1].get("value"), str(self.user.pk))
        self.assertActiveFilterNotRendered(soup)

        # Locked by current user shown in indicator
        self.assertNotContains(response, "indicator--is-dimmed")
        self.assertContains(
            response, 'title="This page is locked, by you, to further editing"'
        )

    def test_get_with_minimal_permissions(self):
        group = Group.objects.create(name="test group")
        self.user.is_superuser = False
        self.user.save()
        self.user.groups.add(group)
        self.user.user_permissions.add(
            Permission.objects.get(
                content_type__app_label="wagtailadmin", codename="access_admin"
            )
        )
        GroupPagePermission.objects.create(
            group=group,
            page=Page.objects.first(),
            permission_type="unlock",
        )

        response = self.get()

        self.assertEqual(response.status_code, 200)
        self.assertTemplateNotUsed(
            response,
            "wagtailadmin/reports/base_page_report.html",
        )
        self.assertTemplateUsed(response, "wagtailadmin/reports/base_report.html")
        self.assertTemplateUsed(
            response,
            "wagtailadmin/reports/locked_pages_results.html",
        )
        self.assertContains(response, "No locked pages found.")

    def test_get_with_no_permissions(self):
        self.user.is_superuser = False
        self.user.save()
        self.user.user_permissions.add(
            Permission.objects.get(
                content_type__app_label="wagtailadmin", codename="access_admin"
            )
        )

        response = self.get()

        self.assertRedirects(response, reverse("wagtailadmin_home"))

    def test_csv_export(self):
        self.page = Page.objects.first()
        self.page.locked = True
        self.page.locked_by = self.user
        if settings.USE_TZ:
            # 12:00 UTC
            self.page.locked_at = "2013-02-01T12:00:00.000Z"
            self.page.latest_revision_created_at = "2013-01-01T12:00:00.000Z"
        else:
            # 12:00 in no specific timezone
            self.page.locked_at = "2013-02-01T12:00:00"
            self.page.latest_revision_created_at = "2013-01-01T12:00:00"
        self.page.save()

        response = self.get(params={"export": "csv"})

        # Check response
        self.assertEqual(response.status_code, 200)
        data_lines = response.getvalue().decode().split("\n")
        self.assertEqual(
            data_lines[0], "Title,Updated,Status,Type,Locked at,Locked by\r"
        )
        if settings.USE_TZ:
            self.assertEqual(
                data_lines[1],
                "Root,2013-01-01 12:00:00+00:00,live,Page,2013-02-01 12:00:00+00:00,test@email.com\r",
            )
        else:
            self.assertEqual(
                data_lines[1],
                "Root,2013-01-01 12:00:00,live,Page,2013-02-01 12:00:00,test@email.com\r",
            )

    def test_xlsx_export(self):
        self.page = Page.objects.first()
        self.page.locked = True
        self.page.locked_by = self.user
        if settings.USE_TZ:
            # 12:00 UTC
            self.page.locked_at = "2013-02-01T12:00:00.000Z"
            self.page.latest_revision_created_at = "2013-01-01T12:00:00.000Z"
        else:
            # 12:00 in no specific timezone
            self.page.locked_at = "2013-02-01T12:00:00"
            self.page.latest_revision_created_at = "2013-01-01T12:00:00"
        self.page.save()

        response = self.get(params={"export": "xlsx"})

        # Check response - the locked page info should be in it
        self.assertEqual(response.status_code, 200)
        workbook_data = response.getvalue()
        worksheet = load_workbook(filename=BytesIO(workbook_data))["Sheet1"]
        cell_array = [[cell.value for cell in row] for row in worksheet.rows]
        self.assertEqual(
            cell_array[0],
            ["Title", "Updated", "Status", "Type", "Locked at", "Locked by"],
        )
        self.assertEqual(
            cell_array[1],
            [
                "Root",
                datetime.datetime(2013, 1, 1, 12, 0),
                "live",
                "Page",
                datetime.datetime(2013, 2, 1, 12, 0),
                "test@email.com",
            ],
        )
        self.assertEqual(len(cell_array), 2)

        self.assertEqual(worksheet["B2"].number_format, ExcelDateFormatter().get())
        self.assertEqual(worksheet["E2"].number_format, ExcelDateFormatter().get())


class TestFilteredLockedPagesView(BaseReportViewTestCase):
    fixtures = ["test.json"]
    url_name = "wagtailadmin_reports:locked_pages"

    def setUp(self):
        self.user = self.login()
        self.unpublished_page = Page.objects.get(
            url_path="/home/events/tentative-unpublished-event/"
        )
        self.unpublished_page.locked = True
        self.unpublished_page.locked_by = self.user
        self.unpublished_page.locked_at = timezone.now()
        self.unpublished_page.save()

        self.christmas_page = Page.objects.get(url_path="/home/events/christmas/")
        self.christmas_page.locked = True
        self.christmas_page.locked_by = self.user
        self.christmas_page.locked_at = timezone.now()
        self.christmas_page.save()

    def test_filter_by_live(self):
        response = self.get(params={"live": "true"})
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Tentative Unpublished Event")
        self.assertContains(response, "My locked page")
        self.assertContains(response, "Christmas")

        response = self.get(params={"live": "false"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Tentative Unpublished Event")
        self.assertNotContains(response, "My locked page")
        self.assertNotContains(response, "Christmas")

        soup = self.get_soup(response.content)
        self.assertActiveFilter(soup, "live", "false")

    def test_filter_by_user(self):
        response = self.get(params={"locked_by": self.user.pk})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Tentative Unpublished Event")
        self.assertContains(response, "Christmas")
        self.assertNotContains(response, "My locked page")


class TestFilteredLockedPagesResultsView(TestFilteredLockedPagesView):
    url_name = "wagtailadmin_reports:locked_pages_results"
    results_only = True


class TestFilteredLogEntriesView(BaseReportViewTestCase):
    fixtures = ["test.json"]
    url_name = "wagtailadmin_reports:site_history"

    def setUp(self):
        self.user = self.login()
        self.home_page = Page.objects.get(url_path="/home/")
        self.custom_model = Advert.objects.get(pk=1)

        self.editor = self.create_user(
            username="the_editor", email="the_editor@example.com", password="password"
        )
        editors = Group.objects.get(name="Editors")
        editors.permissions.add(
            Permission.objects.get(
                content_type__app_label="wagtailadmin", codename="access_admin"
            )
        )
        GroupPagePermission.objects.create(
            group=editors, page=self.home_page, permission_type="change"
        )
        editors.user_set.add(self.editor)

        # timezone matches TIME_ZONE = "Asia/Tokyo" in tests/settings.py
        with freeze_time("2024-05-06 12:00:00+09:00"):
            self.today = timezone.now()

        self.create_log = PageLogEntry.objects.log_action(
            self.home_page,
            "wagtail.create",
            timestamp=self.today - timezone.timedelta(days=4),
            user=self.user,
        )
        self.edit_log_1 = PageLogEntry.objects.log_action(
            self.home_page,
            "wagtail.edit",
            timestamp=self.today - timezone.timedelta(days=3),
        )
        self.edit_log_2 = PageLogEntry.objects.log_action(
            self.home_page,
            "wagtail.edit",
            timestamp=self.today - timezone.timedelta(days=2),
            user=self.editor,
        )
        self.edit_log_3 = PageLogEntry.objects.log_action(
            self.home_page,
            "wagtail.edit",
            timestamp=self.today - timezone.timedelta(days=1),
            title="The FINAL cut",
        )

        self.create_comment_log = PageLogEntry.objects.log_action(
            self.home_page,
            "wagtail.comments.create",
            data={
                "comment": {
                    "contentpath": "title",
                    "text": "Foo",
                }
            },
        )
        self.edit_comment_log = PageLogEntry.objects.log_action(
            self.home_page,
            "wagtail.comments.edit",
            data={
                "comment": {
                    "contentpath": "title",
                    "text": "Edited",
                }
            },
        )
        self.create_reply_log = PageLogEntry.objects.log_action(
            self.home_page,
            "wagtail.comments.create_reply",
            data={
                "comment": {
                    "contentpath": "title",
                    "text": "Foo",
                }
            },
        )

        self.create_custom_log = ModelLogEntry.objects.log_action(
            self.custom_model,
            "wagtail.create",
            timestamp=self.today - timezone.timedelta(days=3),
        )

        self.edit_custom_log = ModelLogEntry.objects.log_action(
            self.custom_model,
            "wagtail.edit",
            timestamp=self.today - timezone.timedelta(days=2),
            title="the final CUT",
        )

    def assert_log_entries(self, response, expected):
        actual = set(response.context["object_list"])
        self.assertSetEqual(actual, set(expected))

    def assert_filter_actions(self, response, expected):
        soup = self.get_soup(response.content)
        actual = {
            choice.get("value")
            for choice in soup.select(
                f"{self.drilldown_selector} input[name='action'][type='checkbox']"
            )
        }
        self.assertSetEqual(actual, set(expected))

    def test_unfiltered(self):
        response = self.get()
        self.assertEqual(response.status_code, 200)
        self.assertBreadcrumbs(
            [{"url": "", "label": "Site history"}],
            response.content,
        )
        self.assert_log_entries(
            response,
            [
                self.create_log,
                self.edit_log_1,
                self.edit_log_2,
                self.edit_log_3,
                self.create_comment_log,
                self.edit_comment_log,
                self.create_reply_log,
                self.create_custom_log,
                self.edit_custom_log,
            ],
        )

        self.assert_filter_actions(
            response,
            [
                "wagtail.create",
                "wagtail.edit",
                "wagtail.comments.create",
                "wagtail.comments.edit",
                "wagtail.comments.create_reply",
            ],
        )

        soup = self.get_soup(response.content)
        self.assertActiveFilterNotRendered(soup)
        self.assertPageTitle(soup, "Site history - Wagtail")

        # The editor should not see the Advert's log entries.
        self.login(user=self.editor)
        response = self.get()
        self.assertEqual(response.status_code, 200)
        self.assert_log_entries(
            response,
            [
                self.create_log,
                self.edit_log_1,
                self.edit_log_2,
                self.edit_log_3,
                self.create_comment_log,
                self.edit_comment_log,
                self.create_reply_log,
            ],
        )

        self.assert_filter_actions(
            response,
            [
                "wagtail.create",
                "wagtail.edit",
                "wagtail.comments.create",
                "wagtail.comments.edit",
                "wagtail.comments.create_reply",
            ],
        )

        soup = self.get_soup(response.content)
        self.assertActiveFilterNotRendered(soup)

    def test_filter_by_action(self):
        response = self.get(params={"action": "wagtail.edit"})
        self.assertEqual(response.status_code, 200)
        self.assert_log_entries(
            response,
            [
                self.edit_log_1,
                self.edit_log_2,
                self.edit_log_3,
                self.edit_custom_log,
            ],
        )

        self.login(user=self.editor)
        response = self.get(params={"action": "wagtail.edit"})
        self.assertEqual(response.status_code, 200)
        self.assert_log_entries(
            response,
            [
                self.edit_log_1,
                self.edit_log_2,
                self.edit_log_3,
            ],
        )

        soup = self.get_soup(response.content)
        self.assertActiveFilter(soup, "action", "wagtail.edit")

    def test_filter_by_action_multiple(self):
        response = self.get(params={"action": ["wagtail.edit", "wagtail.create"]})
        self.assertEqual(response.status_code, 200)
        self.assert_log_entries(
            response,
            [
                self.create_log,
                self.create_custom_log,
                self.edit_log_1,
                self.edit_log_2,
                self.edit_log_3,
                self.edit_custom_log,
            ],
        )

        self.login(user=self.editor)
        response = self.get(params={"action": ["wagtail.edit", "wagtail.create"]})
        self.assertEqual(response.status_code, 200)
        self.assert_log_entries(
            response,
            [
                self.create_log,
                self.edit_log_1,
                self.edit_log_2,
                self.edit_log_3,
            ],
        )

    def test_filter_by_timestamp(self):
        today = self.today.date()
        response = self.get(
            params={"timestamp_from": today - timezone.timedelta(days=3)}
        )
        self.assertEqual(response.status_code, 200)
        self.assert_log_entries(
            response,
            [
                # Doesn't contain self.create_log which was created 4 days ago
                self.edit_log_1,
                self.edit_log_2,
                self.edit_log_3,
                self.create_comment_log,
                self.edit_comment_log,
                self.create_reply_log,
                self.create_custom_log,
                self.edit_custom_log,
            ],
        )

        response = self.get(params={"timestamp_to": today - timezone.timedelta(days=2)})
        self.assertEqual(response.status_code, 200)
        self.assert_log_entries(
            response,
            [
                # Doesn't contain self.edit_log_3 which was created 1 day ago,
                # as well as self.create_comment_log, self.edit_comment_log,
                # and self.create_reply_log which was created without an explicit
                # timestamp (and thus defaults to the current time)
                self.create_log,
                self.edit_log_1,
                self.edit_log_2,
                self.create_custom_log,
                self.edit_custom_log,
            ],
        )

        response = self.get(
            params={
                "timestamp_from": today - timezone.timedelta(days=3),
                "timestamp_to": today - timezone.timedelta(days=2),
            }
        )
        self.assertEqual(response.status_code, 200)
        self.assert_log_entries(
            response,
            [
                # Doesn't contain
                # self.create_log which was created 4 days ago,
                # self.edit_log_3 which was created 1 day ago,
                # as well as self.create_comment_log, self.edit_comment_log,
                # and self.create_reply_log which was created without an explicit
                # timestamp (and thus defaults to the current time)
                self.edit_log_1,
                self.edit_log_2,
                self.create_custom_log,
                self.edit_custom_log,
            ],
        )

    def test_filter_by_user(self):
        response = self.get(params={"user": self.editor.pk})
        self.assertEqual(response.status_code, 200)
        self.assert_log_entries(response, [self.edit_log_2])

        response = self.get(params={"user": [self.user.pk, self.editor.pk]})
        self.assertEqual(response.status_code, 200)
        self.assert_log_entries(response, [self.create_log, self.edit_log_2])

    def test_filter_by_label(self):
        response = self.get(params={"label": "final cut"})
        self.assertEqual(response.status_code, 200)
        self.assert_log_entries(response, [self.edit_log_3, self.edit_custom_log])

    def test_filter_by_object_type(self):
        response = self.get(
            params={"object_type": ContentType.objects.get_for_model(Page).pk}
        )
        self.assertEqual(response.status_code, 200)
        self.assert_log_entries(
            response,
            [
                self.create_log,
                self.edit_log_1,
                self.edit_log_2,
                self.edit_log_3,
                self.create_comment_log,
                self.edit_comment_log,
                self.create_reply_log,
            ],
        )

        response = self.get(
            params={"object_type": ContentType.objects.get_for_model(Advert).pk}
        )
        self.assertEqual(response.status_code, 200)
        self.assert_log_entries(
            response,
            [
                self.create_custom_log,
                self.edit_custom_log,
            ],
        )

    def test_is_commenting_action(self):
        response = self.get(params={"is_commenting_action": "false"})
        self.assertEqual(response.status_code, 200)
        self.assert_log_entries(
            response,
            [
                self.create_log,
                self.edit_log_1,
                self.edit_log_2,
                self.edit_log_3,
                self.create_custom_log,
                self.edit_custom_log,
            ],
        )
        soup = self.get_soup(response.content)
        self.assertActiveFilter(soup, "is_commenting_action", "false")

        response = self.get(params={"is_commenting_action": "true"})
        self.assertEqual(response.status_code, 200)
        self.assert_log_entries(
            response,
            [
                self.create_comment_log,
                self.edit_comment_log,
                self.create_reply_log,
            ],
        )
        soup = self.get_soup(response.content)
        self.assertActiveFilter(soup, "is_commenting_action", "true")

        self.login(user=self.editor)
        response = self.get(params={"is_commenting_action": "false"})
        self.assertEqual(response.status_code, 200)
        self.assert_log_entries(
            response,
            [
                self.create_log,
                self.edit_log_1,
                self.edit_log_2,
                self.edit_log_3,
            ],
        )
        soup = self.get_soup(response.content)
        self.assertActiveFilter(soup, "is_commenting_action", "false")

        response = self.get(params={"is_commenting_action": "true"})
        self.assertEqual(response.status_code, 200)
        self.assert_log_entries(
            response,
            [
                self.create_comment_log,
                self.edit_comment_log,
                self.create_reply_log,
            ],
        )
        soup = self.get_soup(response.content)
        self.assertActiveFilter(soup, "is_commenting_action", "true")

    def test_log_entry_with_stale_content_type(self):
        stale_content_type = ContentType.objects.create(
            app_label="fake_app", model="deleted model"
        )

        ModelLogEntry.objects.create(
            object_id=123,
            content_type=stale_content_type,
            label="This instance's model was deleted, but its content type was not",
            action="wagtail.create",
            timestamp=timezone.now(),
        )

        response = self.get()
        self.assertContains(response, "Deleted model")

    def test_log_entry_with_null_content_type(self):
        ModelLogEntry.objects.create(
            object_id=123,
            content_type=None,
            label="This instance's model was deleted, and so was its content type",
            action="wagtail.create",
            timestamp=timezone.now(),
        )

        response = self.get()
        self.assertContains(response, "Unknown content type")

    def test_decorated_queryset(self):
        # Ensure that decorate_paginated_queryset is only called with the queryset for the current
        # page, instead of all objects over all pages.
        with (
            mock.patch.object(
                LogEntriesView,
                "decorate_paginated_queryset",
                side_effect=LogEntriesView.decorate_paginated_queryset,
                autospec=True,
            ) as decorate_paginated_queryset,
            mock.patch.object(LogEntriesView, "paginate_by", return_value=1),
        ):
            response = self.get()
            decorate_paginated_queryset.assert_called_once()
            queryset = decorate_paginated_queryset.call_args.args[1]
            self.assertEqual(queryset.count(), 1)

        self.assertEqual(response.status_code, 200)


class TestFilteredLogEntriesResultsView(TestFilteredLogEntriesView):
    url_name = "wagtailadmin_reports:site_history_results"
    results_only = True


@override_settings(
    USE_L10N=True,
)
class TestExcelDateFormatter(TestCase):
    def test_all_locales(self):
        formatter = ExcelDateFormatter()

        for lang in LANG_INFO.keys():
            with self.subTest(lang), translation.override(lang):
                self.assertNotEqual(formatter.get(), "")

    def test_format(self):
        formatter = ExcelDateFormatter()

        with self.subTest(format="r"):
            # Format code for RFC 5322 formatted date, e.g. 'Thu, 21 Dec 2000 16:01:07'
            self.assertEqual(formatter.format("r"), "ddd, d mmm yyyy hh:mm:ss")

        with self.subTest(format="m/d/Y g:i A"):
            # Format code for e.g. '12/21/2000 4:01 PM'
            self.assertEqual(formatter.format("m/d/Y g:i A"), "mm/dd/yyyy h:mm AM/PM")


class TestAgingPagesView(BaseReportViewTestCase):
    url_name = "wagtailadmin_reports:aging_pages"

    def setUp(self):
        self.user = self.login()
        self.root = Page.objects.first()
        self.home = Page.objects.get(slug="home")

    def publish_home_page(self):
        self.home.save_revision().publish(user=self.user)

    def test_simple(self):
        response = self.get()
        self.assertEqual(response.status_code, 200)
        self.assertTemplateNotUsed(
            response,
            "wagtailadmin/reports/base_page_report.html",
        )
        self.assertTemplateUsed(response, "wagtailadmin/reports/base_report.html")
        self.assertTemplateUsed(
            response,
            "wagtailadmin/reports/aging_pages_results.html",
        )
        self.assertBreadcrumbs(
            [{"url": "", "label": "Aging pages"}],
            response.content,
        )
        soup = self.get_soup(response.content)
        self.assertActiveFilterNotRendered(soup)
        self.assertPageTitle(soup, "Aging pages - Wagtail")

    def test_displays_only_published_pages(self):
        response = self.get()
        self.assertContains(response, "No pages found.")

        self.publish_home_page()
        response = self.get()

        # Home Page should be listed
        self.assertContains(response, self.home.title)
        # Last published by user is set
        self.assertContains(response, self.user.get_username())

        self.assertNotContains(response, self.root.title)
        self.assertNotContains(response, "No pages found.")

    def test_permissions(self):
        # Publish home page
        self.publish_home_page()

        # Remove privileges from user
        self.user.is_superuser = False
        self.user.user_permissions.add(
            Permission.objects.get(
                content_type__app_label="wagtailadmin", codename="access_admin"
            )
        )
        self.user.save()

        response = self.get()
        self.assertEqual(response.status_code, 302)

    def test_csv_export(self):
        self.publish_home_page()
        if settings.USE_TZ:
            self.home.last_published_at = "2013-01-01T12:00:00.000Z"
        else:
            self.home.last_published_at = "2013-01-01T12:00:00"
        self.home.save()

        response = self.get(params={"export": "csv"})
        self.assertEqual(response.status_code, 200)

        data_lines = response.getvalue().decode().split("\n")
        self.assertEqual(
            data_lines[0], "Title,Status,Last published at,Last published by,Type\r"
        )
        if settings.USE_TZ:
            self.assertEqual(
                data_lines[1],
                "Welcome to your new Wagtail site!,live + draft,2013-01-01 12:00:00+00:00,test@email.com,Page\r",
            )
        else:
            self.assertEqual(
                data_lines[1],
                "Welcome to your new Wagtail site!,live + draft,2013-01-01 12:00:00,test@email.com,Page\r",
            )

    def test_xlsx_export(self):
        self.publish_home_page()

        if settings.USE_TZ:
            self.home.last_published_at = "2013-01-01T12:00:00.000Z"
        else:
            self.home.last_published_at = "2013-01-01T12:00:00"
        self.home.save()

        response = self.get(params={"export": "xlsx"})
        self.assertEqual(response.status_code, 200)

        workbook_data = response.getvalue()
        worksheet = load_workbook(filename=BytesIO(workbook_data))["Sheet1"]
        cell_array = [[cell.value for cell in row] for row in worksheet.rows]

        self.assertEqual(
            cell_array[0],
            ["Title", "Status", "Last published at", "Last published by", "Type"],
        )
        self.assertEqual(
            cell_array[1],
            [
                "Welcome to your new Wagtail site!",
                "live + draft",
                datetime.datetime(2013, 1, 1, 12, 0),
                "test@email.com",
                "Page",
            ],
        )
        self.assertEqual(len(cell_array), 2)

        self.assertEqual(worksheet["C2"].number_format, ExcelDateFormatter().get())

    def test_xlsx_export_without_published_by(self):
        """
        Test that the xlsx export works when a page has no 'published_by' set.
        See https://github.com/wagtail/wagtail/issues/10821
        """

        self.home.save_revision().publish()

        if settings.USE_TZ:
            self.home.last_published_at = "2013-01-01T12:00:00.000Z"
        else:
            self.home.last_published_at = "2013-01-01T12:00:00"

        # mimic a page that does not have a 'published_by' on creation
        self.home.last_published_by = None
        self.home.save()

        response = self.get(params={"export": "xlsx"})
        self.assertEqual(response.status_code, 200)

        workbook_data = response.getvalue()
        worksheet = load_workbook(filename=BytesIO(workbook_data))["Sheet1"]
        cell_array = [[cell.value for cell in row] for row in worksheet.rows]

        self.assertEqual(
            cell_array[0],
            ["Title", "Status", "Last published at", "Last published by", "Type"],
        )
        self.assertEqual(
            cell_array[1],
            [
                "Welcome to your new Wagtail site!",
                "live + draft",
                datetime.datetime(2013, 1, 1, 12, 0),
                None,
                "Page",
            ],
        )
        self.assertEqual(len(cell_array), 2)

        self.assertEqual(worksheet["C2"].number_format, ExcelDateFormatter().get())

    def test_report_renders_when_page_publisher_deleted(self):
        temp_user = self.create_superuser(
            "temp", email="temp@user.com", password="tempuser"
        )
        expected_deleted_string = f"user {temp_user.pk} (deleted)"

        self.home.save_revision().publish(user=temp_user)
        temp_user.delete()

        response = self.get()
        self.assertContains(response, expected_deleted_string)


class TestAgingPagesViewPermissions(BaseReportViewTestCase):
    url_name = "wagtailadmin_reports:aging_pages"

    def setUp(self):
        self.user = self.login()

    def test_simple(self):
        response = self.get()
        self.assertEqual(response.status_code, 200)

    def test_get_with_no_permission(self):
        group = Group.objects.create(name="test group")
        self.user.is_superuser = False
        self.user.save()
        self.user.groups.add(group)
        self.user.user_permissions.add(
            Permission.objects.get(
                content_type__app_label="wagtailadmin", codename="access_admin"
            )
        )
        # No GroupPagePermission created

        response = self.get()
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("wagtailadmin_home"))

    def test_get_with_minimal_permissions(self):
        group = Group.objects.create(name="test group")
        self.user.is_superuser = False
        self.user.save()
        self.user.groups.add(group)
        self.user.user_permissions.add(
            Permission.objects.get(
                content_type__app_label="wagtailadmin", codename="access_admin"
            )
        )
        GroupPagePermission.objects.create(
            group=group,
            page=Page.objects.first(),
            permission_type="add",
        )

        response = self.get()

        self.assertEqual(response.status_code, 200)


class TestFilteredAgingPagesView(BaseReportViewTestCase):
    fixtures = ["test.json"]
    url_name = "wagtailadmin_reports:aging_pages"

    def setUp(self):
        self.user = self.login()
        self.home_page = Page.objects.get(slug="home")
        self.aboutus_page = Page.objects.get(slug="about-us")

    def test_filter_by_live(self):
        response = self.get(params={"live": "true"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.home_page.title)
        self.assertContains(response, self.aboutus_page.title)

        response = self.get(params={"live": "false"})
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, self.home_page.title)
        self.assertNotContains(response, self.aboutus_page.title)

    def test_filter_by_content_type(self):
        ct_pk = self.aboutus_page.specific.content_type.pk
        response = self.get(params={"content_type": ct_pk})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.aboutus_page.title)
        self.assertNotContains(response, self.home_page.title)

        soup = self.get_soup(response.content)
        self.assertActiveFilter(soup, "content_type", ct_pk)

        # Should render the filter inside the drilldown component
        ct_select = soup.select_one(
            f"{self.drilldown_selector} select[name='content_type']"
        )
        self.assertIsNotNone(ct_select)
        selected_option = ct_select.select_one("option[selected]")
        self.assertIsNotNone(selected_option)
        self.assertEqual(selected_option.get("value"), str(ct_pk))

    def test_filter_by_last_published_at(self):
        self.home_page.last_published_at = timezone.now()
        self.home_page.save()

        response = self.get(params={"last_published_at": "2015-01-01"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.aboutus_page.title)
        self.assertNotContains(response, self.home_page.title)


class TestFilteredAgingPagesResultsView(TestFilteredAgingPagesView):
    url_name = "wagtailadmin_reports:aging_pages_results"
    results_only = True


class PageTypesUsageReportViewTest(BaseReportViewTestCase):
    fixtures = ["test.json"]
    url_name = "wagtailadmin_reports:page_types_usage"

    def setUp(self):
        self.user = self.login()

    @staticmethod
    def display_name(content_type):
        return f"{content_type.app_label}.{content_type.model}"

    def test_simple(self):
        response = self.get()
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "wagtailadmin/reports/base_report.html")
        self.assertTemplateUsed(
            response,
            "wagtailadmin/reports/page_types_usage_results.html",
        )
        self.assertBreadcrumbs(
            [{"url": "", "label": "Page types usage"}],
            response.content,
        )
        soup = self.get_soup(response.content)
        self.assertActiveFilterNotRendered(soup)
        self.assertPageTitle(soup, "Page types usage - Wagtail")

    def test_displays_only_page_types(self):
        """Asserts that the correct models are included in the queryset."""
        response = self.get()
        # Assert that the response contains page models:
        event_page_content_type = ContentType.objects.get_for_model(EventPage)
        event_page_content_type_full_name = self.display_name(event_page_content_type)
        self.assertContains(response, event_page_content_type_full_name)
        simple_page_content_type = ContentType.objects.get_for_model(SimplePage)
        simple_page_content_type_full_name = self.display_name(simple_page_content_type)
        self.assertContains(response, simple_page_content_type_full_name)
        # But it should not contain non-page models:
        event_page_speaker_content_type = ContentType.objects.get_for_model(
            EventPageSpeaker
        )
        event_page_speaker_content_type_full_name = self.display_name(
            event_page_speaker_content_type
        )
        self.assertNotContains(response, event_page_speaker_content_type_full_name)

    def test_displays_wagtailcore_page_if_has_instances(self):
        """Asserts that the wagtailcore.Page model is included in the queryset if it has instances."""
        page_content_type = ContentType.objects.get_for_model(Page)
        page_content_type_full_name = self.display_name(page_content_type)

        # Start with no pages:
        Page.objects.filter(depth__gt=1, content_type=page_content_type).delete()

        # There aren't any Page instances and it is not creatable, it shouldn't be included in the report
        response = self.get()
        self.assertNotContains(response, page_content_type_full_name)

        # Create a page:
        page = Page(title="Page")
        Page.get_first_root_node().add_child(instance=page)

        # There is a Page now, so report should include the `Page` ContentType
        response = self.get()
        self.assertContains(response, page_content_type_full_name)


class PageTypesUsageReportViewQuerysetTests(WagtailTestUtils, TestCase):
    def setUp(self):
        super().setUp()
        self.view = page_types_usage.PageTypesUsageReportView()
        self.view.request = RequestFactory().get(
            reverse("wagtailadmin_reports:page_types_usage")
        )
        self.root = Page.objects.first()
        self.home = Page.objects.get(slug="home")

        self.simple_page_a = SimplePage(title="Simple page A", content="hello")
        self.simple_page_b = SimplePage(title="Simple page B", content="hello")
        self.simple_page_c = SimplePage(title="Simple page C", content="hello")

        Page.get_first_root_node().add_child(instance=self.simple_page_a)
        Page.get_first_root_node().add_child(instance=self.simple_page_b)
        Page.get_first_root_node().add_child(instance=self.simple_page_c)
        self.simple_page_a.save_revision().publish()
        self.simple_page_b.save_revision().publish()
        self.simple_page_c.save_revision().publish()

        self.event_page = EventPage(
            title="Event Page",
            audience="public",
            location="foo",
            cost="bar",
            date_from=datetime.date.today(),
        )
        Page.get_first_root_node().add_child(instance=self.event_page)

    def test_queryset_ordering(self):
        """Asserts that the queryset is ordered by page model count."""
        # Test that the queryset is correctly ordered by page model count
        queryset = self.view.get_queryset()
        # Convert queryset to list of ids to make it easier to test
        queryset_list_pks = list(queryset.values_list("pk", flat=True))
        # Get the positions of the content types in the list
        simple_page_content_type = ContentType.objects.get_for_model(SimplePage)
        simple_page_position = queryset_list_pks.index(simple_page_content_type.pk)
        event_page_content_type = ContentType.objects.get_for_model(EventPage)
        event_page_position = queryset_list_pks.index(event_page_content_type.pk)
        # Assert that the SimplePage comes before EventPage, since it has more entries
        self.assertTrue(simple_page_position < event_page_position)
        # There should be 3 SimplePages
        self.assertEqual(queryset.get(id=simple_page_content_type.pk).count, 3)
        # There should be 1 EventPage
        self.assertEqual(queryset.get(id=event_page_content_type.pk).count, 1)

    def test_queryset_last_edited_page(self):
        """Tests that the queryset correctly returns the last edited page."""
        # Edit the first simple page
        revision = self.simple_page_a.save_revision()
        revision.publish()
        # Edit the second simple page
        revision = self.simple_page_b.save_revision()
        revision.publish()
        # Edit the third simple page
        revision = self.simple_page_c.save_revision()
        revision.publish()
        # Re-edit the first simple page
        revision = self.simple_page_a.save_revision()
        revision.publish()
        # Get the queryset:
        queryset = self.view.decorate_paginated_queryset(self.view.get_queryset())
        # Assert that the first simple page is the last edited page
        self.simple_page_a.refresh_from_db()
        self.assertEqual(queryset[0].last_edited_page.specific, self.simple_page_a)


@override_settings(LANGUAGE_CODE="en", WAGTAIL_I18N_ENABLED=True)
class PageTypesReportFiltersTests(BaseReportViewTestCase):
    url_name = "wagtailadmin_reports:page_types_usage"

    def setUp(self):
        self.user = self.login()
        self.default_locale = Locale.get_default()
        self.fr_locale, _ = Locale.objects.get_or_create(language_code="fr")

    def test_locale_filtering(self):
        # Create pages in default locale
        event_page = EventPage(
            title="Event Page",
            audience="public",
            location="foo",
            cost="bar",
            date_from=datetime.date.today(),
        )
        simple_page = SimplePage(title="Simple Page", content="hello")
        Page.get_first_root_node().add_child(instance=event_page)
        Page.get_first_root_node().add_child(instance=simple_page)
        event_page.save_revision().publish()
        simple_page.save_revision().publish()
        # Translate pages to French
        event_page.copy_for_translation(self.fr_locale)
        simple_page.copy_for_translation(self.fr_locale)

        # Edit the simple page in English to make sure that it's the latest
        simple_page.title = "Updated Simple Page English title"
        revision = simple_page.save_revision()
        simple_page.publish(revision)

        response = self.get()
        page_types = {
            content_type.id: content_type
            for content_type in response.context["object_list"]
        }

        event_page_row = page_types.get(ContentType.objects.get_for_model(EventPage).pk)
        simple_page_row = page_types.get(
            ContentType.objects.get_for_model(SimplePage).pk
        )

        self.assertEqual(event_page_row.count, 2)
        self.assertEqual(simple_page_row.count, 2)
        # The last edited page should be the French version
        self.assertEqual(event_page_row.last_edited_page.locale, self.fr_locale)
        # The last edited SimplePage should be the English version
        self.assertEqual(simple_page_row.last_edited_page.locale, self.default_locale)

        # Filter by French locale
        response = self.get({"page_locale": self.fr_locale.language_code})
        page_types = {
            content_type.id: content_type
            for content_type in response.context["object_list"]
        }

        event_page_row = page_types.get(ContentType.objects.get_for_model(EventPage).pk)
        simple_page_row = page_types.get(
            ContentType.objects.get_for_model(SimplePage).pk
        )

        # There should be 1 of each page (only the French locale ones)
        self.assertEqual(event_page_row.count, 1)
        self.assertEqual(simple_page_row.count, 1)
        # The last edited page should be the French version (even though page was later edited in English)
        self.assertEqual(event_page_row.last_edited_page.locale, self.fr_locale)
        self.assertEqual(simple_page_row.last_edited_page.locale, self.fr_locale)

        # Should render the filter inside the drilldown component
        soup = self.get_soup(response.content)
        locale_select = soup.select_one(
            f"{self.drilldown_selector} select[name='page_locale']"
        )
        self.assertIsNotNone(locale_select)
        selected_option = locale_select.select_one("option[selected]")
        self.assertIsNotNone(selected_option)
        self.assertEqual(selected_option.get("value"), "fr")

    def test_site_filtering_with_single_site(self):
        """Asserts that the site filter is not displayed when there is only one site."""
        sites = Site.objects.all()
        self.assertEqual(sites.count(), 1)

        response = self.get()
        filterset = response.context["filters"]

        # Assert that the filterset does not have the site field
        self.assertNotIn("site", filterset.form.fields)
        self.assertNotIn("site", filterset.filters.keys())
        self.assertFalse(filterset.sites_filter_enabled)

    def test_site_filtering_with_multiple_sites(self):
        root_page = Page.get_first_root_node()
        # Create pages in default locale
        event_page = EventPage(
            title="Event Page",
            audience="public",
            location="foo",
            cost="bar",
            date_from=datetime.date.today(),
        )
        simple_page = SimplePage(title="Simple Page", content="hello")
        root_page.add_child(instance=event_page)
        root_page.add_child(instance=simple_page)

        # Create a new site and add the pages to it
        simple_page_site = Site.objects.create(
            hostname="example.com", root_page=simple_page, is_default_site=False
        )
        self.assertEqual(Site.objects.count(), 2)

        response = self.get()
        page_types = {
            content_type.id: content_type
            for content_type in response.context["object_list"]
        }

        event_page_row = page_types.get(ContentType.objects.get_for_model(EventPage).pk)
        simple_page_row = page_types.get(
            ContentType.objects.get_for_model(SimplePage).pk
        )

        self.assertEqual(event_page_row.count, 1)
        self.assertEqual(simple_page_row.count, 1)

        # Filter by the simple_page_site
        response = self.get({"site": simple_page_site.root_page.path})
        page_types = {
            content_type.id: content_type
            for content_type in response.context["object_list"]
        }

        simple_page_row = page_types.get(
            ContentType.objects.get_for_model(SimplePage).pk
        )

        # There should be 1 SimplePage
        self.assertEqual(simple_page_row.count, 1)
        # There shouldn't be a regular Page
        self.assertFalse(ContentType.objects.get_for_model(EventPage).pk in page_types)

    @override_settings(
        WAGTAIL_CONTENT_LANGUAGES=[
            ("en", "English"),
            ("de", "German"),
            ("fr", "French"),
        ],
    )
    def test_get_locale_choices(self):
        choices = page_types_usage._get_locale_choices()

        expected_choices = [
            ("en", "English"),
            ("de", "German"),
            ("fr", "French"),
        ]

        self.assertCountEqual(choices, expected_choices)


class PageTypesReportFiltersResultsTests(PageTypesReportFiltersTests):
    url_name = "wagtailadmin_reports:page_types_usage_results"
    results_only = True


class TestPageTypesUsageReportViewPermissions(BaseReportViewTestCase):
    fixtures = ["test.json"]
    url_name = "wagtailadmin_reports:page_types_usage"

    def test_simple(self):
        response = self.get()
        self.assertEqual(response.status_code, 200)

    def test_get_with_no_permission(self):
        group = Group.objects.create(name="test group")
        self.user.is_superuser = False
        self.user.save()
        self.user.groups.add(group)
        self.user.user_permissions.add(
            Permission.objects.get(
                content_type__app_label="wagtailadmin", codename="access_admin"
            )
        )
        # No GroupPagePermission created

        response = self.get()
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("wagtailadmin_home"))

    def test_get_with_minimal_permissions(self):
        group = Group.objects.create(name="test group")
        self.user.is_superuser = False
        self.user.save()
        self.user.groups.add(group)
        self.user.user_permissions.add(
            Permission.objects.get(
                content_type__app_label="wagtailadmin", codename="access_admin"
            )
        )
        GroupPagePermission.objects.create(
            group=group,
            page=Page.objects.first(),
            permission_type="add",
        )

        response = self.get()

        self.assertEqual(response.status_code, 200)

    def test_get_with_page_specific_permissions(self):
        group = Group.objects.create(name="test group")
        self.user.is_superuser = False
        self.user.save()
        self.user.groups.add(group)
        self.user.user_permissions.add(
            Permission.objects.get(
                content_type__app_label="wagtailadmin", codename="access_admin"
            )
        )
        latest_edited_event_page = EventPage.objects.order_by(
            F("latest_revision_created_at").desc(nulls_last=True), "title", "-pk"
        ).first()
        latest_edited_simple_page = SimplePage.objects.order_by(
            F("latest_revision_created_at").desc(nulls_last=True), "title", "-pk"
        ).first()
        GroupPagePermission.objects.create(
            group=group,
            page=latest_edited_event_page,
            permission_type="change",
        )

        response = self.get()

        self.assertEqual(response.status_code, 200)
        # For pages that the user can edit, it should show the page title and link to edit the page:
        edit_event_page_url = reverse(
            "wagtailadmin_pages:edit", args=(latest_edited_event_page.id,)
        )
        self.assertContains(
            response,
            f"<a href={edit_event_page_url}>{latest_edited_event_page.get_admin_display_title()}</a>",
            html=True,
        )
        # For pages that the user cannot edit, it should only show the page title
        self.assertContains(
            response,
            f"<p>{latest_edited_simple_page.get_admin_display_title()}</p>",
            html=True,
        )
        edit_simple_page_url = reverse(
            "wagtailadmin_pages:edit", args=(latest_edited_simple_page.id,)
        )
        self.assertNotContains(
            response,
            f"<a href={edit_simple_page_url}>{latest_edited_simple_page.get_admin_display_title()}</a>",
            html=True,
        )
