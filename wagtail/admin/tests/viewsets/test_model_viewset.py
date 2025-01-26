import datetime
from io import BytesIO

from django.conf import settings
from django.contrib.admin.utils import quote
from django.contrib.auth import get_permission_codename
from django.contrib.auth.models import Permission
from django.contrib.contenttypes.models import ContentType
from django.test import TestCase
from django.urls import NoReverseMatch, reverse
from django.utils.formats import date_format, localize
from django.utils.html import escape
from django.utils.timezone import make_aware
from openpyxl import load_workbook

from wagtail import hooks
from wagtail.admin.admin_url_finder import AdminURLFinder
from wagtail.log_actions import log
from wagtail.models import ModelLogEntry
from wagtail.test.testapp.models import (
    FeatureCompleteToy,
    JSONStreamModel,
    SearchTestModel,
    VariousOnDeleteModel,
)
from wagtail.test.utils.template_tests import AdminTemplateTestUtils
from wagtail.test.utils.wagtail_tests import WagtailTestUtils
from wagtail.utils.deprecation import RemovedInWagtail70Warning


class TestModelViewSetGroup(WagtailTestUtils, TestCase):
    def setUp(self):
        self.user = self.login()

    def test_menu_items(self):
        response = self.client.get(reverse("wagtailadmin_home"))
        self.assertEqual(response.status_code, 200)
        # Menu label falls back to the title-cased app label
        self.assertContains(
            response,
            '"name": "tests", "label": "Tests", "icon_name": "folder-open-inverse"',
        )
        # Capitalized-first from verbose_name_plural
        self.assertContains(response, "JSON stream models")
        self.assertContains(response, reverse("streammodel:index"))
        self.assertEqual(reverse("streammodel:index"), "/admin/streammodel/")
        # Set on class
        self.assertContains(response, "JSON MinMaxCount StreamModel")
        self.assertContains(response, reverse("minmaxcount_streammodel:index"))
        self.assertEqual(
            reverse("minmaxcount_streammodel:index"),
            "/admin/minmaxcount-streammodel/",
        )
        # Set on instance
        self.assertContains(response, "JSON BlockCounts StreamModel")
        self.assertContains(response, reverse("blockcounts_streammodel:index"))
        self.assertEqual(
            reverse("blockcounts_streammodel:index"),
            "/admin/blockcounts/streammodel/",
        )

    def test_menu_item_with_only_view_permission(self):
        self.user.is_superuser = False
        self.user.save()
        admin_permission = Permission.objects.get(
            content_type__app_label="wagtailadmin",
            codename="access_admin",
        )
        view_permission = Permission.objects.get(
            content_type=ContentType.objects.get_for_model(JSONStreamModel),
            codename=get_permission_codename("view", JSONStreamModel._meta),
        )
        self.user.user_permissions.add(admin_permission, view_permission)

        response = self.client.get(reverse("wagtailadmin_home"))

        # The group menu item is still shown
        self.assertContains(
            response,
            '"name": "tests", "label": "Tests", "icon_name": "folder-open-inverse"',
        )

        # The menu item for the model is shown
        self.assertContains(response, "JSON stream models")
        self.assertContains(response, reverse("streammodel:index"))
        self.assertEqual(reverse("streammodel:index"), "/admin/streammodel/")

        # The other items in the group are not shown as the user doesn't have permission
        self.assertNotContains(response, "JSON MinMaxCount StreamModel")
        self.assertNotContains(response, reverse("minmaxcount_streammodel:index"))
        self.assertNotContains(response, "JSON BlockCounts StreamModel")
        self.assertNotContains(response, reverse("blockcounts_streammodel:index"))


class TestTemplateConfiguration(WagtailTestUtils, TestCase):
    def setUp(self):
        self.user = self.login()

    @classmethod
    def setUpTestData(cls):
        cls.default = JSONStreamModel.objects.create(
            body='[{"type": "text", "value": "foo"}]',
        )
        cls.custom = FeatureCompleteToy.objects.create(name="Test Toy")

    def get_default_url(self, view_name, args=()):
        return reverse(f"streammodel:{view_name}", args=args)

    def get_custom_url(self, view_name, args=()):
        return reverse(f"feature_complete_toy:{view_name}", args=args)

    def test_default_templates(self):
        pk = quote(self.default.pk)
        cases = {
            "index": (
                [],
                "wagtailadmin/generic/index.html",
            ),
            "index_results": (
                [],
                "wagtailadmin/generic/listing_results.html",
            ),
            "add": (
                [],
                "wagtailadmin/generic/create.html",
            ),
            "edit": (
                [pk],
                "wagtailadmin/generic/edit.html",
            ),
            "delete": (
                [pk],
                "wagtailadmin/generic/confirm_delete.html",
            ),
        }
        for view_name, (args, template_name) in cases.items():
            with self.subTest(view_name=view_name):
                response = self.client.get(self.get_default_url(view_name, args=args))
                self.assertTemplateUsed(response, template_name)

    def test_custom_template_lookups(self):
        pk = quote(self.custom.pk)
        cases = {
            "override with index_template_name": (
                "index",
                [],
                "tests/fctoy_index.html",
            ),
            "with app label and model name": (
                "add",
                [],
                "customprefix/tests/featurecompletetoy/create.html",
            ),
            "with app label": (
                "edit",
                [pk],
                "customprefix/tests/edit.html",
            ),
            "without app label and model name": (
                "delete",
                [pk],
                "customprefix/confirm_delete.html",
            ),
        }
        for case, (view_name, args, template_name) in cases.items():
            with self.subTest(case=case):
                response = self.client.get(self.get_custom_url(view_name, args=args))
                self.assertTemplateUsed(response, template_name)
                self.assertContains(
                    response, "<p>Some extra custom content</p>", html=True
                )

    def test_wagtail_admin_template_mixin_variables_with_legacy_header(self):
        pk = quote(self.custom.pk)
        cases = {
            "delete": ([pk], "Delete", str(self.custom)),
        }
        for view_name, (args, title, subtitle) in cases.items():
            with self.subTest(view_name=view_name):
                response = self.client.get(self.get_custom_url(view_name, args=args))
                soup = self.get_soup(response.content)
                h1 = soup.select_one("h1")
                self.assertIsNotNone(h1)
                self.assertEqual(
                    "".join(h1.find_all(string=True, recursive=False)).strip(), title
                )
                subtitle_el = h1.select_one("span")
                if subtitle:
                    self.assertIsNotNone(subtitle_el)
                    self.assertEqual(subtitle_el.string, subtitle)
                else:
                    self.assertIsNone(subtitle_el)
                icon = h1.select_one("svg use[href='#icon-media']")
                self.assertIsNotNone(icon)

    def test_wagtail_admin_template_mixin_variables(self):
        pk = quote(self.custom.pk)
        cases = {
            "index": ([], "Feature complete toys", None),
            "add": ([], "New", "Feature complete toy"),
            "edit": ([pk], "Editing", str(self.custom)),
        }
        for view_name, (args, title, subtitle) in cases.items():
            with self.subTest(view_name=view_name):
                response = self.client.get(self.get_custom_url(view_name, args=args))
                soup = self.get_soup(response.content)
                h1 = soup.select_one("h1")
                expected_h1 = title
                if subtitle:
                    expected_h1 = f"{title}: {subtitle}"
                self.assertIsNotNone(h1)
                self.assertEqual(h1.get_text(strip=True), expected_h1)
                icon = h1.select_one("svg use[href='#icon-media']")
                # Icon is no longer rendered in the h1 with the slim header in place
                self.assertIsNone(icon)


class TestCustomColumns(WagtailTestUtils, TestCase):
    def setUp(self):
        self.user = self.login()

    @classmethod
    def setUpTestData(cls):
        FeatureCompleteToy.objects.create(name="Racecar")
        FeatureCompleteToy.objects.create(name="level")
        FeatureCompleteToy.objects.create(name="Lotso")

    def test_list_display(self):
        index_url = reverse("feature_complete_toy:index")
        response = self.client.get(index_url)
        # "name" column
        self.assertContains(response, "Racecar")
        self.assertContains(response, "level")
        self.assertContains(response, "Lotso")
        # BooleanColumn("is_cool")
        soup = self.get_soup(response.content)

        help = soup.select_one("td:has(svg.icon-help)")
        self.assertIsNotNone(help)
        self.assertEqual(help.text.strip(), "None")

        success = soup.select_one("td:has(svg.icon-success.w-text-positive-100)")
        self.assertIsNotNone(success)
        self.assertEqual(success.text.strip(), "True")

        error = soup.select_one("td:has(svg.icon-error.w-text-critical-100)")
        self.assertIsNotNone(error)
        self.assertEqual(error.text.strip(), "False")

        updated_at = soup.select("th a")[-1]
        self.assertEqual(updated_at.text.strip(), "Updated")
        self.assertEqual(updated_at["href"], f"{index_url}?ordering=_updated_at")


class TestListFilter(WagtailTestUtils, TestCase):
    cases = {
        "list": ("feature_complete_toy", "release_date", "Release date"),
        "dict": ("fctoy_alt1", "name__icontains", "Name contains"),
        "filterset_class": (
            "fctoy-alt2",
            "release_date__year__lte",
            "Release date year is less than or equal to",
        ),
    }

    def setUp(self):
        self.user = self.login()

    def get(self, url_namespace, params=None):
        return self.client.get(reverse(f"{url_namespace}:index"), params)

    @classmethod
    def setUpTestData(cls):
        FeatureCompleteToy.objects.create(
            name="Buzz Lightyear",
            release_date=datetime.date(1995, 11, 19),
        )
        FeatureCompleteToy.objects.create(
            name="Forky",
            release_date=datetime.date(2019, 6, 11),
        )

    def test_unfiltered_no_results(self):
        FeatureCompleteToy.objects.all().delete()
        for case, (url_namespace, lookup, label_text) in self.cases.items():
            with self.subTest(case=case):
                response = self.get(url_namespace)
                self.assertContains(
                    response,
                    "There are no feature complete toys to display",
                )
                self.assertNotContains(
                    response,
                    "No feature complete toys match your query",
                )
                self.assertNotContains(response, "Buzz Lightyear")
                self.assertNotContains(response, "Forky")

                soup = self.get_soup(response.content)
                label = soup.select_one(f"label#id_{lookup}-label")
                self.assertIsNotNone(label)
                self.assertEqual(label.string.strip(), label_text)
                input = soup.select_one(f"input#id_{lookup}")
                self.assertIsNotNone(input)

    def test_unfiltered_with_results(self):
        for case, (url_namespace, lookup, label_text) in self.cases.items():
            with self.subTest(case=case):
                response = self.get(url_namespace)
                self.assertContains(response, "Buzz Lightyear")
                self.assertContains(response, "Forky")
                self.assertNotContains(response, "There are 2 matches")
                self.assertNotContains(
                    response,
                    "There are no feature complete toys to display",
                )
                self.assertNotContains(
                    response,
                    "No feature complete toys match your query",
                )

                soup = self.get_soup(response.content)
                label = soup.select_one(f"label#id_{lookup}-label")
                self.assertIsNotNone(label)
                self.assertEqual(label.string.strip(), label_text)
                input = soup.select_one(f"input#id_{lookup}")
                self.assertIsNotNone(input)

    def test_empty_filter_with_results(self):
        for case, (url_namespace, lookup, label_text) in self.cases.items():
            with self.subTest(case=case):
                response = self.get(url_namespace, {lookup: ""})
                self.assertContains(response, "Buzz Lightyear")
                self.assertContains(response, "Forky")
                self.assertNotContains(response, "There are 2 matches")
                self.assertNotContains(
                    response,
                    "No feature complete toys match your query",
                )

                soup = self.get_soup(response.content)
                label = soup.select_one(f"label#id_{lookup}-label")
                self.assertIsNotNone(label)
                self.assertEqual(label.string.strip(), label_text)
                input = soup.select_one(f"input#id_{lookup}")
                self.assertIsNotNone(input)
                self.assertFalse(input.attrs.get("value"))

    def test_filtered_no_results(self):
        lookup_values = {
            "release_date": "1999-09-09",
            "name__icontains": "Woody",
            "release_date__year__lte": "1990",
        }
        for case, (url_namespace, lookup, label_text) in self.cases.items():
            with self.subTest(case=case):
                value = lookup_values[lookup]
                response = self.get(url_namespace, {lookup: value})
                self.assertContains(
                    response,
                    "No feature complete toys match your query",
                )
                self.assertNotContains(response, "Buzz Lightyear")
                self.assertNotContains(response, "Forky")
                self.assertNotContains(response, "There are 2 matches")

                soup = self.get_soup(response.content)
                label = soup.select_one(f"label#id_{lookup}-label")
                self.assertIsNotNone(label)
                self.assertEqual(label.string.strip(), label_text)
                input = soup.select_one(f"input#id_{lookup}")
                self.assertIsNotNone(input)
                self.assertEqual(input.attrs.get("value"), value)

                # Should render the active filters even when there are no results
                active_filters = soup.select_one(".w-active-filters")
                self.assertIsNotNone(active_filters)
                clear = active_filters.select_one(".w-pill__remove")
                self.assertIsNotNone(clear)
                url, params = clear.attrs.get("data-w-swap-src-value").split("?", 1)
                self.assertEqual(url, reverse(f"{url_namespace}:index_results"))
                self.assertNotIn(f"{lookup}={value}", params)

    def test_filtered_with_results(self):
        lookup_values = {
            "release_date": "1995-11-19",
            "name__icontains": "Ightyear",
            "release_date__year__lte": "2017",
        }
        for case, (url_namespace, lookup, label_text) in self.cases.items():
            with self.subTest(case=case):
                value = lookup_values[lookup]
                response = self.get(url_namespace, {lookup: value})
                self.assertContains(response, "Buzz Lightyear")
                self.assertContains(response, "There is 1 match")
                self.assertNotContains(response, "Forky")
                self.assertNotContains(
                    response,
                    "No feature complete toys match your query",
                )

                soup = self.get_soup(response.content)
                label = soup.select_one(f"label#id_{lookup}-label")
                self.assertIsNotNone(label)
                self.assertEqual(label.string.strip(), label_text)
                input = soup.select_one(f"input#id_{lookup}")
                self.assertIsNotNone(input)
                self.assertEqual(input.attrs.get("value"), value)

                # Should render the active filters
                active_filters = soup.select_one(".w-active-filters")
                self.assertIsNotNone(active_filters)
                clear = active_filters.select_one(".w-pill__remove")
                self.assertIsNotNone(clear)
                url, params = clear.attrs.get("data-w-swap-src-value").split("?", 1)
                self.assertEqual(url, reverse(f"{url_namespace}:index_results"))
                self.assertNotIn(f"{lookup}={value}", params)


class TestSearchIndexView(WagtailTestUtils, TestCase):
    url_name = "index"
    cases = {
        # With the default search backend
        "default": ("feature_complete_toy", "release_date"),
        # With Django ORM
        None: ("fctoy-alt2", "release_date__year__lte"),
    }

    def setUp(self):
        self.user = self.login()

    @classmethod
    def setUpTestData(cls):
        FeatureCompleteToy.objects.create(
            name="Buzz Lightyear",
            release_date=datetime.date(1995, 11, 19),
        )
        FeatureCompleteToy.objects.create(
            name="Forky",
            release_date=datetime.date(2019, 6, 11),
        )

    def assertInputRendered(self, response, search_q):
        soup = self.get_soup(response.content)
        input = soup.select_one("input#id_q")
        self.assertIsNotNone(input)
        self.assertEqual(input.attrs.get("value"), search_q)

    def get(self, url_namespace, params=None):
        return self.client.get(reverse(f"{url_namespace}:{self.url_name}"), params)

    def test_search_disabled(self):
        response = self.get("fctoy_alt1", {"q": "ork"})
        self.assertFalse(response.context.get("search_form"))
        self.assertContains(response, "Forky")
        self.assertContains(response, "Buzz Lightyear")
        self.assertNotContains(response, "There are 2 matches")
        soup = self.get_soup(response.content)
        input = soup.select_one("input#id_q")
        self.assertIsNone(input)

    def test_search_no_results(self):
        for backend, (url_namespace, _) in self.cases.items():
            with self.subTest(backend=backend):
                response = self.get(url_namespace, {"q": "Woody"})
                self.assertContains(
                    response,
                    "No feature complete toys match your query",
                )
                self.assertNotContains(response, "Buzz Lightyear")
                self.assertNotContains(response, "Forky")
                self.assertInputRendered(response, "Woody")

    def test_search_with_results(self):
        for backend, (url_namespace, _) in self.cases.items():
            with self.subTest(backend=backend):
                response = self.get(url_namespace, {"q": "ork"})
                self.assertContains(response, "Forky")
                self.assertNotContains(response, "Buzz Lightyear")
                self.assertContains(response, "There is 1 match")
                self.assertInputRendered(response, "ork")

    def test_filtered_searched_no_results(self):
        lookup_values = {
            "release_date": "2019-06-11",
            "release_date__year__lte": "2023",
        }
        for backend, (url_namespace, lookup) in self.cases.items():
            with self.subTest(backend=backend):
                value = lookup_values[lookup]
                response = self.get(url_namespace, {"q": "Woody", lookup: value})
                self.assertContains(
                    response,
                    "No feature complete toys match your query",
                )
                self.assertNotContains(response, "Buzz Lightyear")
                self.assertNotContains(response, "Forky")
                self.assertInputRendered(response, "Woody")

    def test_filtered_searched_with_results(self):
        lookup_values = {
            "release_date": "2019-06-11",
            "release_date__year__lte": "2023",
        }
        for backend, (url_namespace, lookup) in self.cases.items():
            with self.subTest(backend=backend):
                value = lookup_values[lookup]
                response = self.get(url_namespace, {"q": "ork", lookup: value})
                self.assertContains(response, "Forky")
                self.assertNotContains(response, "Buzz Lightyear")
                self.assertContains(response, "There is 1 match")
                self.assertInputRendered(response, "ork")


class TestSearchIndexResultsView(TestSearchIndexView):
    url_name = "index_results"

    def assertInputRendered(self, response, search_q):
        # index_results view doesn't render the search input
        pass


class TestSearchFields(WagtailTestUtils, TestCase):
    @classmethod
    def setUpTestData(cls):
        objects = [
            SearchTestModel(title="Hello World", body="This one is classic"),
            SearchTestModel(title="Hello Anime", body="We love anime (opinions vary)"),
            SearchTestModel(title="Food", body="I like food, do you?"),
        ]
        SearchTestModel.objects.bulk_create(objects)

    def setUp(self):
        self.login()

    def get(self, q):
        return self.client.get(reverse("searchtest:index"), {"q": q})

    def test_single_result_with_body(self):
        response = self.get("IkE")
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Hello World")
        self.assertNotContains(response, "Hello Anime")
        self.assertContains(response, "Food")

    def test_multiple_results_with_title(self):
        response = self.get("ELlo")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Hello World")
        self.assertContains(response, "Hello Anime")
        self.assertNotContains(response, "Food")

    def test_no_results(self):
        response = self.get("Abra Kadabra")
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Hello World")
        self.assertNotContains(response, "Hello Anime")
        self.assertNotContains(response, "Food")


class TestListExport(WagtailTestUtils, TestCase):
    def setUp(self):
        self.user = self.login()

    @classmethod
    def setUpTestData(cls):
        FeatureCompleteToy.objects.create(
            name="Racecar",
            release_date=datetime.date(1995, 11, 19),
        )
        FeatureCompleteToy.objects.create(
            name="LEVEL",
            release_date=datetime.date(2010, 6, 18),
        )
        FeatureCompleteToy.objects.create(
            name="Catso",
            release_date=datetime.date(2010, 6, 18),
        )

    def test_export_disabled(self):
        index_url = reverse("fctoy_alt1:index")
        response = self.client.get(index_url)
        soup = self.get_soup(response.content)

        csv_link = soup.select_one(f"a[href='{index_url}?export=csv']")
        self.assertIsNone(csv_link)
        xlsx_link = soup.select_one(f"a[href='{index_url}?export=xlsx']")
        self.assertIsNone(xlsx_link)

    def test_get_not_export_shows_export_buttons(self):
        index_url = reverse("feature_complete_toy:index")
        response = self.client.get(index_url)
        soup = self.get_soup(response.content)

        csv_link = soup.select_one(f"a[href='{index_url}?export=csv']")
        self.assertIsNotNone(csv_link)
        self.assertEqual(csv_link.text.strip(), "Download CSV")
        xlsx_link = soup.select_one(f"a[href='{index_url}?export=xlsx']")
        self.assertIsNotNone(xlsx_link)
        self.assertEqual(xlsx_link.text.strip(), "Download XLSX")

    def test_get_filtered_shows_export_buttons_with_filters(self):
        index_url = reverse("feature_complete_toy:index")
        response = self.client.get(index_url, {"release_date": "2010-06-18"})
        soup = self.get_soup(response.content)

        csv_link = soup.select_one(
            f"a[href='{index_url}?release_date=2010-06-18&export=csv']"
        )
        self.assertIsNotNone(csv_link)
        self.assertEqual(csv_link.text.strip(), "Download CSV")
        xlsx_link = soup.select_one(
            f"a[href='{index_url}?release_date=2010-06-18&export=xlsx']"
        )
        self.assertIsNotNone(xlsx_link)
        self.assertEqual(xlsx_link.text.strip(), "Download XLSX")

    def test_csv_export(self):
        index_url = reverse("feature_complete_toy:index")
        response = self.client.get(index_url, {"export": "csv"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.get("Content-Disposition"),
            'attachment; filename="feature-complete-toys.csv"',
        )

        data_lines = response.getvalue().decode().strip().split("\r\n")
        self.assertEqual(data_lines[0], "Name,Launch date,Is cool")
        self.assertEqual(data_lines[1], "Catso,2010-06-18,False")
        self.assertEqual(data_lines[2], "LEVEL,2010-06-18,True")
        self.assertEqual(data_lines[3], "Racecar,1995-11-19,")
        self.assertEqual(len(data_lines), 4)

    def test_csv_export_filtered(self):
        index_url = reverse("feature_complete_toy:index")
        response = self.client.get(
            index_url,
            {"release_date": "2010-06-18", "export": "csv"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.get("Content-Disposition"),
            'attachment; filename="feature-complete-toys.csv"',
        )

        data_lines = response.getvalue().decode().strip().split("\r\n")
        self.assertEqual(data_lines[0], "Name,Launch date,Is cool")
        self.assertEqual(data_lines[1], "Catso,2010-06-18,False")
        self.assertEqual(data_lines[2], "LEVEL,2010-06-18,True")
        self.assertEqual(len(data_lines), 3)

    def test_xlsx_export(self):
        index_url = reverse("feature_complete_toy:index")
        response = self.client.get(index_url, {"export": "xlsx"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.get("Content-Disposition"),
            'attachment; filename="feature-complete-toys.xlsx"',
        )

        workbook_data = response.getvalue()
        worksheet = load_workbook(filename=BytesIO(workbook_data)).active
        cell_array = [[cell.value for cell in row] for row in worksheet.rows]
        self.assertEqual(cell_array[0], ["Name", "Launch date", "Is cool"])
        self.assertEqual(cell_array[1], ["Catso", datetime.date(2010, 6, 18), False])
        self.assertEqual(cell_array[2], ["LEVEL", datetime.date(2010, 6, 18), True])
        self.assertEqual(cell_array[3], ["Racecar", datetime.date(1995, 11, 19), None])
        self.assertEqual(len(cell_array), 4)

    def test_xlsx_export_filtered(self):
        index_url = reverse("feature_complete_toy:index")
        response = self.client.get(
            index_url,
            {"release_date": "2010-06-18", "export": "xlsx"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.get("Content-Disposition"),
            'attachment; filename="feature-complete-toys.xlsx"',
        )

        workbook_data = response.getvalue()
        worksheet = load_workbook(filename=BytesIO(workbook_data)).active
        cell_array = [[cell.value for cell in row] for row in worksheet.rows]
        self.assertEqual(cell_array[0], ["Name", "Launch date", "Is cool"])
        self.assertEqual(cell_array[1], ["Catso", datetime.date(2010, 6, 18), False])
        self.assertEqual(cell_array[2], ["LEVEL", datetime.date(2010, 6, 18), True])
        self.assertEqual(len(cell_array), 3)


class TestPagination(WagtailTestUtils, TestCase):
    def setUp(self):
        self.user = self.login()

    @classmethod
    def setUpTestData(cls):
        objects = [FeatureCompleteToy(name=f"Frisbee {i}") for i in range(32)]
        FeatureCompleteToy.objects.bulk_create(objects)

    def test_default_list_pagination(self):
        list_url = reverse("fctoy_alt1:index")
        response = self.client.get(list_url)

        # Default is 20 per page
        self.assertEqual(FeatureCompleteToy.objects.all().count(), 32)
        self.assertContains(response, "Page 1 of 2")
        self.assertContains(response, "Next")
        self.assertContains(response, list_url + "?p=2")

    def test_custom_list_pagination(self):
        list_url = reverse("feature_complete_toy:index")
        response = self.client.get(list_url)

        # Custom is set to display 5 per page
        self.assertEqual(FeatureCompleteToy.objects.all().count(), 32)
        self.assertContains(response, "Page 1 of 7")
        self.assertContains(response, "Next")
        self.assertContains(response, list_url + "?p=2")


class TestOrdering(WagtailTestUtils, TestCase):
    def setUp(self):
        self.user = self.login()

    @classmethod
    def setUpTestData(cls):
        objects = [
            FeatureCompleteToy(name="CCCCCCCCCC", strid="1"),
            FeatureCompleteToy(name="AAAAAAAAAA", strid="2"),
            FeatureCompleteToy(name="DDDDDDDDDD", strid="3"),
            FeatureCompleteToy(name="BBBBBBBBBB", strid="4"),
        ]
        FeatureCompleteToy.objects.bulk_create(objects)

    def test_default_order(self):
        response = self.client.get(reverse("fctoy_alt1:index"))
        # Without ordering on the model, should be ordered descending by pk
        self.assertFalse(FeatureCompleteToy._meta.ordering)
        self.assertEqual(
            [obj.name for obj in response.context["object_list"]],
            [
                "BBBBBBBBBB",
                "DDDDDDDDDD",
                "AAAAAAAAAA",
                "CCCCCCCCCC",
            ],
        )

    def test_custom_order_from_query_args(self):
        response = self.client.get(reverse("fctoy-alt3:index") + "?ordering=-name")
        self.assertFalse(FeatureCompleteToy._meta.ordering)
        self.assertEqual(
            [obj.name for obj in response.context["object_list"]],
            [
                "DDDDDDDDDD",
                "CCCCCCCCCC",
                "BBBBBBBBBB",
                "AAAAAAAAAA",
            ],
        )

    def test_custom_order_from_view(self):
        response = self.client.get(reverse("feature_complete_toy:index"))
        # Should respect the view's ordering
        self.assertFalse(FeatureCompleteToy._meta.ordering)
        self.assertEqual(
            [obj.name for obj in response.context["object_list"]],
            [
                "AAAAAAAAAA",
                "BBBBBBBBBB",
                "CCCCCCCCCC",
                "DDDDDDDDDD",
            ],
        )

    def test_custom_order_from_from_viewset(self):
        response = self.client.get(reverse("fctoy-alt3:index"))
        # The view has an ordering but it is overwritten by the viewset
        self.assertFalse(FeatureCompleteToy._meta.ordering)
        self.assertEqual(
            [obj.name for obj in response.context["object_list"]],
            [
                "CCCCCCCCCC",
                "AAAAAAAAAA",
                "DDDDDDDDDD",
                "BBBBBBBBBB",
            ],
        )


class TestBreadcrumbs(AdminTemplateTestUtils, WagtailTestUtils, TestCase):
    def setUp(self):
        self.user = self.login()

    @classmethod
    def setUpTestData(cls):
        cls.object = FeatureCompleteToy.objects.create(name="Test Toy")

    def test_index_view(self):
        response = self.client.get(reverse("feature_complete_toy:index"))
        items = [
            {
                "url": "",
                "label": "Feature complete toys",
            }
        ]
        self.assertBreadcrumbsItemsRendered(items, response.content)

    def test_add_view(self):
        response = self.client.get(reverse("feature_complete_toy:add"))
        items = [
            {
                "url": reverse("feature_complete_toy:index"),
                "label": "Feature complete toys",
            },
            {
                "url": "",
                "label": "New: Feature complete toy",
            },
        ]
        self.assertBreadcrumbsItemsRendered(items, response.content)

    def test_edit_view(self):
        edit_url = reverse("feature_complete_toy:edit", args=(quote(self.object.pk),))
        response = self.client.get(edit_url)
        items = [
            {
                "url": reverse("feature_complete_toy:index"),
                "label": "Feature complete toys",
            },
            {
                "url": "",
                "label": str(self.object),
            },
        ]
        self.assertBreadcrumbsItemsRendered(items, response.content)

    def test_delete_view(self):
        delete_url = reverse(
            "feature_complete_toy:delete",
            args=(quote(self.object.pk),),
        )
        response = self.client.get(delete_url)
        self.assertBreadcrumbsNotRendered(response.content)

    def test_history_view(self):
        history_url = reverse(
            "feature_complete_toy:history",
            args=(quote(self.object.pk),),
        )
        response = self.client.get(history_url)
        items = [
            {
                "url": reverse("feature_complete_toy:index"),
                "label": "Feature complete toys",
            },
            {
                "url": reverse(
                    "feature_complete_toy:edit", args=(quote(self.object.pk),)
                ),
                "label": str(self.object),
            },
            {
                "url": "",
                "label": "History",
                "sublabel": str(self.object),
            },
        ]
        self.assertBreadcrumbsItemsRendered(items, response.content)

    def test_history_view_pagination(self):
        for i in range(25):
            log(instance=self.object, action="wagtail.edit", user=self.user)

        history_url = reverse(
            "feature_complete_toy:history",
            args=(quote(self.object.pk),),
        )
        response = self.client.get(history_url)
        self.assertContains(response, "Page 1 of 2")
        self.assertContains(response, f'<a href="{history_url}?p=2">')

    def test_usage_view(self):
        usage_url = reverse(
            "feature_complete_toy:usage",
            args=(quote(self.object.pk),),
        )
        response = self.client.get(usage_url)
        items = [
            {
                "url": reverse("feature_complete_toy:index"),
                "label": "Feature complete toys",
            },
            {
                "url": reverse(
                    "feature_complete_toy:edit", args=(quote(self.object.pk),)
                ),
                "label": str(self.object),
            },
            {
                "url": "",
                "label": "Usage",
                "sublabel": str(self.object),
            },
        ]
        self.assertBreadcrumbsItemsRendered(items, response.content)

    def test_usage_view_pagination(self):
        with self.captureOnCommitCallbacks(execute=True):
            for i in range(25):
                VariousOnDeleteModel.objects.create(
                    text=f"Toybox {i}", cascading_toy=self.object
                )

        usage_url = reverse(
            "feature_complete_toy:usage",
            args=(quote(self.object.pk),),
        )
        response = self.client.get(usage_url)
        self.assertContains(response, "Page 1 of 2")
        self.assertContains(response, f'<a href="{usage_url}?p=2">')

    def test_inspect_view(self):
        inspect_url = reverse(
            "feature_complete_toy:inspect",
            args=(quote(self.object.pk),),
        )
        response = self.client.get(inspect_url)
        items = [
            {
                "url": reverse("feature_complete_toy:index"),
                "label": "Feature complete toys",
            },
            {
                "url": reverse(
                    "feature_complete_toy:edit", args=(quote(self.object.pk),)
                ),
                "label": str(self.object),
            },
            {
                "url": "",
                "label": "Inspect",
                "sublabel": str(self.object),
            },
        ]
        self.assertBreadcrumbsItemsRendered(items, response.content)


class TestLegacyPatterns(WagtailTestUtils, TestCase):
    # RemovedInWagtail70Warning: legacy integer pk-based URLs will be removed

    def setUp(self):
        self.user = self.login()

    @classmethod
    def setUpTestData(cls):
        cls.object = JSONStreamModel.objects.create(
            body='[{"type": "text", "value": "foo"}]',
        )

    def test_legacy_edit(self):
        edit_url = reverse("streammodel:edit", args=(quote(self.object.pk),))
        legacy_edit_url = "/admin/streammodel/1/"
        with self.assertWarnsRegex(
            RemovedInWagtail70Warning,
            "`/<pk>/` edit view URL pattern has been deprecated in favour of /edit/<pk>/.",
        ):
            response = self.client.get(legacy_edit_url)
        self.assertEqual(edit_url, "/admin/streammodel/edit/1/")
        self.assertRedirects(response, edit_url, 301)

    def test_legacy_delete(self):
        delete_url = reverse("streammodel:delete", args=(quote(self.object.pk),))
        legacy_delete_url = "/admin/streammodel/1/delete/"
        with self.assertWarnsRegex(
            RemovedInWagtail70Warning,
            "`/<pk>/delete/` delete view URL pattern has been deprecated in favour of /delete/<pk>/.",
        ):
            response = self.client.get(legacy_delete_url)
        self.assertEqual(delete_url, "/admin/streammodel/delete/1/")
        self.assertRedirects(response, delete_url, 301)


class TestHistoryView(WagtailTestUtils, TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = cls.create_test_user()
        cls.object = FeatureCompleteToy.objects.create(name="Buzz")
        cls.url = reverse(
            "feature_complete_toy:history",
            args=(quote(cls.object.pk),),
        )

        content_type = ContentType.objects.get_for_model(FeatureCompleteToy)
        cls.timestamp_1 = datetime.datetime(2021, 9, 30, 10, 1, 0)
        cls.timestamp_2 = datetime.datetime(2022, 5, 10, 12, 34, 0)
        if settings.USE_TZ:
            cls.timestamp_1 = make_aware(cls.timestamp_1)
            cls.timestamp_2 = make_aware(cls.timestamp_2)
        ModelLogEntry.objects.create(
            content_type=content_type,
            label="Test Buzz",
            action="wagtail.create",
            user=cls.user,
            timestamp=cls.timestamp_1,
            object_id=cls.object.pk,
        )
        ModelLogEntry.objects.create(
            content_type=content_type,
            label="Test Buzz Updated",
            action="wagtail.edit",
            user=cls.user,
            timestamp=cls.timestamp_2,
            object_id=cls.object.pk,
        )

    def setUp(self):
        self.login(self.user)

    def test_simple(self):
        expected = (
            ("Edited", str(self.user), date_format(self.timestamp_2, "c")),
            ("Created", str(self.user), date_format(self.timestamp_1, "c")),
        )
        response = self.client.get(self.url)
        soup = self.get_soup(response.content)
        rows = soup.select("#listing-results tbody tr")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(rows), 2)

        rendered_rows = []
        for row in rows:
            cells = []
            tds = row.select("td")
            self.assertEqual(len(tds), 3)
            cells.append(tds[0].text.strip())
            cells.append(tds[1].text.strip())
            cells.append(tds[2].select_one("time").attrs.get("datetime"))
            rendered_rows.append(cells)

        for rendered_row, expected_row in zip(rendered_rows, expected):
            self.assertSequenceEqual(rendered_row, expected_row)

        # History view is not searchable
        input = soup.select_one("input#id_q")
        self.assertIsNone(input)
        self.assertFalse(response.context.get("search_form"))

    def test_action_filter(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)

        # Should only show the created and edited options for the filter
        soup = self.get_soup(response.content)
        options = soup.select('input[name="action"][type="checkbox"]')
        self.assertEqual(len(options), 2)
        self.assertEqual(
            {option.attrs.get("value") for option in options},
            {"wagtail.create", "wagtail.edit"},
        )
        # Should not show the heading when not searching
        heading = soup.select_one('h2[role="alert"]')
        self.assertIsNone(heading)

        # Should only show the edited log
        response = self.client.get(self.url, {"action": "wagtail.edit"})
        self.assertEqual(response.status_code, 200)
        soup = self.get_soup(response.content)
        rows = soup.select("#listing-results tbody tr")
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].select_one("td").text.strip(), "Edited")

        # Should only show the created log
        response = self.client.get(self.url, {"action": "wagtail.create"})
        self.assertEqual(response.status_code, 200)
        soup = self.get_soup(response.content)
        rows = soup.select("#listing-results tbody tr")
        heading = soup.select_one('h2[role="alert"]')
        self.assertEqual(heading.string.strip(), "There is 1 match")
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].select_one("td").text.strip(), "Created")

        # Should display the heading when there are results
        heading = soup.select_one('h2[role="alert"]')
        self.assertEqual(heading.string.strip(), "There is 1 match")

        response = self.client.get(
            self.url,
            {"action": ["wagtail.create", "wagtail.edit"]},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context["object_list"]), 2)

    def test_user_filter(self):
        # A user who absolutely has no permissions to the model
        self.create_user("no_power")

        # A user who has permissions to the model,
        # but only has logs for a different object
        has_power = self.create_superuser("has_power")
        other_obj = FeatureCompleteToy.objects.create(name="Woody")
        log(
            instance=other_obj,
            action="wagtail.create",
            content_changed=True,
            user=has_power,
        )

        # A user who does not have permissions to the model,
        # but has logs for the object (e.g. maybe they used to have permissions)
        previously_has_power = self.create_user("previously_has_power")
        log(
            instance=self.object,
            action="wagtail.edit",
            content_changed=True,
            user=previously_has_power,
        )

        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        soup = self.get_soup(response.content)

        # Should only show the current user and the previously_has_power user
        options = soup.select('input[name="user"][type="checkbox"]')
        self.assertEqual(len(options), 2)
        self.assertEqual(
            {option.attrs.get("value") for option in options},
            {str(self.user.pk), str(previously_has_power.pk)},
        )

        response = self.client.get(self.url, {"user": str(self.user.pk)})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context["object_list"]), 2)

        response = self.client.get(self.url, {"user": str(previously_has_power.pk)})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context["object_list"]), 1)

        # Should allow filtering by multiple users
        response = self.client.get(
            self.url,
            {"user": [str(self.user.pk), str(previously_has_power.pk)]},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context["object_list"]), 3)

    def test_filtered_no_results(self):
        response = self.client.get(self.url, {"timestamp_to": "2020-01-01"})
        soup = self.get_soup(response.content)
        results = soup.select_one("#listing-results")
        table = soup.select_one("#listing-results table")
        p = results.select_one("p")
        self.assertEqual(response.status_code, 200)
        self.assertIsNotNone(results)
        self.assertIsNone(table)
        self.assertIsNotNone(p)
        self.assertEqual(p.text.strip(), "No log entries match your query.")

    def test_empty(self):
        ModelLogEntry.objects.all().delete()
        response = self.client.get(self.url)
        soup = self.get_soup(response.content)
        results = soup.select_one("#listing-results")
        table = soup.select_one("#listing-results table")
        self.assertEqual(response.status_code, 200)
        self.assertIsNotNone(results)
        self.assertEqual(results.text.strip(), "There are no log entries to display.")
        self.assertIsNone(table)

        # Should hide the action and user filters as there are no choices for
        # these filters (since the choices are based on the current queryset)
        action_inputs = soup.select('input[name="action"]')
        self.assertEqual(len(action_inputs), 0)
        user_inputs = soup.select('input[name="user"]')
        self.assertEqual(len(user_inputs), 0)

        # Should still render the timestamp filter as it is always available
        timestamp_before_input = soup.select_one('input[name="timestamp_to"]')
        self.assertIsNotNone(timestamp_before_input)
        timestamp_after_input = soup.select_one('input[name="timestamp_from"]')
        self.assertIsNotNone(timestamp_after_input)

    def test_edit_view_links_to_history_view(self):
        edit_url = reverse("feature_complete_toy:edit", args=(quote(self.object.pk),))
        response = self.client.get(edit_url)
        soup = self.get_soup(response.content)
        header = soup.select_one(".w-slim-header")
        history_link = header.find("a", attrs={"href": self.url})
        self.assertIsNotNone(history_link)

    def test_deleted_user(self):
        to_be_deleted = self.create_user("to_be_deleted")
        user_id = to_be_deleted.pk
        log(
            instance=self.object,
            action="wagtail.edit",
            content_changed=True,
            user=to_be_deleted,
        )
        to_be_deleted.delete()
        response = self.client.get(self.url)
        self.assertContains(response, f"user {user_id} (deleted)")


class TestUsageView(WagtailTestUtils, TestCase):
    @classmethod
    def setUpTestData(cls):
        with cls.captureOnCommitCallbacks(execute=True):
            cls.user = cls.create_test_user()
            cls.object = FeatureCompleteToy.objects.create(name="Buzz")
            cls.url = reverse(
                "feature_complete_toy:usage",
                args=(quote(cls.object.pk),),
            )
            cls.tbx = VariousOnDeleteModel.objects.create(
                text="Toybox", cascading_toy=cls.object
            )

    def setUp(self):
        self.user = self.login(self.user)

    def test_simple(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)

        soup = self.get_soup(response.content)
        h1 = soup.select_one("h1")
        self.assertEqual(h1.text.strip(), f"Usage: {self.object}")

        tds = soup.select("#listing-results tbody tr td")
        self.assertEqual(len(tds), 3)
        self.assertEqual(tds[0].text.strip(), str(self.tbx))
        self.assertEqual(tds[1].text.strip(), "Various on delete model")
        self.assertEqual(tds[2].text.strip(), "Cascading toy")

        tbx_edit_url = AdminURLFinder(self.user).get_edit_url(self.tbx)

        # Link to referrer's edit view
        link = tds[0].select_one("a")
        self.assertIsNotNone(link)
        self.assertEqual(link.attrs.get("href"), tbx_edit_url)
        content_path_link = tds[-1].select_one("a")
        self.assertEqual(
            content_path_link.attrs.get("href"),
            tbx_edit_url + "#:w:contentpath=cascading_toy",
        )

        # Link to referrer's edit view with parameters for the specific field
        link = tds[2].select_one("a")
        self.assertIsNotNone(link)
        self.assertIn(tbx_edit_url, link.attrs.get("href"))

        # Usage view is not searchable
        input = soup.select_one("input#id_q")
        self.assertIsNone(input)
        self.assertFalse(response.context.get("search_form"))

    def test_usage_without_permission(self):
        self.user.is_superuser = False
        self.user.save()
        admin_permission = Permission.objects.get(
            content_type__app_label="wagtailadmin", codename="access_admin"
        )
        self.user.user_permissions.add(admin_permission)

        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("wagtailadmin_home"))

    def test_usage_without_permission_on_referrer(self):
        self.user.is_superuser = False
        self.user.save()
        admin_permission = Permission.objects.get(
            content_type__app_label="wagtailadmin", codename="access_admin"
        )
        toy_edit_permission = Permission.objects.get(
            content_type__app_label="tests", codename="change_featurecompletetoy"
        )
        self.user.user_permissions.add(admin_permission, toy_edit_permission)

        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)

        soup = self.get_soup(response.content)
        h1 = soup.select_one("h1")
        self.assertEqual(h1.text.strip(), f"Usage: {self.object}")

        tds = soup.select("#listing-results tbody tr td")
        self.assertEqual(len(tds), 3)
        self.assertEqual(tds[0].text.strip(), "(Private various on delete model)")
        self.assertEqual(tds[1].text.strip(), "Various on delete model")
        self.assertEqual(tds[2].text.strip(), "Cascading toy")

        # Not link to referrer's edit view
        link = tds[0].select_one("a")
        self.assertIsNone(link)

        # Not link to referrer's edit view
        link = tds[2].select_one("a")
        self.assertIsNone(link)

    def test_usage_with_describe_on_delete(self):
        response = self.client.get(self.url + "?describe_on_delete=1")
        self.assertEqual(response.status_code, 200)

        soup = self.get_soup(response.content)
        h1 = soup.select_one("h1")
        self.assertEqual(h1.text.strip(), f"Usage: {self.object}")

        tds = soup.select("#listing-results tbody tr td")
        self.assertEqual(len(tds), 3)
        self.assertEqual(tds[0].text.strip(), str(self.tbx))
        self.assertEqual(tds[1].text.strip(), "Various on delete model")
        self.assertEqual(
            tds[2].text.strip(),
            "Cascading toy: the various on delete model will also be deleted",
        )

        tbx_edit_url = AdminURLFinder(self.user).get_edit_url(self.tbx)

        # Link to referrer's edit view
        link = tds[0].select_one("a")
        self.assertIsNotNone(link)
        self.assertEqual(link.attrs.get("href"), tbx_edit_url)

        # Link to referrer's edit view with parameters for the specific field
        link = tds[2].select_one("a")
        self.assertIsNotNone(link)
        self.assertIn(tbx_edit_url, link.attrs.get("href"))

    def test_empty(self):
        self.tbx.delete()
        response = self.client.get(self.url)
        soup = self.get_soup(response.content)
        results = soup.select_one("#listing-results")
        table = soup.select_one("#listing-results table")
        self.assertEqual(response.status_code, 200)
        self.assertIsNotNone(results)
        self.assertEqual(results.text.strip(), "There are no results.")
        self.assertIsNone(table)

    def test_edit_view_links_to_usage_view(self):
        edit_url = reverse("feature_complete_toy:edit", args=(quote(self.object.pk),))
        response = self.client.get(edit_url)
        soup = self.get_soup(response.content)
        side_panel = soup.select_one("[data-side-panel='status']")
        usage_link = side_panel.find("a", attrs={"href": self.url})
        self.assertIsNotNone(usage_link)

    def test_delete_view_links_to_usage_view(self):
        edit_url = reverse("feature_complete_toy:delete", args=(quote(self.object.pk),))
        response = self.client.get(edit_url)
        soup = self.get_soup(response.content)
        usage_link = soup.find("a", attrs={"href": self.url + "?describe_on_delete=1"})
        self.assertIsNotNone(usage_link)


class TestInspectView(WagtailTestUtils, TestCase):
    def setUp(self):
        self.user = self.login()

    @classmethod
    def setUpTestData(cls):
        cls.object = FeatureCompleteToy.objects.create(name="Test Toy")
        cls.url = reverse("feature_complete_toy:inspect", args=(quote(cls.object.pk),))
        cls.edit_url = reverse(
            "feature_complete_toy:edit", args=(quote(cls.object.pk),)
        )
        cls.delete_url = reverse(
            "feature_complete_toy:delete", args=(quote(cls.object.pk),)
        )

    def test_simple(self):
        response = self.client.get(self.url)
        expected_fields = ["Strid", "Release date"]
        expected_values = [
            # The pk may contain whitespace at the start/end, it's hard to
            # distinguish from the whitespace in the HTML so just strip it
            self.object.pk.strip(),
            localize(self.object.release_date),
        ]
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "wagtailadmin/generic/inspect.html")
        soup = self.get_soup(response.content)
        fields = [dt.text.strip() for dt in soup.select("dt")]
        values = [dd.text.strip() for dd in soup.select("dd")]
        self.assertEqual(fields, expected_fields)
        self.assertEqual(values, expected_values)
        # One in the breadcrumb, one at the bottom
        self.assertEqual(len(soup.find_all("a", attrs={"href": self.edit_url})), 2)
        self.assertEqual(len(soup.find_all("a", attrs={"href": self.delete_url})), 1)

    def test_inspect_view_fields(self):
        # The alt1 viewset has a custom inspect_view_fields and inspect_view_fields_exclude
        response = self.client.get(
            reverse("fctoy_alt1:inspect", args=(quote(self.object.pk),))
        )
        expected_fields = ["Name"]
        expected_values = ["Test Toy"]
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "wagtailadmin/generic/inspect.html")
        soup = self.get_soup(response.content)
        fields = [dt.text.strip() for dt in soup.select("dt")]
        values = [dd.text.strip() for dd in soup.select("dd")]
        self.assertEqual(fields, expected_fields)
        self.assertEqual(values, expected_values)

    def test_view_permission_registered(self):
        content_type = ContentType.objects.get_for_model(FeatureCompleteToy)
        qs = Permission.objects.none()
        for fn in hooks.get_hooks("register_permissions"):
            qs |= fn()
        registered_user_permissions = qs.filter(content_type=content_type)
        self.assertEqual(
            set(registered_user_permissions.values_list("codename", flat=True)),
            {
                "add_featurecompletetoy",
                "change_featurecompletetoy",
                "delete_featurecompletetoy",
                # The "view" permission should be registered if inspect view is enabled
                "view_featurecompletetoy",
                # Any custom permissions should be registered too
                "can_set_release_date",
            },
        )

    def test_disabled(self):
        # An alternate viewset for the same model without inspect_view_enabled = True
        with self.assertRaises(NoReverseMatch):
            reverse("fctoy-alt2:inspect", args=(quote(self.object.pk),))

    def test_without_permission(self):
        self.user.is_superuser = False
        self.user.save()
        admin_permission = Permission.objects.get(
            content_type__app_label="wagtailadmin", codename="access_admin"
        )
        self.user.user_permissions.add(admin_permission)

        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("wagtailadmin_home"))

    def assert_minimal_permission(self, permission):
        self.user.is_superuser = False
        self.user.user_permissions.add(
            Permission.objects.get(
                content_type__app_label="wagtailadmin", codename="access_admin"
            ),
            Permission.objects.get(
                content_type__app_label=self.object._meta.app_label,
                codename=get_permission_codename(permission, self.object._meta),
            ),
        )
        self.user.save()

        response = self.client.get(self.url)
        expected_fields = ["Strid", "Release date"]
        expected_values = [
            # The pk may contain whitespace at the start/end, it's hard to
            # distinguish from the whitespace in the HTML so just strip it
            self.object.pk.strip(),
            localize(self.object.release_date),
        ]
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "wagtailadmin/generic/inspect.html")
        soup = self.get_soup(response.content)
        fields = [dt.text.strip() for dt in soup.select("dt")]
        values = [dd.text.strip() for dd in soup.select("dd")]
        self.assertEqual(fields, expected_fields)
        self.assertEqual(values, expected_values)
        self.assertEqual(len(soup.find_all("a", attrs={"href": self.edit_url})), 0)
        self.assertEqual(len(soup.find_all("a", attrs={"href": self.delete_url})), 0)

    def test_only_add_permission(self):
        self.assert_minimal_permission("add")

    def test_only_view_permission(self):
        self.assert_minimal_permission("view")


class TestListingButtons(WagtailTestUtils, TestCase):
    def setUp(self):
        self.user = self.login()

    @classmethod
    def setUpTestData(cls):
        cls.object = FeatureCompleteToy.objects.create(name="Test Toy")

    def test_simple(self):
        response = self.client.get(reverse("feature_complete_toy:index"))

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "wagtailadmin/shared/buttons.html")

        soup = self.get_soup(response.content)
        actions = soup.select_one("tbody tr td ul.actions")
        more_dropdown = actions.select_one("li [data-controller='w-dropdown']")
        self.assertIsNotNone(more_dropdown)
        more_button = more_dropdown.select_one("button")
        self.assertEqual(
            more_button.attrs.get("aria-label").strip(),
            f"More options for '{self.object}'",
        )

        expected_buttons = [
            (
                "Edit",
                f"Edit '{self.object}'",
                reverse("feature_complete_toy:edit", args=[quote(self.object.pk)]),
            ),
            (
                "Copy",
                f"Copy '{self.object}'",
                reverse("feature_complete_toy:copy", args=[quote(self.object.pk)]),
            ),
            (
                "Inspect",
                f"Inspect '{self.object}'",
                reverse("feature_complete_toy:inspect", args=[quote(self.object.pk)]),
            ),
            (
                "Delete",
                f"Delete '{self.object}'",
                reverse("feature_complete_toy:delete", args=[quote(self.object.pk)]),
            ),
        ]

        rendered_buttons = more_dropdown.select("a")
        self.assertEqual(len(rendered_buttons), len(expected_buttons))

        for rendered_button, (label, aria_label, url) in zip(
            rendered_buttons, expected_buttons
        ):
            self.assertEqual(rendered_button.text.strip(), label)
            self.assertEqual(rendered_button.attrs.get("aria-label"), aria_label)
            self.assertEqual(rendered_button.attrs.get("href"), url)

    def test_title_cell_not_link_to_edit_view_when_no_edit_permission(self):
        self.user.is_superuser = False
        self.user.save()
        admin_permission = Permission.objects.get(
            content_type__app_label="wagtailadmin",
            codename="access_admin",
        )
        add_permission = Permission.objects.get(
            content_type__app_label=self.object._meta.app_label,
            codename=get_permission_codename("add", self.object._meta),
        )
        self.user.user_permissions.add(admin_permission, add_permission)

        response = self.client.get(reverse("fctoy-alt2:index"))
        self.assertEqual(response.status_code, 200)
        soup = self.get_soup(response.content)
        title_wrapper = soup.select_one("#listing-results td.title .title-wrapper")
        self.assertIsNotNone(title_wrapper)

        # fctoy-alt2 doesn't have inspect view enabled, so the title cell should
        # not link anywhere
        self.assertIsNone(title_wrapper.select_one("a"))
        self.assertEqual(title_wrapper.text.strip(), str(self.object))

        # There should be no edit link at all on the page
        self.assertNotContains(
            response,
            reverse("fctoy-alt2:edit", args=[quote(self.object.pk)]),
        )

    def test_title_cell_links_to_inspect_view_when_no_edit_permission(self):
        self.user.is_superuser = False
        self.user.save()
        admin_permission = Permission.objects.get(
            content_type__app_label="wagtailadmin",
            codename="access_admin",
        )
        view_permission = Permission.objects.get(
            content_type__app_label=self.object._meta.app_label,
            codename=get_permission_codename("view", self.object._meta),
        )
        self.user.user_permissions.add(admin_permission, view_permission)

        response = self.client.get(reverse("feature_complete_toy:index"))
        self.assertEqual(response.status_code, 200)
        soup = self.get_soup(response.content)
        title_wrapper = soup.select_one("#listing-results td.title .title-wrapper")
        self.assertIsNotNone(title_wrapper)
        link = title_wrapper.select_one("a")
        self.assertIsNotNone(link)
        self.assertEqual(link.text.strip(), self.object.name)
        self.assertEqual(
            link.get("href"),
            reverse("feature_complete_toy:inspect", args=[quote(self.object.pk)]),
        )

        # Should contain the inspect link twice:
        # once in the title cell and once in the dropdown
        self.assertContains(
            response,
            reverse("feature_complete_toy:inspect", args=[quote(self.object.pk)]),
            count=2,
        )

        # There should be no edit link at all on the page
        self.assertNotContains(
            response,
            reverse("feature_complete_toy:edit", args=[quote(self.object.pk)]),
        )

    def test_copy_disabled(self):
        response = self.client.get(reverse("fctoy_alt1:index"))

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "wagtailadmin/shared/buttons.html")

        soup = self.get_soup(response.content)
        actions = soup.select_one("tbody tr td ul.actions")
        more_dropdown = actions.select_one("li [data-controller='w-dropdown']")
        self.assertIsNotNone(more_dropdown)
        more_button = more_dropdown.select_one("button")
        self.assertEqual(
            more_button.attrs.get("aria-label").strip(),
            f"More options for '{self.object}'",
        )

        expected_buttons = [
            (
                "Edit",
                f"Edit '{self.object}'",
                reverse("fctoy_alt1:edit", args=[quote(self.object.pk)]),
            ),
            (
                "Inspect",
                f"Inspect '{self.object}'",
                reverse("fctoy_alt1:inspect", args=[quote(self.object.pk)]),
            ),
            (
                "Delete",
                f"Delete '{self.object}'",
                reverse("fctoy_alt1:delete", args=[quote(self.object.pk)]),
            ),
        ]

        rendered_buttons = more_dropdown.select("a")
        self.assertEqual(len(rendered_buttons), len(expected_buttons))

        for rendered_button, (label, aria_label, url) in zip(
            rendered_buttons, expected_buttons
        ):
            self.assertEqual(rendered_button.text.strip(), label)
            self.assertEqual(rendered_button.attrs.get("aria-label"), aria_label)
            self.assertEqual(rendered_button.attrs.get("href"), url)

    def test_dropdown_not_rendered_when_no_child_buttons_exist(self):
        self.user.is_superuser = False
        self.user.save()
        admin_permission = Permission.objects.get(
            content_type__app_label="wagtailadmin",
            codename="access_admin",
        )
        add_permission = Permission.objects.get(
            content_type__app_label=self.object._meta.app_label,
            codename=get_permission_codename("add", self.object._meta),
        )
        self.user.user_permissions.add(admin_permission, add_permission)

        # The alt3 viewset doesn't have "copy" and "inspect" views enabled,
        # so when only "add" permission is granted, the dropdown should have
        # no items and thus not be rendered at all
        response = self.client.get(reverse("fctoy-alt3:index"))
        soup = self.get_soup(response.content)
        actions = soup.select_one("tbody tr td ul.actions")
        self.assertIsNone(actions)


class TestCopyView(WagtailTestUtils, TestCase):
    def setUp(self):
        self.user = self.login()
        self.url = reverse("feature_complete_toy:copy", args=[quote(self.object.pk)])

    @classmethod
    def setUpTestData(cls):
        cls.object = FeatureCompleteToy.objects.create(name="Test Toy")

    def test_without_permission(self):
        self.user.is_superuser = False
        self.user.save()
        admin_permission = Permission.objects.get(
            content_type__app_label="wagtailadmin", codename="access_admin"
        )
        self.user.user_permissions.add(admin_permission)

        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("wagtailadmin_home"))

    def test_form_is_prefilled(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        soup = self.get_soup(response.content)
        name_input = soup.select_one('input[name="name"]')
        self.assertEqual(name_input.attrs.get("value"), "Test Toy")


class TestEditHandler(WagtailTestUtils, TestCase):
    def setUp(self):
        self.user = self.login()

    @classmethod
    def setUpTestData(cls):
        cls.object = FeatureCompleteToy.objects.create(name="Test Toy")
        cls.url = reverse("feature_complete_toy:edit", args=(quote(cls.object.pk),))

    def test_edit_form_rendered_with_panels(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "wagtailadmin/shared/panel.html")
        # Many features from the Panels API are powered by client-side JS in
        # _editor_js.html. We need to make sure that this template is included
        # in the response for now.
        self.assertTemplateUsed(response, "wagtailadmin/pages/_editor_js.html", count=1)

        soup = self.get_soup(response.content)

        # Minimap should be rendered
        minimap_container = soup.select_one("[data-minimap-container]")
        self.assertIsNotNone(minimap_container)

        # Form should be rendered using panels
        panels = soup.select("[data-panel]")
        self.assertEqual(len(panels), 2)
        headings = ["Name", "Release date"]
        for expected_heading, panel in zip(headings, panels):
            rendered_heading = panel.select_one("[data-panel-heading-text]")
            self.assertIsNotNone(rendered_heading)
            self.assertEqual(rendered_heading.text.strip(), expected_heading)

    def test_field_permissions(self):
        self.user.is_superuser = False
        self.user.save()
        self.user.user_permissions.add(
            Permission.objects.get(
                content_type__app_label="wagtailadmin", codename="access_admin"
            ),
            Permission.objects.get(
                content_type__app_label=self.object._meta.app_label,
                codename=get_permission_codename("change", self.object._meta),
            ),
        )

        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(list(response.context["form"].fields), ["name"])

        self.user.user_permissions.add(
            Permission.objects.get(
                codename="can_set_release_date",
            )
        )

        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            list(response.context["form"].fields), ["name", "release_date"]
        )


class TestDefaultMessages(WagtailTestUtils, TestCase):
    def setUp(self):
        self.user = self.login()

    @classmethod
    def setUpTestData(cls):
        cls.object = FeatureCompleteToy.objects.create(name="Test Toy")
        cls.create_url = reverse("feature_complete_toy:add")
        cls.edit_url = reverse(
            "feature_complete_toy:edit", args=(quote(cls.object.pk),)
        )
        cls.delete_url = reverse(
            "feature_complete_toy:delete", args=(quote(cls.object.pk),)
        )

    def test_create_error(self):
        response = self.client.post(
            self.create_url,
            data={"name": "", "release_date": "2024-01-11"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            escape("The feature complete toy could not be created due to errors."),
        )

    def test_create_success(self):
        response = self.client.post(
            self.create_url,
            data={"name": "Pink Flamingo", "release_date": "2024-01-11"},
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            escape("Feature complete toy 'Pink Flamingo (2024-01-11)' created."),
        )

    def test_edit_error(self):
        response = self.client.post(
            self.edit_url, data={"name": "", "release_date": "2024-01-11"}
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            escape("The feature complete toy could not be saved due to errors."),
        )

    def test_edit_success(self):
        response = self.client.post(
            self.edit_url,
            data={"name": "rubberduck", "release_date": "2024-02-01"},
            follow=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            escape("Feature complete toy 'rubberduck (2024-02-01)' updated."),
        )

    def test_delete_success(self):
        response = self.client.post(self.delete_url, follow=True)
        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            escape(f"Feature complete toy '{self.object}' deleted."),
        )


class TestHeaderButtons(WagtailTestUtils, TestCase):
    def setUp(self):
        self.user = self.login()

    @classmethod
    def setUpTestData(cls):
        cls.object = FeatureCompleteToy.objects.create(name="Test Toy")
        cls.edit_url = reverse(
            "feature_complete_toy:edit", args=(quote(cls.object.pk),)
        )
        cls.copy_url = reverse(
            "feature_complete_toy:copy", args=(quote(cls.object.pk),)
        )
        cls.delete_url = reverse(
            "feature_complete_toy:delete", args=(quote(cls.object.pk),)
        )
        cls.inspect_url = reverse(
            "feature_complete_toy:inspect", args=(quote(cls.object.pk),)
        )

    def test_header_buttons_in_edit_view(self):
        response = self.client.get(self.edit_url)
        self.assertEqual(response.status_code, 200)
        soup = self.get_soup(response.content)
        header_buttons = soup.select(".w-slim-header .w-dropdown a")
        expected_buttons = [
            ("Copy", self.copy_url),
            ("Delete", self.delete_url),
            ("Inspect", self.inspect_url),
        ]
        self.assertEqual(
            [(a.text.strip(), a.get("href")) for a in header_buttons],
            expected_buttons,
        )
